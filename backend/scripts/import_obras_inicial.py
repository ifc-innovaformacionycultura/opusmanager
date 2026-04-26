"""
Importación inicial del Excel REGISTRO_DE_REPERTORIO.xlsx en obras + obra_originales.

Busca el archivo en:
  1. /app/REGISTRO_DE_REPERTORIO.xlsx
  2. /mnt/user-data/uploads/REGISTRO_DE_REPERTORIO__respuestas_.xlsx

Idempotente: omite obras que ya existen por (autor, titulo).
Ejecutar:
    cd /app/backend && python scripts/import_obras_inicial.py
"""
import os
import re
import sys
import unicodedata
from pathlib import Path
from datetime import datetime

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
from dotenv import load_dotenv
load_dotenv(ROOT / '.env')

from openpyxl import load_workbook
from supabase import create_client


CANDIDATES = [
    '/app/REGISTRO_DE_REPERTORIO.xlsx',
    '/mnt/user-data/uploads/REGISTRO_DE_REPERTORIO__respuestas_.xlsx',
]

GENERO_VALIDOS = {'SINF.', 'SINF.COR.', 'ESC.', 'COR.'}
PROCEDENCIAS_VALIDAS = {'PROPIO', 'COMPRADO', 'ALQUILER', 'INTERNET', 'INTERNET-LIBRE', 'CESIÓN'}


def strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFKD', s or '') if not unicodedata.combining(c))


def iniciales(autor: str) -> str:
    autor = (autor or '').strip()
    if not autor:
        return 'XX'
    if ',' in autor:
        ap, nom = autor.split(',', 1)
        return ((strip_accents(ap.strip())[:1] or 'X') + (strip_accents(nom.strip())[:1] or 'X')).upper()
    parts = [p for p in re.split(r'\s+', strip_accents(autor)) if p]
    if len(parts) == 1:
        return (parts[0][:2] or 'XX').upper()
    return ((parts[0][:1] + parts[-1][:1]) or 'XX').upper()


def norm_estado(v):
    s = (str(v or '').strip().lower())
    if s in ('si', 'sí', 'yes', 'true', '1', 'x'):
        return 'si'
    if 'revisi' in s or 'revisar' in s:
        return 'necesita_revision'
    return 'no'


def norm_fecha(v):
    if v is None:
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    return s or None


def norm_genero(v):
    s = (str(v or '').strip().upper())
    if not s:
        return None
    return s if s in GENERO_VALIDOS else None


def norm_procedencia(v):
    s = (str(v or '').strip().upper())
    if not s:
        return None
    return s if s in PROCEDENCIAS_VALIDAS else None


def main():
    src = next((p for p in CANDIDATES if os.path.exists(p)), None)
    if not src:
        print(f"❌ No se encontró el Excel en: {CANDIDATES}")
        return 1
    print(f"📄 Leyendo {src}")

    wb = load_workbook(src, data_only=True)
    ws = wb.active
    headers = [str(c.value or '').strip() for c in ws[1]]
    print(f"  columnas: {headers[:5]}…")

    def col(row, *names):
        for n in names:
            if n in headers:
                v = row[headers.index(n)]
                if v is not None and str(v).strip():
                    return v
        return None

    supa = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_KEY'])
    existentes = supa.table('obras').select('autor,titulo').execute().data or []
    ya = {(o.get('autor', '').strip().lower(), o.get('titulo', '').strip().lower()) for o in existentes}

    # Pre-cargar contadores existentes para preservar correlativo de códigos
    todos = supa.table('obras').select('codigo').like('codigo', '%/N%').execute().data or []
    nums_por_inicial = {}
    for r in todos:
        m = re.match(r'^([A-Z]{2})/Nº?(\d+)', r.get('codigo') or '')
        if m:
            ini, n = m.group(1), int(m.group(2))
            nums_por_inicial[ini] = max(nums_por_inicial.get(ini, 0), n)

    insertadas = 0
    duplicadas = 0
    errores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        autor = (str(col(row, 'A.AUTOR', 'autor') or '').strip())
        titulo = (str(col(row, 'B.TITULO DE LA OBRA', 'titulo') or '').strip())
        if not autor or not titulo:
            continue
        key = (autor.lower(), titulo.lower())
        if key in ya:
            duplicadas += 1
            continue

        ini = iniciales(autor)
        nums_por_inicial[ini] = nums_por_inicial.get(ini, 0) + 1
        codigo = f"{ini}/Nº{nums_por_inicial[ini]:03d}"

        payload = {
            "codigo": codigo,
            "autor": autor,
            "titulo": titulo,
            "arreglista": (str(col(row, 'ARREGLISTA') or '').strip() or None),
            "co_autor": (str(col(row, 'CO-AUTOR/LETRISTA', 'co_autor') or '').strip() or None),
            "movimiento": (str(col(row, 'C.MOVIMIENTO/SECCIÓN', 'C.MOVIMIENTO/SECCION', 'movimiento') or '').strip() or None),
            "genero": norm_genero(col(row, 'D.GÉNERO', 'D.GENERO', 'genero')),
            "subgenero": (str(col(row, 'E.SUBGÉNERO', 'E.SUBGENERO', 'subgenero') or '').strip() or None),
            "procedencia": norm_procedencia(col(row, 'PROCEDENCIA DEL MATERIAL', 'procedencia')),
            "fecha_registro": norm_fecha(col(row, 'FECHA DE REGISTRO', 'fecha_registro')),
            "observaciones": (str(col(row, 'OBSERVACIONES', 'observaciones') or '').strip() or None),
        }
        try:
            res = supa.table('obras').insert(payload).execute()
            obra_id = res.data[0]['id'] if res.data else None
            if obra_id:
                originales = {
                    'general': norm_estado(col(row, 'COPIAS ORIGINALES - GENERAL', 'original_general')),
                    'partes':  norm_estado(col(row, 'COPIAS ORIGINALES - PARTES', 'original_partes')),
                    'arcos':   norm_estado(col(row, 'COPIAS ORIGINALES ARCOS', 'COPIAS ORIGINALES - ARCOS', 'original_arcos')),
                }
                for tipo, est in originales.items():
                    supa.table('obra_originales').insert({"obra_id": obra_id, "tipo": tipo, "estado": est}).execute()
                # Enlace digital → guarda en observaciones para no perderlo
                enlace = (str(col(row, 'ENLACE DIGITAL', 'enlace_digital') or '').strip() or None)
                if enlace:
                    base = payload.get('observaciones') or ''
                    nueva_obs = (base + ' | ' if base else '') + f"Drive: {enlace}"
                    supa.table('obras').update({"observaciones": nueva_obs}).eq('id', obra_id).execute()
            insertadas += 1
            ya.add(key)
        except Exception as e:
            errores.append({"autor": autor, "titulo": titulo, "error": str(e)})

    print(f"\n✅ Insertadas: {insertadas}")
    print(f"⏭️  Duplicadas (ya existían): {duplicadas}")
    print(f"❌ Errores: {len(errores)}")
    for e in errores[:10]:
        print(f"   - {e['autor']!r} / {e['titulo']!r}: {e['error']}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
