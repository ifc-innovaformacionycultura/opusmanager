# Gestor Routes - Admin/Manager endpoints
from fastapi import APIRouter, HTTPException, Depends, status, UploadFile, File, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from supabase_client import supabase, create_user_profile
from auth_utils import get_current_user, get_current_gestor, is_super_admin
from typing import List, Optional, Literal, Dict, Any
from datetime import datetime, timezone
from io import BytesIO, StringIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
import logging
import secrets
import string
from email_service import send_musico_credentials_email

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/gestor", tags=["gestor"])

# ==================== Models ====================

class EventoCreate(BaseModel):
    nombre: str
    temporada: Optional[str] = None
    descripcion: Optional[str] = None
    fecha_inicio: Optional[str] = None
    hora_inicio: Optional[str] = None
    fecha_inicio_preparacion: Optional[str] = None
    fecha_fin: Optional[str] = None
    tipo: Optional[str] = None
    lugar: Optional[str] = None
    notas: Optional[str] = None
    # Fechas secundarias de función (punto 2)
    fecha_secundaria_1: Optional[str] = None
    hora_secundaria_1: Optional[str] = None
    fecha_secundaria_2: Optional[str] = None
    hora_secundaria_2: Optional[str] = None
    fecha_secundaria_3: Optional[str] = None
    hora_secundaria_3: Optional[str] = None
    fecha_secundaria_4: Optional[str] = None
    hora_secundaria_4: Optional[str] = None
    # Partituras por sección instrumental (punto 3)
    partitura_cuerda: Optional[str] = None
    partitura_viento_madera: Optional[str] = None
    partitura_viento_metal: Optional[str] = None
    partitura_percusion: Optional[str] = None
    partitura_coro: Optional[str] = None
    partitura_teclados: Optional[str] = None
    # Notas y enlaces para músicos (punto 4)
    notas_musicos: Optional[str] = None
    info_adicional_url_1: Optional[str] = None
    info_adicional_url_2: Optional[str] = None
    info_adicional_url_3: Optional[str] = None

class EventoUpdate(BaseModel):
    nombre: Optional[str] = None
    temporada: Optional[str] = None
    descripcion: Optional[str] = None
    fecha_inicio: Optional[str] = None
    hora_inicio: Optional[str] = None
    fecha_inicio_preparacion: Optional[str] = None
    fecha_fin: Optional[str] = None
    estado: Optional[str] = None
    tipo: Optional[str] = None
    lugar: Optional[str] = None
    notas: Optional[str] = None
    # Fechas secundarias de función (punto 2)
    fecha_secundaria_1: Optional[str] = None
    hora_secundaria_1: Optional[str] = None
    fecha_secundaria_2: Optional[str] = None
    hora_secundaria_2: Optional[str] = None
    fecha_secundaria_3: Optional[str] = None
    hora_secundaria_3: Optional[str] = None
    fecha_secundaria_4: Optional[str] = None
    hora_secundaria_4: Optional[str] = None
    # Partituras por sección instrumental (punto 3)
    partitura_cuerda: Optional[str] = None
    partitura_viento_madera: Optional[str] = None
    partitura_viento_metal: Optional[str] = None
    partitura_percusion: Optional[str] = None
    partitura_coro: Optional[str] = None
    partitura_teclados: Optional[str] = None
    # Notas y enlaces para músicos (punto 4)
    notas_musicos: Optional[str] = None
    info_adicional_url_1: Optional[str] = None
    info_adicional_url_2: Optional[str] = None
    info_adicional_url_3: Optional[str] = None

class AsignacionCreate(BaseModel):
    usuario_id: str
    evento_id: str
    importe: Optional[float] = 0
    comentarios: Optional[str] = None

class EnsayoCreate(BaseModel):
    evento_id: str
    fecha: str  # ISO date string
    hora: str  # HH:MM format
    hora_fin: Optional[str] = None
    tipo: str = "ensayo"  # 'ensayo', 'concierto', 'funcion'
    obligatorio: bool = True
    lugar: Optional[str] = None
    notas: Optional[str] = None

class EnsayoUpdate(BaseModel):
    fecha: Optional[str] = None
    hora: Optional[str] = None
    hora_fin: Optional[str] = None
    tipo: Optional[str] = None
    obligatorio: Optional[bool] = None
    lugar: Optional[str] = None
    notas: Optional[str] = None

class MusicoCreate(BaseModel):
    email: EmailStr
    nombre: str
    apellidos: str
    instrumento: Optional[str] = None
    telefono: Optional[str] = None

# ==================== Eventos ====================

@router.get("/eventos")
async def get_eventos(
    estado: Optional[str] = None,
    temporada: Optional[str] = None,
    current_user: dict = Depends(get_current_gestor)
):
    """Get all eventos (with optional filters). Incluye ensayos[] por evento."""
    try:
        query = supabase.table('eventos').select('*')
        
        if estado:
            query = query.eq('estado', estado)
        if temporada:
            query = query.eq('temporada', temporada)
        
        response = query.order('created_at', desc=True).execute()
        eventos = response.data or []

        # Adjuntar ensayos por evento (una sola query)
        if eventos:
            evento_ids = [e['id'] for e in eventos]
            ens_res = supabase.table('ensayos').select('*') \
                .in_('evento_id', evento_ids) \
                .execute().data or []
            by_evento = {}
            for ens in ens_res:
                by_evento.setdefault(ens['evento_id'], []).append(ens)
            # Orden: ensayos fecha ASC → conciertos/funciones fecha ASC
            for evid, lst in by_evento.items():
                lst.sort(key=lambda x: (0 if (x.get('tipo') or 'ensayo') == 'ensayo' else 1,
                                         x.get('fecha') or '', x.get('hora') or ''))
            for ev in eventos:
                ev['ensayos'] = by_evento.get(ev['id'], [])

        return {"eventos": eventos}
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al cargar eventos: {str(e)}"
        )

@router.post("/eventos")
async def create_evento(
    data: EventoCreate,
    current_user: dict = Depends(get_current_gestor)
):
    """Create new evento"""
    try:
        gestor_id = current_user.get("profile", {}).get("id")
        
        evento_data = {
            **data.model_dump(exclude_none=True),
            "gestor_id": gestor_id,
            "estado": "abierto"
        }
        
        response = supabase.table('eventos').insert(evento_data).execute()
        
        return {
            "message": "Evento creado",
            "evento": response.data[0] if response.data else None
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear evento: {str(e)}"
        )

@router.get("/eventos/{evento_id}")
async def get_evento(
    evento_id: str,
    current_user: dict = Depends(get_current_gestor)
):
    """Get evento by ID with related data"""
    try:
        # Get evento with asignaciones and ensayos
        evento = supabase.table('eventos').select('*').eq('id', evento_id).single().execute()
        
        if not evento.data:
            raise HTTPException(status_code=404, detail="Evento no encontrado")
        
        # Get asignaciones
        asignaciones = supabase.table('asignaciones') \
            .select('*, usuario:usuarios(*)') \
            .eq('evento_id', evento_id) \
            .execute()
        
        # Get ensayos
        ensayos = supabase.table('ensayos') \
            .select('*') \
            .eq('evento_id', evento_id) \
            .order('fecha', desc=False) \
            .execute()
        
        return {
            "evento": evento.data,
            "asignaciones": asignaciones.data or [],
            "ensayos": ensayos.data or []
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al cargar evento: {str(e)}"
        )

@router.put("/eventos/{evento_id}")
async def update_evento(
    evento_id: str,
    data: EventoUpdate,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_gestor)
):
    """Update evento"""
    try:
        # exclude_unset: permite borrar un campo enviándolo explícitamente como null.
        # Normalizamos strings vacíos a None para los campos de fecha/hora/url,
        # ya que PostgreSQL rechazaría "" en TIMESTAMPTZ/TIME.
        raw = data.model_dump(exclude_unset=True)
        null_on_empty = {
            'fecha_inicio', 'fecha_fin',
            'fecha_secundaria_1', 'fecha_secundaria_2', 'fecha_secundaria_3', 'fecha_secundaria_4',
            'hora_secundaria_1', 'hora_secundaria_2', 'hora_secundaria_3', 'hora_secundaria_4',
            'partitura_cuerda', 'partitura_viento_madera', 'partitura_viento_metal',
            'partitura_percusion', 'partitura_coro', 'partitura_teclados',
            'info_adicional_url_1', 'info_adicional_url_2', 'info_adicional_url_3',
        }
        for key in null_on_empty:
            if key in raw and raw[key] == '':
                raw[key] = None

        update_data = {
            **raw,
            "updated_at": datetime.now().isoformat()
        }

        # Regla crítica de verificación (Bloque 2 — refinado):
        # Si el evento pasa de 'abierto' a 'borrador', resetear verificaciones a 'pendiente'
        # para que el director vuelva a revisarlas antes de re-publicar.
        if 'estado' in raw and raw['estado'] == 'borrador':
            try:
                prev = supabase.table('eventos').select('estado').eq('id', evento_id).limit(1).execute().data or []
                if prev and prev[0].get('estado') == 'abierto':
                    supabase.table('evento_verificaciones').delete().eq('evento_id', evento_id).execute()
            except Exception:
                pass

        # Hook certificados: si pasa a 'finalizado', generar certificados en background
        if 'estado' in raw and raw['estado'] == 'finalizado':
            try:
                prev = supabase.table('eventos').select('estado').eq('id', evento_id).limit(1).execute().data or []
                if not prev or prev[0].get('estado') != 'finalizado':
                    from routes_documentos import hook_evento_finalizado
                    background_tasks.add_task(hook_evento_finalizado, evento_id)
            except Exception:
                pass

        response = supabase.table('eventos') \
            .update(update_data) \
            .eq('id', evento_id) \
            .execute()
        
        return {
            "message": "Evento actualizado",
            "evento": response.data[0] if response.data else None
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al actualizar evento: {str(e)}"
        )

# ==================== Seguimiento de Plantillas (Bloque D) ====================

from instrumentos import (
    SECCIONES_ORDER,
    seccion_de_instrumento,
    instrumento_sort_key,
)

def _funciones_del_evento(ev: dict) -> List[dict]:
    """Devuelve la lista de 'funciones' (fecha principal + secundarias) de un evento."""
    funciones = []
    if ev.get('fecha_inicio'):
        funciones.append({"fecha": ev['fecha_inicio'], "hora": None, "label": "Principal"})
    for i in range(1, 5):
        f = ev.get(f'fecha_secundaria_{i}')
        h = ev.get(f'hora_secundaria_{i}')
        if f:
            funciones.append({"fecha": f, "hora": h, "label": f"Función {i + 1}"})
    return funciones


def _nivel_estudios_efectivo(u: dict) -> Optional[str]:
    """Fallback: si no hay nivel_estudios, usa especialidad."""
    return u.get('nivel_estudios') or u.get('especialidad') or None


def _localidad_efectiva(u: dict) -> Optional[str]:
    """Fallback: si no hay localidad, extrae de direccion (última coma)."""
    if u.get('localidad'):
        return u['localidad']
    direccion = (u.get('direccion') or '').strip()
    if ',' in direccion:
        return direccion.rsplit(',', 1)[-1].strip() or None
    return None


@router.get("/seguimiento")
async def get_seguimiento(current_user: dict = Depends(get_current_gestor)):
    """
    Seguimiento de Plantillas — Estructura pivot Músicos × Eventos.

    Respuesta:
      {
        eventos: [{id, nombre, estado ('borrador'|'abierto'), fechas[], ensayos:[{id,tipo,fecha,hora,obligatorio}]}],
        musicos: [{
          id, nombre, apellidos, email, instrumento, especialidad,
          nivel_estudios (efectivo), baremo, localidad (efectiva), anos_experiencia,
          asignaciones: {
            evento_id: {
              asignacion_id, estado, publicado_musico,
              disponibilidad: { ensayo_id: {asiste, asistencia_real} },
              porcentaje_disponibilidad, porcentaje_asistencia_real
            }
          }
        }]
      }
    """
    try:
        # 1) Eventos publicados (estado='abierto')
        eventos_res = supabase.table('eventos') \
            .select('*') \
            .eq('estado', 'abierto') \
            .order('fecha_inicio', desc=False) \
            .execute()
        eventos_raw = eventos_res.data or []
        evento_ids = [e['id'] for e in eventos_raw]

        # 2) Ensayos de esos eventos. ORDEN: primero tipo='ensayo' por fecha ASC,
        #    luego los demás (concierto/funcion) por fecha ASC. Dentro de cada grupo, por fecha+hora.
        ensayos_map = {}  # evento_id -> [ensayos ordenados]
        if evento_ids:
            ens_res = supabase.table('ensayos') \
                .select('id,evento_id,tipo,fecha,hora,obligatorio,lugar') \
                .in_('evento_id', evento_ids) \
                .execute()
            tmp_by_evento = {}
            for e in (ens_res.data or []):
                tmp_by_evento.setdefault(e['evento_id'], []).append(e)
            for evid, lst in tmp_by_evento.items():
                # sort key: (0 si tipo='ensayo' else 1, fecha, hora)
                lst.sort(key=lambda x: (
                    0 if (x.get('tipo') or 'ensayo') == 'ensayo' else 1,
                    x.get('fecha') or '',
                    x.get('hora') or ''
                ))
                ensayos_map[evid] = lst

        eventos_out = []
        for ev in eventos_raw:
            eventos_out.append({
                "id": ev['id'],
                "nombre": ev.get('nombre'),
                "estado": ev.get('estado'),
                "tipo": ev.get('tipo'),
                "lugar": ev.get('lugar'),
                "temporada": ev.get('temporada'),
                "fecha_inicio": ev.get('fecha_inicio'),
                "fecha_fin": ev.get('fecha_fin'),
                "fechas": _funciones_del_evento(ev),
                "ensayos": ensayos_map.get(ev['id'], []),
            })

        # 3) Músicos activos
        musicos_res = supabase.table('usuarios') \
            .select('id,nombre,apellidos,email,instrumento,especialidad,nivel_estudios,baremo,localidad,direccion,anos_experiencia,estado,estado_invitacion') \
            .eq('rol', 'musico') \
            .eq('estado', 'activo') \
            .order('apellidos', desc=False) \
            .execute()
        musicos_raw = musicos_res.data or []

        # 4) Asignaciones
        asigs_list = []
        if evento_ids:
            a_res = supabase.table('asignaciones') \
                .select('id,usuario_id,evento_id,estado,publicado_musico,cache_presupuestado,importe') \
                .in_('evento_id', evento_ids) \
                .execute()
            asigs_list = a_res.data or []

        # 5) Disponibilidades de esos usuarios (solo para los ensayos conocidos)
        ensayo_ids = [e['id'] for evlist in ensayos_map.values() for e in evlist]
        disp_map = {}  # (usuario_id, ensayo_id) -> disponibilidad
        if ensayo_ids:
            d_res = supabase.table('disponibilidad') \
                .select('id,usuario_id,ensayo_id,asiste,asistencia_real') \
                .in_('ensayo_id', ensayo_ids) \
                .execute()
            for d in (d_res.data or []):
                disp_map[(d['usuario_id'], d['ensayo_id'])] = d

        # Convocatoria por instrumento por ensayo
        ensayo_instr_map = {}
        if ensayo_ids:
            ei_res = supabase.table('ensayo_instrumentos') \
                .select('ensayo_id,instrumento,convocado') \
                .in_('ensayo_id', ensayo_ids).execute()
            for row in (ei_res.data or []):
                ensayo_instr_map.setdefault(row['ensayo_id'], {})[row['instrumento']] = bool(row.get('convocado', True))

        # Index asignaciones por (usuario_id, evento_id)
        asig_index = {(a['usuario_id'], a['evento_id']): a for a in asigs_list}

        # CRM contactos (Bloque 1) — resumen por (usuario_id, evento_id)
        crm_index = {}  # (usuario_id, evento_id) -> {total, ultimo_tipo, ultimo_estado, ultima_fecha}
        if evento_ids:
            try:
                cr = supabase.table('contactos_musico') \
                    .select('usuario_id,evento_id,tipo,estado_respuesta,fecha_contacto') \
                    .in_('evento_id', evento_ids) \
                    .order('fecha_contacto', desc=True) \
                    .execute().data or []
                for row in cr:
                    key = (row['usuario_id'], row['evento_id'])
                    if key not in crm_index:
                        crm_index[key] = {
                            "total_contactos": 0,
                            "ultimo_tipo": row.get('tipo'),
                            "ultimo_estado": row.get('estado_respuesta'),
                            "ultima_fecha": row.get('fecha_contacto'),
                        }
                    crm_index[key]["total_contactos"] += 1
            except Exception:
                pass

        # Total ensayos por evento
        total_ensayos_por_evento = {eid: len(evs) for eid, evs in ensayos_map.items()}

        musicos_out = []
        for u in musicos_raw:
            m = {
                "id": u['id'],
                "nombre": u.get('nombre') or '',
                "apellidos": u.get('apellidos') or '',
                "email": u.get('email') or '',
                "instrumento": u.get('instrumento'),
                "especialidad": u.get('especialidad'),
                "nivel_estudios": _nivel_estudios_efectivo(u),
                "baremo": u.get('baremo'),
                "localidad": _localidad_efectiva(u),
                "anos_experiencia": u.get('anos_experiencia'),
                "estado_invitacion": u.get('estado_invitacion') or 'pendiente',
                # Shape unificado con /plantillas-definitivas: lista ordenada por evento.
                # Cada item incluye `evento_id` para lookup directo en el frontend.
                "asignaciones": []
            }
            for ev in eventos_raw:
                asig = asig_index.get((u['id'], ev['id']))
                ensayos = ensayos_map.get(ev['id'], [])
                instr_musico = u.get('instrumento')
                # Shape unificado con /plantillas-definitivas: LISTA ordenada por ensayo
                disp_list = []
                si_disp = 0
                si_real = 0
                convocados_count = 0
                for e in ensayos:
                    convocado = _is_convocado(ensayo_instr_map, e['id'], instr_musico)
                    if convocado:
                        convocados_count += 1
                    d = disp_map.get((u['id'], e['id']))
                    disp_list.append({
                        "ensayo_id": e['id'],
                        "asiste": d.get('asiste') if d else None,
                        "asistencia_real": d.get('asistencia_real') if d else None,
                        "disponibilidad_id": d.get('id') if d else None,
                        "convocado": convocado,
                    })
                    if convocado and d:
                        if d.get('asiste') is True:
                            si_disp += 1
                        if d.get('asistencia_real') is True:
                            si_real += 1
                pct_disp = round((si_disp / convocados_count) * 100) if convocados_count else 0
                pct_real = round((si_real / convocados_count) * 100) if convocados_count else 0

                m["asignaciones"].append({
                    "evento_id": ev['id'],
                    "asignacion_id": asig['id'] if asig else None,
                    "estado": asig['estado'] if asig else None,
                    "publicado_musico": bool(asig.get('publicado_musico')) if asig else False,
                    "disponibilidad": disp_list,
                    "porcentaje_disponibilidad": pct_disp,
                    "porcentaje_asistencia_real": pct_real,
                    "crm": crm_index.get((u['id'], ev['id'])) or {
                        "total_contactos": 0,
                        "ultimo_tipo": None,
                        "ultimo_estado": "no_contactado",
                        "ultima_fecha": None,
                    },
                })
            musicos_out.append(m)

        return {"eventos": eventos_out, "musicos": musicos_out}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en seguimiento: {str(e)}")


class PublicarRequest(BaseModel):
    usuario_ids: List[str]
    evento_id: str
    publicar: bool


@router.post("/seguimiento/publicar")
async def seguimiento_publicar(
    data: PublicarRequest,
    current_user: dict = Depends(get_current_gestor)
):
    """
    Crea o actualiza asignaciones para publicar/despublicar un evento a un conjunto de músicos.
    - publicar=True:  crea asignación con estado='pendiente', publicado_musico=True, fecha_publicacion=NOW.
                     Si ya existe, sólo actualiza publicado_musico=True + fecha_publicacion.
    - publicar=False: si existe, publicado_musico=False. Si no existe, no hace nada.
    """
    if not data.usuario_ids:
        return {"publicados": 0, "despublicados": 0, "creados": 0}

    try:
        existing_res = supabase.table('asignaciones') \
            .select('id,usuario_id,publicado_musico') \
            .eq('evento_id', data.evento_id) \
            .in_('usuario_id', data.usuario_ids) \
            .execute()
        existing_by_user = {a['usuario_id']: a for a in (existing_res.data or [])}

        publicados = 0
        despublicados = 0
        creados = 0
        now = datetime.now().isoformat()
        # Push (Bloque PWA): obtener nombre del evento sólo una vez
        evento_nombre = None
        if data.publicar:
            try:
                ev_res = supabase.table('eventos').select('nombre').eq('id', data.evento_id).limit(1).execute().data or []
                evento_nombre = ev_res[0].get('nombre') if ev_res else None
            except Exception:
                pass

        for uid in data.usuario_ids:
            existing = existing_by_user.get(uid)
            if data.publicar:
                if existing:
                    supabase.table('asignaciones').update({
                        "publicado_musico": True,
                        "fecha_publicacion": now,
                        "updated_at": now,
                    }).eq('id', existing['id']).execute()
                    publicados += 1
                else:
                    supabase.table('asignaciones').insert({
                        "usuario_id": uid,
                        "evento_id": data.evento_id,
                        "estado": "pendiente",
                        "publicado_musico": True,
                        "fecha_publicacion": now,
                    }).execute()
                    creados += 1
                # Push al músico
                try:
                    from routes_push import notify_push
                    notify_push(
                        uid,
                        f"🎼 Nueva convocatoria: {evento_nombre or 'Nuevo evento'}",
                        "Confirma tu disponibilidad en el portal.",
                        '/portal',
                        tipo='convocatoria',
                    )
                except Exception:
                    pass
            else:
                if existing:
                    supabase.table('asignaciones').update({
                        "publicado_musico": False,
                        "updated_at": now,
                    }).eq('id', existing['id']).execute()
                    despublicados += 1

        return {"publicados": publicados, "despublicados": despublicados, "creados": creados}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al publicar: {str(e)}")


class BulkAccionRequest(BaseModel):
    usuario_ids: List[str]
    evento_id: str
    accion: str  # 'pendiente' | 'confirmado' | 'no_disponible' | 'excluido' | 'rechazado'


@router.post("/seguimiento/bulk-accion")
async def seguimiento_bulk_accion(
    data: BulkAccionRequest,
    current_user: dict = Depends(get_current_gestor)
):
    """Aplica un cambio de estado a varios músicos para un evento (UPSERT)."""
    valid = {'pendiente', 'confirmado', 'rechazado', 'no_disponible', 'excluido'}
    if data.accion not in valid:
        raise HTTPException(status_code=400, detail=f"Acción inválida. Usa: {sorted(valid)}")
    if not data.usuario_ids:
        return {"actualizados": 0, "creados": 0}

    try:
        existing_res = supabase.table('asignaciones') \
            .select('id,usuario_id') \
            .eq('evento_id', data.evento_id) \
            .in_('usuario_id', data.usuario_ids) \
            .execute()
        existing_by_user = {a['usuario_id']: a['id'] for a in (existing_res.data or [])}

        actualizados = 0
        creados = 0
        now = datetime.now().isoformat()
        for uid in data.usuario_ids:
            if uid in existing_by_user:
                supabase.table('asignaciones') \
                    .update({"estado": data.accion, "updated_at": now}) \
                    .eq('id', existing_by_user[uid]) \
                    .execute()
                actualizados += 1
            else:
                supabase.table('asignaciones').insert({
                    "usuario_id": uid,
                    "evento_id": data.evento_id,
                    "estado": data.accion,
                }).execute()
                creados += 1

        return {"actualizados": actualizados, "creados": creados}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en acción masiva: {str(e)}")


# ==================== Plantillas Definitivas (Bloque D) ====================

def _cachet_lookup(cachets_rows: List[dict], instrumento: Optional[str], nivel: Optional[str]) -> Optional[float]:
    """Busca en cachets_config el importe para (instrumento, nivel). Case-insensitive."""
    res = _cachet_lookup_with_source(cachets_rows, instrumento, nivel)
    return res[0] if res else None


def _cachet_lookup_with_source(cachets_rows: List[dict], instrumento: Optional[str], nivel: Optional[str]):
    """Igual que _cachet_lookup pero devuelve (importe, fuente) donde fuente indica
    si fue match exacto, por_instrumento, base_exacto, base_por_instrumento.
    La lista cachets_rows puede mezclar filas con evento_id (específicas) y evento_id=NULL (base).
    Prioridad: evento+instrumento+nivel → evento+instrumento → base+instrumento+nivel → base+instrumento."""
    if not instrumento:
        return None
    i = str(instrumento).strip().lower()
    n = str(nivel or '').strip().lower()

    # Separar event-specific vs base
    specific = [r for r in cachets_rows if r.get('evento_id')]
    base = [r for r in cachets_rows if not r.get('evento_id')]

    def match(rows, want_nivel):
        for r in rows:
            if (r.get('instrumento') or '').strip().lower() != i:
                continue
            nivel_r = (r.get('nivel_estudios') or '').strip().lower()
            if want_nivel and nivel_r != n:
                continue
            if not want_nivel and nivel_r and n and nivel_r != n:
                continue
            if r.get('importe') is None:
                continue
            return float(r['importe'])
        return None

    # 1) evento+instrumento+nivel
    v = match(specific, True)
    if v is not None:
        return (v, 'exacto')
    # 2) evento+instrumento (cualquier nivel)
    for r in specific:
        if (r.get('instrumento') or '').strip().lower() == i and r.get('importe') is not None:
            return (float(r['importe']), 'por_instrumento')
    # 3) base+instrumento+nivel
    v = match(base, True)
    if v is not None:
        return (v, 'base_exacto')
    # 4) base+instrumento
    for r in base:
        if (r.get('instrumento') or '').strip().lower() == i and r.get('importe') is not None:
            return (float(r['importe']), 'base_por_instrumento')
    return None


@router.get("/plantillas-definitivas")
async def get_plantillas_definitivas(current_user: dict = Depends(get_current_gestor)):
    """
    Devuelve eventos que tienen al menos un músico confirmado,
    agrupados por sección instrumental, con todos los cálculos
    necesarios para la pantalla (disponibilidad, asistencia real,
    cachet previsto/real, gastos adicionales).
    """
    try:
        # Asignaciones confirmadas
        # Iter E1 — añadidos campos de cierre de plantilla.
        a_res = supabase.table('asignaciones') \
            .select('id,usuario_id,evento_id,estado,cache_presupuestado,importe,numero_atril,letra,comentarios,nivel_estudios,porcentaje_asistencia,estado_cierre,cerrado_plantilla_por,cerrado_plantilla_at') \
            .eq('estado', 'confirmado') \
            .execute()
        confirmadas = a_res.data or []
        if not confirmadas:
            return {"eventos": []}

        evento_ids = list({a['evento_id'] for a in confirmadas})
        usuario_ids = list({a['usuario_id'] for a in confirmadas})

        # Iter E1 — Eventos visibles:
        # 1) estado='abierto' (comportamiento original, plantilla editable).
        # 2) Cualquier evento cuyas asignaciones tengan estado_cierre IN
        #    ('cerrado_plantilla','cerrado_economico') — plantilla en modo solo lectura.
        eventos_cerrados_ids = {
            a['evento_id'] for a in confirmadas
            if (a.get('estado_cierre') or 'abierto') in ('cerrado_plantilla', 'cerrado_economico')
        }
        eventos_res = supabase.table('eventos') \
            .select('*') \
            .in_('id', evento_ids) \
            .order('fecha_inicio', desc=False) \
            .execute()
        eventos_raw = [
            e for e in (eventos_res.data or [])
            if e.get('estado') == 'abierto' or e['id'] in eventos_cerrados_ids
        ]
        # Recalcular evento_ids después del filtro (puede haber quedado vacío)
        evento_ids = [e['id'] for e in eventos_raw]
        if not evento_ids:
            return {"eventos": [], "total_temporada": 0}

        # Iter E1 — Mapa de cierre por evento. Si cualquier asignacion del evento
        # tiene estado_cierre != 'abierto', consideramos el evento como cerrado.
        cierre_by_evento: Dict[str, Dict] = {}
        for a in confirmadas:
            evid = a['evento_id']
            if evid not in evento_ids:
                continue
            ec = a.get('estado_cierre') or 'abierto'
            cur = cierre_by_evento.get(evid)
            if cur is None or (cur.get('estado_cierre') == 'abierto' and ec != 'abierto'):
                cierre_by_evento[evid] = {
                    "estado_cierre": ec,
                    "cerrado_plantilla_por": a.get('cerrado_plantilla_por'),
                    "cerrado_plantilla_at": a.get('cerrado_plantilla_at'),
                }
        # Resolver nombre del gestor que cerró cada evento.
        gestor_ids_cierre = list({
            v.get('cerrado_plantilla_por') for v in cierre_by_evento.values()
            if v.get('cerrado_plantilla_por')
        })
        gestores_cierre_by_id: Dict[str, str] = {}
        if gestor_ids_cierre:
            try:
                gres = supabase.table('usuarios').select('id,nombre,apellidos') \
                    .in_('id', gestor_ids_cierre).execute().data or []
                gestores_cierre_by_id = {
                    g['id']: f"{g.get('nombre','')} {g.get('apellidos','')}".strip() or g.get('id')
                    for g in gres
                }
            except Exception:
                pass

        # Iter E1.1 — Eventos que tienen al menos una entrada en registro_actividad
        # Iter E2 — ampliado a tipos económicos para mostrar el botón "🕒 Historial"
        # también cuando solo hay entradas de cierre/reapertura económica.
        eventos_con_historial: set = set()
        try:
            ra = supabase.table('registro_actividad').select('entidad_id,tipo') \
                .eq('entidad_tipo', 'evento') \
                .in_('tipo', ['evento_concluido', 'evento_reabierto', 'economico_cerrado', 'economico_reabierto']) \
                .in_('entidad_id', evento_ids).execute().data or []
            for r in ra:
                if r.get('entidad_id'):
                    eventos_con_historial.add(r['entidad_id'])
        except Exception:
            pass

        ens_res = supabase.table('ensayos') \
            .select('id,evento_id,tipo,fecha,hora,obligatorio,lugar') \
            .in_('evento_id', evento_ids) \
            .order('fecha', desc=False) \
            .execute()
        ensayos_by_evento = {}
        for e in (ens_res.data or []):
            ensayos_by_evento.setdefault(e['evento_id'], []).append(e)

        usuarios_res = supabase.table('usuarios') \
            .select('id,nombre,apellidos,email,instrumento,especialidad,nivel_estudios,baremo,localidad,direccion,anos_experiencia') \
            .in_('id', usuario_ids) \
            .execute()
        usuarios_by_id = {u['id']: u for u in (usuarios_res.data or [])}

        ensayo_ids = [e['id'] for evs in ensayos_by_evento.values() for e in evs]
        disp_by_pair = {}  # (usuario_id, ensayo_id) -> row
        if ensayo_ids:
            d_res = supabase.table('disponibilidad') \
                .select('id,usuario_id,ensayo_id,asiste,asistencia_real') \
                .in_('ensayo_id', ensayo_ids) \
                .in_('usuario_id', usuario_ids) \
                .execute()
            for d in (d_res.data or []):
                disp_by_pair[(d['usuario_id'], d['ensayo_id'])] = d

        gastos_res = supabase.table('gastos_adicionales') \
            .select('id,usuario_id,evento_id,cache_extra,transporte_importe,transporte_justificante_url,alojamiento_importe,alojamiento_justificante_url,otros_importe,otros_justificante_url,notas,cache_extra_provisional,cache_extra_validado_por,cache_extra_validado_at,transporte_provisional,transporte_validado_por,transporte_validado_at') \
            .in_('evento_id', evento_ids) \
            .in_('usuario_id', usuario_ids) \
            .execute()
        gastos_by_pair = {(g['usuario_id'], g['evento_id']): g for g in (gastos_res.data or [])}

        # Iter F1 — Resolver nombres de validadores (cache_extra + transporte).
        validador_ids = set()
        for g in (gastos_res.data or []):
            if g.get('cache_extra_validado_por'): validador_ids.add(g['cache_extra_validado_por'])
            if g.get('transporte_validado_por'): validador_ids.add(g['transporte_validado_por'])
        validadores_by_id: Dict[str, str] = {}
        if validador_ids:
            try:
                vrows = supabase.table('usuarios').select('id,nombre,apellidos') \
                    .in_('id', list(validador_ids)).execute().data or []
                validadores_by_id = {
                    v['id']: f"{v.get('nombre','')} {v.get('apellidos','')}".strip() or v.get('id')
                    for v in vrows
                }
            except Exception:
                pass

        # Iter D · Comedor descontable: importe a descontar del TOTAL por (usuario, evento)
        # Suma de las comidas confirmadas (precio_menu + precio_cafe si toma_cafe) por músico.
        comida_by_pair: Dict[tuple, float] = {}
        try:
            com_rows = supabase.table('evento_comidas') \
                .select('id,evento_id,incluye_cafe,precio_menu,precio_cafe') \
                .in_('evento_id', evento_ids).execute().data or []
            comida_ids = [c['id'] for c in com_rows]
            if comida_ids:
                comida_meta = {c['id']: c for c in com_rows}
                conf_rows = supabase.table('confirmaciones_comida') \
                    .select('comida_id,usuario_id,confirmado,toma_cafe') \
                    .in_('comida_id', comida_ids) \
                    .in_('usuario_id', usuario_ids).execute().data or []
                for cr in conf_rows:
                    if cr.get('confirmado') is not True:
                        continue
                    meta = comida_meta.get(cr.get('comida_id'))
                    if not meta:
                        continue
                    importe = float(meta.get('precio_menu') or 0)
                    if meta.get('incluye_cafe') and cr.get('toma_cafe'):
                        importe += float(meta.get('precio_cafe') or 0)
                    key = (cr['usuario_id'], meta['evento_id'])
                    comida_by_pair[key] = comida_by_pair.get(key, 0.0) + importe
        except Exception as e:
            print(f"[WARN] plantillas-definitivas: no se pudo calcular comida_importe: {e}")

        # Convocatoria por instrumento por ensayo
        ensayo_instr_map = {}
        if ensayo_ids:
            ei_res = supabase.table('ensayo_instrumentos') \
                .select('ensayo_id,instrumento,convocado') \
                .in_('ensayo_id', ensayo_ids).execute()
            for row in (ei_res.data or []):
                ensayo_instr_map.setdefault(row['ensayo_id'], {})[row['instrumento']] = bool(row.get('convocado', True))

        cachets_res = supabase.table('cachets_config') \
            .select('id,evento_id,instrumento,nivel_estudios,importe') \
            .in_('evento_id', evento_ids) \
            .execute()
        cachets_by_evento = {}
        for c in (cachets_res.data or []):
            cachets_by_evento.setdefault(c['evento_id'], []).append(c)
        # Cachets base (evento_id IS NULL) — plantilla global
        cachets_base_res = supabase.table('cachets_config') \
            .select('id,evento_id,instrumento,nivel_estudios,importe') \
            .is_('evento_id', 'null').execute()
        cachets_base = cachets_base_res.data or []

        eventos_out = []
        for ev in eventos_raw:
            ensayos = ensayos_by_evento.get(ev['id'], [])
            total_e = len(ensayos)
            asigs_ev = [a for a in confirmadas if a['evento_id'] == ev['id']]

            # Construir músicos agrupados por sección
            secciones_map = {key: [] for key, _label in SECCIONES_ORDER}
            sin_seccion = []
            total_ev = {
                "cache_previsto": 0.0, "cache_real": 0.0, "extras": 0.0,
                "transporte": 0.0, "alojamiento": 0.0, "otros": 0.0, "comida": 0.0, "total": 0.0,
            }
            for a in asigs_ev:
                u = usuarios_by_id.get(a['usuario_id'])
                if not u:
                    continue
                instr_musico = u.get('instrumento')
                disp_list = []
                asist_list = []
                for e in ensayos:
                    convocado = _is_convocado(ensayo_instr_map, e['id'], instr_musico)
                    d = disp_by_pair.get((u['id'], e['id']))
                    disp_list.append({
                        "ensayo_id": e['id'],
                        "asiste": d.get('asiste') if d else None,
                        "disponibilidad_id": d.get('id') if d else None,
                        "convocado": convocado,
                    })
                    asist_list.append({
                        "ensayo_id": e['id'],
                        "asistencia_real": d.get('asistencia_real') if d else None,
                        "convocado": convocado,
                    })
                # % disponibilidad calculado SÓLO sobre ensayos convocados
                ensayos_convocados = [x for x in disp_list if x["convocado"]]
                total_convocados = len(ensayos_convocados)
                si_disp = sum(1 for x in ensayos_convocados if x["asiste"] is True)
                pct_disp = round((si_disp / total_convocados) * 100) if total_convocados else 0

                # pct_real = promedio de asistencia_real (no NULL) SÓLO en ensayos convocados.
                valores_real = [
                    float(x["asistencia_real"]) for x in asist_list
                    if x["convocado"] and x["asistencia_real"] is not None
                ]
                pct_real = round(sum(valores_real) / len(valores_real), 2) if valores_real else 0

                # Caché previsto: cachets_config específico → base → fallback asignaciones
                nivel_efectivo = a.get('nivel_estudios') or _nivel_estudios_efectivo(u)
                combined_cachets = cachets_by_evento.get(ev['id'], []) + cachets_base
                lookup = _cachet_lookup_with_source(combined_cachets, u.get('instrumento'), nivel_efectivo)
                if lookup is not None:
                    cache_prev, cache_fuente = lookup
                else:
                    cache_prev = float(a.get('cache_presupuestado') or a.get('importe') or 0)
                    cache_fuente = 'asignacion' if cache_prev > 0 else 'sin_datos'

                cache_real = round(cache_prev * (pct_real / 100.0), 2)

                g = gastos_by_pair.get((u['id'], ev['id'])) or {}
                extras_visibles = float(g.get('cache_extra') or 0)
                transp_visibles = float(g.get('transporte_importe') or 0)
                # Iter F1 — provisionales NO suman al TOTAL.
                cache_extra_prov = bool(g.get('cache_extra_provisional'))
                transporte_prov = bool(g.get('transporte_provisional'))
                extras = 0.0 if cache_extra_prov else extras_visibles
                transp = 0.0 if transporte_prov else transp_visibles
                aloj = float(g.get('alojamiento_importe') or 0)
                otros = float(g.get('otros_importe') or 0)
                comida = float(comida_by_pair.get((u['id'], ev['id'])) or 0)
                total = round(cache_real + extras + transp + aloj + otros - comida, 2)

                total_ev["cache_previsto"] += cache_prev
                total_ev["cache_real"]     += cache_real
                total_ev["extras"]         += extras
                total_ev["transporte"]     += transp
                total_ev["alojamiento"]    += aloj
                total_ev["otros"]          += otros
                total_ev["comida"]         += comida
                total_ev["total"]          += total

                musico_row = {
                    "asignacion_id": a['id'],
                    "usuario_id": u['id'],
                    "nombre": u.get('nombre') or '',
                    "apellidos": u.get('apellidos') or '',
                    "email": u.get('email') or '',
                    "instrumento": u.get('instrumento'),
                    "especialidad": u.get('especialidad'),
                    "nivel_estudios": nivel_efectivo,
                    "numero_atril": a.get('numero_atril'),
                    "letra": a.get('letra'),
                    "comentario": a.get('comentarios') or '',
                    "disponibilidad": disp_list,
                    "asistencia": asist_list,
                    "porcentaje_disponibilidad": pct_disp,
                    "porcentaje_asistencia_real": pct_real,
                    "cache_previsto": round(cache_prev, 2),
                    "cache_fuente": cache_fuente,
                    "cache_real": cache_real,
                    "cache_extra": extras_visibles,
                    "motivo_extra": g.get('notas') or '',
                    "transporte_importe": transp_visibles,
                    "transporte_justificante_url": g.get('transporte_justificante_url'),
                    "alojamiento_importe": aloj,
                    "alojamiento_justificante_url": g.get('alojamiento_justificante_url'),
                    "otros_importe": otros,
                    "otros_justificante_url": g.get('otros_justificante_url'),
                    "comida_importe": comida,
                    "total": total,
                    # Iter F1 — Provisional / Validación
                    "gasto_id": g.get('id'),
                    "cache_extra_provisional": cache_extra_prov,
                    "cache_extra_validado_por_nombre": validadores_by_id.get(g.get('cache_extra_validado_por')) if g.get('cache_extra_validado_por') else None,
                    "cache_extra_validado_at": g.get('cache_extra_validado_at'),
                    "transporte_provisional": transporte_prov,
                    "transporte_validado_por_nombre": validadores_by_id.get(g.get('transporte_validado_por')) if g.get('transporte_validado_por') else None,
                    "transporte_validado_at": g.get('transporte_validado_at'),
                    # Iter F1 — auxiliares (uso interno para subtotales coherentes con TOTAL).
                    "_extras_efectivo": extras,
                    "_transp_efectivo": transp,
                }
                sec_key = seccion_de_instrumento(u.get('instrumento'))
                if sec_key:
                    secciones_map[sec_key].append(musico_row)
                else:
                    sin_seccion.append(musico_row)

            # Ordenar dentro de cada sección por instrumento → apellidos
            for key in secciones_map:
                secciones_map[key].sort(key=lambda m: (instrumento_sort_key(m.get('instrumento')),
                                                       (m.get('apellidos') or '').lower()))
            sin_seccion.sort(key=lambda m: (m.get('apellidos') or '').lower())

            secciones_out = []
            for key, label in SECCIONES_ORDER:
                musicos_sec = secciones_map[key]
                if not musicos_sec:
                    continue
                sec_totals = {k: 0.0 for k in ("cache_previsto","cache_real","extras","transporte","alojamiento","otros","total")}
                for m in musicos_sec:
                    sec_totals["cache_previsto"] += m["cache_previsto"]
                    sec_totals["cache_real"]     += m["cache_real"]
                    sec_totals["extras"]         += m["_extras_efectivo"]
                    sec_totals["transporte"]     += m["_transp_efectivo"]
                    sec_totals["alojamiento"]    += m["alojamiento_importe"]
                    sec_totals["otros"]          += m["otros_importe"]
                    sec_totals["total"]          += m["total"]
                secciones_out.append({
                    "key": key, "label": label,
                    "count": len(musicos_sec),
                    "musicos": musicos_sec,
                    "totales": {k: round(v, 2) for k, v in sec_totals.items()},
                })
            if sin_seccion:
                sec_totals = {k: 0.0 for k in ("cache_previsto","cache_real","extras","transporte","alojamiento","otros","total")}
                for m in sin_seccion:
                    sec_totals["cache_previsto"] += m["cache_previsto"]
                    sec_totals["cache_real"]     += m["cache_real"]
                    sec_totals["extras"]         += m["_extras_efectivo"]
                    sec_totals["transporte"]     += m["_transp_efectivo"]
                    sec_totals["alojamiento"]    += m["alojamiento_importe"]
                    sec_totals["otros"]          += m["otros_importe"]
                    sec_totals["total"]          += m["total"]
                secciones_out.append({
                    "key": "otros", "label": "Sin sección",
                    "count": len(sin_seccion),
                    "musicos": sin_seccion,
                    "totales": {k: round(v, 2) for k, v in sec_totals.items()},
                })

            # Iter F1 — Limpiar campos auxiliares antes de serializar.
            for sec in secciones_out:
                for m in sec.get('musicos', []):
                    m.pop('_extras_efectivo', None)
                    m.pop('_transp_efectivo', None)

            cierre_info = cierre_by_evento.get(ev['id']) or {}
            eventos_out.append({
                "id": ev['id'],
                "nombre": ev.get('nombre'),
                "estado": ev.get('estado'),
                "fecha_inicio": ev.get('fecha_inicio'),
                "fechas": _funciones_del_evento(ev),
                "lugar": ev.get('lugar'),
                "ensayos": ensayos,
                "total_musicos": len(asigs_ev),
                "totales": {k: round(v, 2) for k, v in total_ev.items()},
                "secciones": secciones_out,
                # Iter E1 — info de cierre de plantilla
                "estado_cierre": cierre_info.get("estado_cierre") or "abierto",
                "cerrado_plantilla_at": cierre_info.get("cerrado_plantilla_at"),
                "cerrado_plantilla_por_nombre": gestores_cierre_by_id.get(
                    cierre_info.get("cerrado_plantilla_por")
                ) if cierre_info.get("cerrado_plantilla_por") else None,
                # Iter E1.1 — flag de historial (botón visible si True)
                "tiene_historial_cierre": ev['id'] in eventos_con_historial,
            })

        return {"eventos": eventos_out}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en plantillas definitivas: {str(e)}")


class AsistenciaItem(BaseModel):
    disponibilidad_id: Optional[str] = None
    usuario_id: Optional[str] = None
    ensayo_id: Optional[str] = None
    asistencia_real: Optional[float] = None  # porcentaje 0..100


class GastoItem(BaseModel):
    usuario_id: str
    evento_id: str
    transporte_importe: Optional[float] = None
    transporte_justificante_url: Optional[str] = None
    alojamiento_importe: Optional[float] = None
    alojamiento_justificante_url: Optional[str] = None
    otros_importe: Optional[float] = None
    otros_justificante_url: Optional[str] = None
    cache_extra: Optional[float] = None
    notas: Optional[str] = None


class AnotacionItem(BaseModel):
    asignacion_id: str
    numero_atril: Optional[int] = None
    letra: Optional[str] = None
    comentario: Optional[str] = None


class GuardarPlantillasRequest(BaseModel):
    asistencias: List[AsistenciaItem] = []
    gastos: List[GastoItem] = []
    anotaciones: List[AnotacionItem] = []


@router.put("/plantillas-definitivas/guardar")
async def guardar_plantillas_definitivas(
    data: GuardarPlantillasRequest,
    current_user: dict = Depends(get_current_gestor)
):
    """
    Persiste cambios de Plantillas Definitivas en una sola llamada:
      - asistencias: UPDATE/UPSERT disponibilidad.asistencia_real
      - gastos: UPSERT gastos_adicionales (usuario_id + evento_id UNIQUE)
      - anotaciones: UPDATE asignaciones.numero_atril / letra / comentarios
      - Actualiza asignaciones.porcentaje_asistencia por cada asignación tocada
    """
    now = datetime.now().isoformat()
    resumen = {"asistencias": 0, "gastos": 0, "anotaciones": 0}
    try:
        # ============================================================
        # Iter E1 — Defensa backend: bloquear cambios sobre eventos
        # cuya plantilla esté concluida (estado_cierre != 'abierto').
        # ============================================================
        evento_ids_tocados = set()
        for g in data.gastos:
            if g.evento_id:
                evento_ids_tocados.add(g.evento_id)
        ensayo_ids_payload = list({a.ensayo_id for a in data.asistencias if a.ensayo_id})
        if ensayo_ids_payload:
            er = supabase.table('ensayos').select('id,evento_id') \
                .in_('id', ensayo_ids_payload).execute().data or []
            for e in er:
                if e.get('evento_id'):
                    evento_ids_tocados.add(e['evento_id'])
        asig_ids_payload = [n.asignacion_id for n in data.anotaciones if n.asignacion_id]
        if asig_ids_payload:
            ar = supabase.table('asignaciones').select('id,evento_id') \
                .in_('id', asig_ids_payload).execute().data or []
            for a in ar:
                if a.get('evento_id'):
                    evento_ids_tocados.add(a['evento_id'])
        if evento_ids_tocados:
            cerr = supabase.table('asignaciones').select('evento_id,estado_cierre') \
                .in_('evento_id', list(evento_ids_tocados)).execute().data or []
            evs_cerrados = {
                c['evento_id'] for c in cerr
                if (c.get('estado_cierre') or 'abierto') != 'abierto'
            }
            if evs_cerrados:
                # Cargar nombres para mensaje claro
                names = {}
                try:
                    nr = supabase.table('eventos').select('id,nombre') \
                        .in_('id', list(evs_cerrados)).execute().data or []
                    names = {n['id']: n.get('nombre') for n in nr}
                except Exception:
                    pass
                etiquetas = ", ".join(
                    f"'{names.get(eid, eid)}'" for eid in evs_cerrados
                )
                raise HTTPException(
                    status_code=403,
                    detail=(
                        f"No se permiten cambios: el evento {etiquetas} está concluido. "
                        f"Para editar, un administrador debe reabrir la plantilla."
                    ),
                )
        # 1) Asistencias
        for a in data.asistencias:
            if a.disponibilidad_id:
                supabase.table('disponibilidad').update({
                    "asistencia_real": a.asistencia_real,
                    "updated_at": now,
                }).eq('id', a.disponibilidad_id).execute()
                resumen["asistencias"] += 1
            elif a.usuario_id and a.ensayo_id:
                # UPSERT por (usuario_id, ensayo_id) — la tabla tiene UNIQUE sobre estos dos campos
                ex = supabase.table('disponibilidad') \
                    .select('id') \
                    .eq('usuario_id', a.usuario_id).eq('ensayo_id', a.ensayo_id).limit(1).execute()
                if ex.data:
                    supabase.table('disponibilidad').update({
                        "asistencia_real": a.asistencia_real,
                        "updated_at": now,
                    }).eq('id', ex.data[0]['id']).execute()
                else:
                    supabase.table('disponibilidad').insert({
                        "usuario_id": a.usuario_id,
                        "ensayo_id": a.ensayo_id,
                        "asistencia_real": a.asistencia_real,
                    }).execute()
                resumen["asistencias"] += 1

        # 2) Gastos — UPSERT por (usuario_id, evento_id)
        # Iter F1 — Marcar provisional / validado en cache_extra y transporte_importe.
        is_admin_user = is_super_admin(current_user)
        admin_profile = current_user.get('profile') or {}
        admin_id_for_validar = admin_profile.get('id')
        admin_nombre = (
            f"{admin_profile.get('nombre','')} {admin_profile.get('apellidos','')}".strip()
            or (admin_profile.get('email') or '')
        )
        notif_provisional_pendientes = []  # acumulamos para enviar push tras commit
        for g in data.gastos:
            payload = {k: v for k, v in g.model_dump(exclude_unset=True).items() if k not in ('usuario_id', 'evento_id')}
            if not payload:
                payload = {}
            payload["updated_at"] = now

            # Iter F1 — leer fila actual para detectar cambios reales en cache_extra/transporte
            ex = supabase.table('gastos_adicionales') \
                .select('id,cache_extra,cache_extra_provisional,transporte_importe,transporte_provisional') \
                .eq('usuario_id', g.usuario_id).eq('evento_id', g.evento_id).limit(1).execute()
            ex_row = ex.data[0] if ex.data else None
            now_iso = datetime.now(timezone.utc).isoformat()

            def _flo(x):
                try: return float(x) if x is not None else 0.0
                except Exception: return 0.0

            for campo, prov_col, val_por_col, val_at_col in (
                ('cache_extra', 'cache_extra_provisional', 'cache_extra_validado_por', 'cache_extra_validado_at'),
                ('transporte_importe', 'transporte_provisional', 'transporte_validado_por', 'transporte_validado_at'),
            ):
                if campo not in payload:
                    continue
                nuevo = _flo(payload.get(campo))
                actual = _flo((ex_row or {}).get(campo))
                cambio = abs(nuevo - actual) > 0.001
                if not cambio:
                    continue
                if is_admin_user:
                    payload[prov_col] = False
                    payload[val_por_col] = admin_id_for_validar
                    payload[val_at_col] = now_iso
                else:
                    payload[prov_col] = True
                    payload[val_por_col] = None
                    payload[val_at_col] = None
                    notif_provisional_pendientes.append({
                        'usuario_id': g.usuario_id,
                        'evento_id': g.evento_id,
                        'campo': campo,
                        'importe': nuevo,
                    })

            if ex_row:
                supabase.table('gastos_adicionales').update(payload).eq('id', ex_row['id']).execute()
                gasto_id_after = ex_row['id']
            else:
                insert_payload = {"usuario_id": g.usuario_id, "evento_id": g.evento_id, **payload}
                ins_res = supabase.table('gastos_adicionales').insert(insert_payload).execute()
                gasto_id_after = (ins_res.data or [{}])[0].get('id')
            # Registro de actividad para cada importe nuevo provisional (para notificar al validar).
            for n in [x for x in notif_provisional_pendientes if x['usuario_id'] == g.usuario_id and x['evento_id'] == g.evento_id]:
                if not gasto_id_after:
                    continue
                try:
                    supabase.table('registro_actividad').insert({
                        "tipo": "importe_provisional_creado",
                        "descripcion": f"Importe provisional {n['campo']}={n['importe']}€ creado por {admin_nombre or 'gestor'}",
                        "usuario_id": admin_id_for_validar,
                        "usuario_nombre": admin_nombre,
                        "entidad_tipo": "gasto",
                        "entidad_id": gasto_id_after,
                    }).execute()
                except Exception:
                    pass
                n['gasto_id'] = gasto_id_after
            resumen["gastos"] += 1

        # Iter F1 — Notificar a super admins sobre importes pendientes (después del commit).
        if notif_provisional_pendientes and not is_admin_user:
            try:
                # Lookup nombres de músico + evento (cacheado por id).
                u_ids = list({n['usuario_id'] for n in notif_provisional_pendientes})
                e_ids = list({n['evento_id'] for n in notif_provisional_pendientes})
                u_rows = supabase.table('usuarios').select('id,nombre,apellidos') \
                    .in_('id', u_ids).execute().data or []
                e_rows = supabase.table('eventos').select('id,nombre') \
                    .in_('id', e_ids).execute().data or []
                u_by_id = {u['id']: f"{u.get('nombre','')} {u.get('apellidos','')}".strip() for u in u_rows}
                e_by_id = {e['id']: e.get('nombre') for e in e_rows}
                admins_rows = supabase.table('usuarios').select('id') \
                    .in_('rol', ['admin', 'director_general']).execute().data or []
                try:
                    from routes_push import notify_push as _notify_push
                except Exception:
                    _notify_push = None
                gestor_nombre = admin_nombre or 'Un gestor'
                for n in notif_provisional_pendientes:
                    tipo_label = 'caché extra' if n['campo'] == 'cache_extra' else 'transporte'
                    titulo = "⚠️ Importe pendiente validación"
                    body = (
                        f"{gestor_nombre} ha introducido {n['importe']:.2f}€ de "
                        f"{tipo_label} para {u_by_id.get(n['usuario_id'], 'músico')} "
                        f"en {e_by_id.get(n['evento_id'], 'evento')}"
                    )
                    for a in admins_rows:
                        aid = a.get('id')
                        if not aid:
                            continue
                        try:
                            supabase.table('notificaciones_gestor').insert({
                                "gestor_id": aid,
                                "tipo": "importe_pendiente_validacion",
                                "titulo": titulo,
                                "descripcion": body,
                                "entidad_tipo": "gasto",
                                "entidad_id": n.get('gasto_id'),
                                "leida": False,
                            }).execute()
                        except Exception:
                            pass
                        if _notify_push:
                            try:
                                _notify_push(aid, titulo, body, '/plantillas-definitivas', tipo='general')
                            except Exception:
                                pass
            except Exception as e:
                logger.warning(f"Iter F1 notif provisionales: {e}")

        # 3) Anotaciones
        asigs_ids_tocadas = set()
        for n in data.anotaciones:
            payload = {k: v for k, v in {
                "numero_atril": n.numero_atril,
                "letra": n.letra,
                "comentarios": n.comentario,
                "updated_at": now,
            }.items() if v is not None or k == "updated_at"}
            supabase.table('asignaciones').update(payload).eq('id', n.asignacion_id).execute()
            resumen["anotaciones"] += 1
            asigs_ids_tocadas.add(n.asignacion_id)

        # 4) Recalcular porcentaje_asistencia para las asignaciones tocadas por asistencias
        #    (buscamos las asignaciones de esos usuario+evento)
        pares = {(a.usuario_id, a.ensayo_id) for a in data.asistencias if a.usuario_id and a.ensayo_id}
        if pares:
            usuario_ids = list({u for (u, _) in pares})
            ensayo_ids = list({e for (_, e) in pares})
            # evento_id de esos ensayos
            e_res = supabase.table('ensayos').select('id,evento_id').in_('id', ensayo_ids).execute()
            eventos_de_ens = {e['id']: e['evento_id'] for e in (e_res.data or [])}
            evento_ids = list({eventos_de_ens[eid] for eid in ensayo_ids if eid in eventos_de_ens})
            # Para cada (usuario, evento) recalcular
            for uid in usuario_ids:
                for evid in evento_ids:
                    # contar ensayos + asistencias reales
                    all_e = supabase.table('ensayos').select('id').eq('evento_id', evid).execute().data or []
                    total = len(all_e)
                    if not total:
                        continue
                    all_ids = [x['id'] for x in all_e]
                    d = supabase.table('disponibilidad').select('asistencia_real') \
                        .eq('usuario_id', uid).in_('ensayo_id', all_ids).execute().data or []
                    # Nueva lógica: promedio de los porcentajes asistencia_real no-NULL
                    vals = [float(x['asistencia_real']) for x in d if x.get('asistencia_real') is not None]
                    pct = round(sum(vals) / len(vals), 2) if vals else 0.0
                    supabase.table('asignaciones').update({
                        "porcentaje_asistencia": pct, "updated_at": now
                    }).eq('usuario_id', uid).eq('evento_id', evid).execute()

        return {"ok": True, "resumen": resumen}
    except HTTPException:
        # Iter E1 — Re-elevar 403 (evento concluido) sin convertir a 500.
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al guardar: {str(e)}")


# ============================================================
# Iter E1 — Concluir / Reabrir plantilla del evento
# ============================================================

@router.post("/eventos/{evento_id}/concluir-plantilla")
async def concluir_plantilla(
    evento_id: str,
    current_user: dict = Depends(get_current_gestor),
):
    """Marca todas las asignaciones del evento con estado_cierre='cerrado_plantilla'.

    Cualquier gestor puede concluir un evento. Notifica vía push y notificaciones_gestor.
    Si existían recibos con regenerar_pendiente=TRUE, los regenera y notifica a admins.
    """
    profile = current_user.get('profile') or {}
    gestor_id = profile.get('id')
    gestor_nombre = (
        f"{profile.get('nombre','')} {profile.get('apellidos','')}".strip()
        or (profile.get('email') or 'Gestor')
    )

    ev_row = supabase.table('eventos').select('id,nombre').eq('id', evento_id).limit(1).execute().data or []
    if not ev_row:
        raise HTTPException(status_code=404, detail="Evento no encontrado")
    evento = ev_row[0]

    # 1) UPDATE asignaciones
    now_iso = datetime.now(timezone.utc).isoformat()
    upd = supabase.table('asignaciones').update({
        "estado_cierre": "cerrado_plantilla",
        "cerrado_plantilla_por": gestor_id,
        "cerrado_plantilla_at": now_iso,
        "updated_at": now_iso,
    }).eq('evento_id', evento_id).execute()
    actualizadas = len(upd.data or [])

    # 2) Regenerar recibos marcados como pendientes (tras una reapertura previa).
    regenerados = 0
    try:
        recibos_pend = supabase.table('recibos').select('id,asignacion_id') \
            .eq('evento_id', evento_id).eq('regenerar_pendiente', True).execute().data or []
    except Exception:
        recibos_pend = []
    if recibos_pend:
        try:
            from routes_documentos import generar_recibo as _gen_recibo
        except Exception as e:
            _gen_recibo = None
            logger.warning(f"concluir_plantilla: generar_recibo no disponible: {e}")
        for r in recibos_pend:
            try:
                if _gen_recibo:
                    _gen_recibo(r['asignacion_id'], force=True)
                supabase.table('recibos').update({
                    "regenerar_pendiente": False,
                    "actualizado_at": now_iso,
                }).eq('id', r['id']).execute()
                regenerados += 1
            except Exception as e:
                logger.warning(f"Error regenerando recibo {r['id']}: {e}")

    # 3) Notificar a todos los gestores (push + notificaciones_gestor)
    try:
        gestores_rows = supabase.table('usuarios').select('id') \
            .in_('rol', ['gestor', 'archivero', 'director_general', 'admin']) \
            .execute().data or []
    except Exception:
        gestores_rows = []
    titulo = "🏁 Evento concluido"
    body = (
        f"La plantilla de {evento.get('nombre','evento')} ha sido concluida. "
        f"Ya puedes procesar los pagos en Gestión Económica."
    )
    try:
        from routes_push import notify_push as _notify_push
    except Exception:
        _notify_push = None
    for g in gestores_rows:
        gid = g.get('id')
        if not gid:
            continue
        try:
            supabase.table('notificaciones_gestor').insert({
                "gestor_id": gid,
                "tipo": "evento_concluido",
                "titulo": titulo,
                "descripcion": body,
                "entidad_tipo": "evento",
                "entidad_id": evento_id,
                "leida": False,
            }).execute()
        except Exception:
            pass
        if _notify_push:
            try:
                _notify_push(gid, titulo, body, '/admin/asistencia-pagos', tipo='general')
            except Exception:
                pass

    # 4) Si se regeneraron recibos, notificar a admins/director.
    if regenerados > 0:
        try:
            admins_rows = supabase.table('usuarios').select('id') \
                .in_('rol', ['admin', 'director_general']).execute().data or []
        except Exception:
            admins_rows = []
        rt = "📄 Recibos regenerados"
        rb = (
            f"Se han regenerado {regenerados} recibo"
            f"{'s' if regenerados != 1 else ''} del evento "
            f"{evento.get('nombre','evento')} tras la reapertura."
        )
        for a in admins_rows:
            aid = a.get('id')
            if not aid:
                continue
            try:
                supabase.table('notificaciones_gestor').insert({
                    "gestor_id": aid,
                    "tipo": "recibos_regenerados",
                    "titulo": rt,
                    "descripcion": rb,
                    "entidad_tipo": "evento",
                    "entidad_id": evento_id,
                    "leida": False,
                }).execute()
            except Exception:
                pass
            if _notify_push:
                try:
                    _notify_push(aid, rt, rb, '/admin/recibos', tipo='general')
                except Exception:
                    pass

    # 5) Registro de actividad
    try:
        supabase.table('registro_actividad').insert({
            "tipo": "evento_concluido",
            "descripcion": f"Plantilla del evento '{evento.get('nombre','')}' concluida por {gestor_nombre}",
            "usuario_id": gestor_id,
            "usuario_nombre": gestor_nombre,
            "entidad_tipo": "evento",
            "entidad_id": evento_id,
        }).execute()
    except Exception:
        pass

    return {
        "ok": True,
        "evento_id": evento_id,
        "actualizadas": actualizadas,
        "recibos_regenerados": regenerados,
        "cerrado_plantilla_at": now_iso,
        "cerrado_plantilla_por_nombre": gestor_nombre,
    }


@router.post("/eventos/{evento_id}/reabrir-plantilla")
async def reabrir_plantilla(
    evento_id: str,
    current_user: dict = Depends(get_current_gestor),
):
    """Reabre una plantilla concluida. Solo super admins (director_general / admin /
    admin@convocatorias.com) pueden ejecutar esto.

    Si existen recibos para el evento, los marca con regenerar_pendiente=TRUE
    para que se regeneren automáticamente al volver a concluir.
    """
    if not is_super_admin(current_user):
        raise HTTPException(
            status_code=403,
            detail="Solo el director general o administradores pueden reabrir plantillas.",
        )
    profile = current_user.get('profile') or {}
    gestor_id = profile.get('id')
    gestor_nombre = (
        f"{profile.get('nombre','')} {profile.get('apellidos','')}".strip()
        or (profile.get('email') or 'Admin')
    )

    ev_row = supabase.table('eventos').select('id,nombre').eq('id', evento_id).limit(1).execute().data or []
    if not ev_row:
        raise HTTPException(status_code=404, detail="Evento no encontrado")
    evento = ev_row[0]

    now_iso = datetime.now(timezone.utc).isoformat()
    upd = supabase.table('asignaciones').update({
        "estado_cierre": "abierto",
        "cerrado_plantilla_por": None,
        "cerrado_plantilla_at": None,
        "updated_at": now_iso,
    }).eq('evento_id', evento_id).execute()
    actualizadas = len(upd.data or [])

    # Marcar recibos para regenerar si existen.
    recibos_marcados = 0
    try:
        rec = supabase.table('recibos').update({
            "regenerar_pendiente": True,
            "actualizado_at": now_iso,
        }).eq('evento_id', evento_id).execute()
        recibos_marcados = len(rec.data or [])
    except Exception as e:
        logger.warning(f"Error marcando recibos para regenerar: {e}")

    # Registro de actividad
    try:
        supabase.table('registro_actividad').insert({
            "tipo": "evento_reabierto",
            "descripcion": f"Plantilla del evento '{evento.get('nombre','')}' reabierta por {gestor_nombre}",
            "usuario_id": gestor_id,
            "usuario_nombre": gestor_nombre,
            "entidad_tipo": "evento",
            "entidad_id": evento_id,
        }).execute()
    except Exception:
        pass

    return {
        "ok": True,
        "evento_id": evento_id,
        "actualizadas": actualizadas,
        "recibos_marcados_regenerar": recibos_marcados,
    }


@router.get("/eventos/{evento_id}/historial-cierres")
async def get_historial_cierres(
    evento_id: str,
    current_user: dict = Depends(get_current_gestor),
):
    """Iter E1.1 — Historial de cierres y reaperturas del evento.

    Iter E2 — Ampliado a 4 tipos: cierre/reapertura de plantilla y de económico.
    """
    try:
        rows = supabase.table('registro_actividad') \
            .select('id,tipo,descripcion,usuario_nombre,created_at') \
            .eq('entidad_tipo', 'evento') \
            .eq('entidad_id', evento_id) \
            .in_('tipo', ['evento_concluido', 'evento_reabierto', 'economico_cerrado', 'economico_reabierto']) \
            .order('created_at', desc=True) \
            .execute().data or []
        return {"evento_id": evento_id, "entries": rows}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error cargando historial: {str(e)}")


# ============================================================
# Iter E2 — Cerrar / Reabrir económico del evento
# ============================================================

@router.post("/eventos/{evento_id}/cerrar-economico")
async def cerrar_economico(
    evento_id: str,
    current_user: dict = Depends(get_current_gestor),
):
    """Cierra económicamente un evento. Solo super admins.

    Pre-condición: la plantilla debe estar concluida (estado_cierre='cerrado_plantilla').
    Acción: estado_cierre='cerrado_economico', genera recibos faltantes para asignaciones
    pagadas, notifica push + notificaciones_gestor a admins/director_general.
    """
    if not is_super_admin(current_user):
        raise HTTPException(
            status_code=403,
            detail="Solo el director general o administradores pueden cerrar económicamente.",
        )
    profile = current_user.get('profile') or {}
    gestor_id = profile.get('id')
    gestor_nombre = (
        f"{profile.get('nombre','')} {profile.get('apellidos','')}".strip()
        or (profile.get('email') or 'Admin')
    )

    ev_row = supabase.table('eventos').select('id,nombre').eq('id', evento_id).limit(1).execute().data or []
    if not ev_row:
        raise HTTPException(status_code=404, detail="Evento no encontrado")
    evento = ev_row[0]

    # Pre-condición: plantilla concluida
    asigs = supabase.table('asignaciones').select(
        'id,estado_pago,estado_cierre'
    ).eq('evento_id', evento_id).eq('estado', 'confirmado').execute().data or []
    if not asigs:
        raise HTTPException(status_code=400, detail="El evento no tiene asignaciones confirmadas.")
    estados = {(a.get('estado_cierre') or 'abierto') for a in asigs}
    if estados != {'cerrado_plantilla'}:
        # Bien todas en abierto, bien mezcla, bien ya cerrado_economico → bloquear con mensaje claro
        if 'cerrado_economico' in estados:
            raise HTTPException(status_code=400, detail="El económico ya está cerrado.")
        raise HTTPException(
            status_code=400,
            detail="Debes concluir primero la plantilla del evento antes de cerrar el económico.",
        )

    now_iso = datetime.now(timezone.utc).isoformat()

    # 1) UPDATE asignaciones
    upd = supabase.table('asignaciones').update({
        "estado_cierre": "cerrado_economico",
        "cerrado_economico_por": gestor_id,
        "cerrado_economico_at": now_iso,
        "updated_at": now_iso,
    }).eq('evento_id', evento_id).eq('estado', 'confirmado').execute()
    actualizadas = len(upd.data or [])

    # 2) Generar recibos faltantes (solo para asignaciones pagadas sin recibo).
    recibos_generados = 0
    try:
        pagadas = [a for a in asigs if (a.get('estado_pago') == 'pagado')]
        if pagadas:
            asig_ids_pagadas = [a['id'] for a in pagadas]
            existentes = supabase.table('recibos').select('asignacion_id') \
                .in_('asignacion_id', asig_ids_pagadas).execute().data or []
            con_recibo = {r['asignacion_id'] for r in existentes if r.get('asignacion_id')}
            faltan = [aid for aid in asig_ids_pagadas if aid not in con_recibo]
            if faltan:
                try:
                    from routes_documentos import generar_recibo as _gen_recibo
                except Exception as e:
                    _gen_recibo = None
                    logger.warning(f"cerrar_economico: generar_recibo no disponible: {e}")
                if _gen_recibo:
                    for aid in faltan:
                        try:
                            res = _gen_recibo(aid)
                            if res:
                                recibos_generados += 1
                        except Exception as e:
                            logger.warning(f"Error generando recibo {aid}: {e}")
    except Exception as e:
        logger.warning(f"cerrar_economico: bloque recibos falló: {e}")

    # 3) Notificación push + notificaciones_gestor a admins + director_general.
    try:
        admins_rows = supabase.table('usuarios').select('id') \
            .in_('rol', ['admin', 'director_general']).execute().data or []
    except Exception:
        admins_rows = []
    titulo = "💰 Económico cerrado"
    body = f"El económico del evento {evento.get('nombre','evento')} ha sido cerrado."
    try:
        from routes_push import notify_push as _notify_push
    except Exception:
        _notify_push = None
    for a in admins_rows:
        aid = a.get('id')
        if not aid:
            continue
        try:
            supabase.table('notificaciones_gestor').insert({
                "gestor_id": aid,
                "tipo": "economico_cerrado",
                "titulo": titulo,
                "descripcion": body,
                "entidad_tipo": "evento",
                "entidad_id": evento_id,
                "leida": False,
            }).execute()
        except Exception:
            pass
        if _notify_push:
            try:
                _notify_push(aid, titulo, body, '/admin/asistencia-pagos', tipo='general')
            except Exception:
                pass

    # 4) Registro de actividad
    try:
        supabase.table('registro_actividad').insert({
            "tipo": "economico_cerrado",
            "descripcion": f"Económico del evento '{evento.get('nombre','')}' cerrado por {gestor_nombre}",
            "usuario_id": gestor_id,
            "usuario_nombre": gestor_nombre,
            "entidad_tipo": "evento",
            "entidad_id": evento_id,
        }).execute()
    except Exception:
        pass

    return {
        "ok": True,
        "evento_id": evento_id,
        "actualizadas": actualizadas,
        "recibos_generados": recibos_generados,
        "cerrado_economico_at": now_iso,
        "cerrado_economico_por_nombre": gestor_nombre,
    }


@router.post("/eventos/{evento_id}/reabrir-economico")
async def reabrir_economico(
    evento_id: str,
    current_user: dict = Depends(get_current_gestor),
):
    """Reabre el económico de un evento. Solo super admins.

    Vuelve estado_cierre a 'cerrado_plantilla' (NO a 'abierto'). Marca recibos del
    evento con regenerar_pendiente=TRUE para que se regeneren al volver a cerrar.
    """
    if not is_super_admin(current_user):
        raise HTTPException(
            status_code=403,
            detail="Solo el director general o administradores pueden reabrir económicamente.",
        )
    profile = current_user.get('profile') or {}
    gestor_id = profile.get('id')
    gestor_nombre = (
        f"{profile.get('nombre','')} {profile.get('apellidos','')}".strip()
        or (profile.get('email') or 'Admin')
    )

    ev_row = supabase.table('eventos').select('id,nombre').eq('id', evento_id).limit(1).execute().data or []
    if not ev_row:
        raise HTTPException(status_code=404, detail="Evento no encontrado")
    evento = ev_row[0]

    now_iso = datetime.now(timezone.utc).isoformat()
    upd = supabase.table('asignaciones').update({
        "estado_cierre": "cerrado_plantilla",
        "cerrado_economico_por": None,
        "cerrado_economico_at": None,
        "updated_at": now_iso,
    }).eq('evento_id', evento_id).eq('estado_cierre', 'cerrado_economico').execute()
    actualizadas = len(upd.data or [])

    # Marcar recibos para regenerar al volver a cerrar.
    recibos_marcados = 0
    try:
        rec = supabase.table('recibos').update({
            "regenerar_pendiente": True,
            "actualizado_at": now_iso,
        }).eq('evento_id', evento_id).execute()
        recibos_marcados = len(rec.data or [])
    except Exception as e:
        logger.warning(f"reabrir_economico: error marcando recibos: {e}")

    # Registro de actividad
    try:
        supabase.table('registro_actividad').insert({
            "tipo": "economico_reabierto",
            "descripcion": f"Económico del evento '{evento.get('nombre','')}' reabierto por {gestor_nombre}",
            "usuario_id": gestor_id,
            "usuario_nombre": gestor_nombre,
            "entidad_tipo": "evento",
            "entidad_id": evento_id,
        }).execute()
    except Exception:
        pass

    return {
        "ok": True,
        "evento_id": evento_id,
        "actualizadas": actualizadas,
        "recibos_marcados_regenerar": recibos_marcados,
    }


# ==================== Cachets Config & Upload justificantes ====================
# NOTE: Modelos CachetRow y CachetBaseItem + endpoints /cachets-config y /cachets-base
# movidos a routes_economia.py durante el refactor de feb 2026.


@router.post("/plantillas-definitivas/justificante")
async def upload_justificante(
    archivo: UploadFile = File(...),
    usuario_id: str = "",
    evento_id: str = "",
    tipo: str = "otros",  # 'transporte' | 'alojamiento' | 'otros'
    current_user: dict = Depends(get_current_gestor)
):
    """Sube un justificante al bucket `justificantes` y devuelve la URL pública."""
    import os
    from supabase import create_client
    if tipo not in ("transporte", "alojamiento", "otros"):
        raise HTTPException(status_code=400, detail="Tipo inválido")
    if not usuario_id or not evento_id:
        raise HTTPException(status_code=400, detail="usuario_id y evento_id son obligatorios")

    content = await archivo.read()
    ext = (archivo.filename or "").rsplit(".", 1)[-1].lower() or "bin"
    ts = int(datetime.now().timestamp())
    path = f"{evento_id}/{usuario_id}/{tipo}_{ts}.{ext}"

    admin_client = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_KEY'])
    try:
        admin_client.storage.from_('justificantes').upload(
            path, content,
            {"content-type": archivo.content_type or "application/octet-stream", "upsert": "true"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error subiendo archivo: {str(e)}")

    public_url = admin_client.storage.from_('justificantes').get_public_url(path)
    return {"url": public_url, "path": path}


# NOTE: /cachets-config/{evento_id} GET/PUT movidos a routes_economia.py


# ==================================================================
# CACHETS BASE (evento_id IS NULL)
# NOTE: endpoints /cachets-base y /cachets-config/{id}/copy-from-base
# movidos a routes_economia.py durante el refactor de feb 2026.
# ==================================================================


# ==================================================================
# ENSAYO_INSTRUMENTOS — Convocatoria por instrumento
# ==================================================================

SECCIONES_INSTR = {
    'cuerda':        ['Violín', 'Viola', 'Violonchelo', 'Contrabajo'],
    'viento_madera': ['Flauta', 'Oboe', 'Clarinete', 'Fagot'],
    'viento_metal':  ['Trompa', 'Trompeta', 'Trombón', 'Tuba'],
    'percusion':     ['Percusión'],
    'teclados':      ['Piano', 'Órgano'],
    'coro':          ['Soprano', 'Alto', 'Tenor', 'Barítono'],
}
ALL_INSTR_FLAT = [i for lst in SECCIONES_INSTR.values() for i in lst]


class EnsayoInstrumentoRow(BaseModel):
    instrumento: str
    convocado: bool = True


@router.get("/ensayos/{ensayo_id}/instrumentos")
async def get_ensayo_instrumentos(
    ensayo_id: str,
    current_user: dict = Depends(get_current_gestor),
):
    """Devuelve lista de {instrumento, convocado} para un ensayo.
    Si no hay filas: todos convocados (default TRUE) para los 18 instrumentos estándar."""
    try:
        r = supabase.table('ensayo_instrumentos') \
            .select('instrumento,convocado') \
            .eq('ensayo_id', ensayo_id).execute()
        rows = r.data or []
        existing = {row['instrumento']: bool(row.get('convocado', True)) for row in rows}
        out = []
        for instr in ALL_INSTR_FLAT:
            out.append({"instrumento": instr, "convocado": existing.get(instr, True)})
        return {"ensayo_id": ensayo_id, "instrumentos": out, "ha_custom": len(existing) > 0}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener convocatoria: {str(e)}")


@router.put("/ensayos/{ensayo_id}/instrumentos")
async def put_ensayo_instrumentos(
    ensayo_id: str,
    rows: List[EnsayoInstrumentoRow],
    current_user: dict = Depends(get_current_gestor),
):
    """UPSERT masivo de la convocatoria del ensayo."""
    try:
        existing = supabase.table('ensayo_instrumentos') \
            .select('id,instrumento').eq('ensayo_id', ensayo_id).execute().data or []
        idx = {e['instrumento']: e['id'] for e in existing}
        creados = 0; actualizados = 0
        for row in rows:
            payload = {
                "ensayo_id": ensayo_id,
                "instrumento": row.instrumento,
                "convocado": bool(row.convocado),
            }
            if row.instrumento in idx:
                supabase.table('ensayo_instrumentos').update(payload) \
                    .eq('id', idx[row.instrumento]).execute()
                actualizados += 1
            else:
                supabase.table('ensayo_instrumentos').insert(payload).execute()
                creados += 1
        return {"ok": True, "creados": creados, "actualizados": actualizados}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al guardar convocatoria: {str(e)}")


@router.get("/ensayo-instrumentos-bulk")
async def get_ensayo_instrumentos_bulk(
    ensayo_ids: str,  # coma-separada
    current_user: dict = Depends(get_current_gestor),
):
    """Devuelve mapping {ensayo_id: {instrumento: convocado}} para lista de ensayos.
    Si no hay filas para un ensayo/instrumento → se asume convocado=TRUE.
    Sólo se devuelven overrides (filas != default), el consumidor interpreta el default."""
    try:
        ids = [x.strip() for x in ensayo_ids.split(',') if x.strip()]
        if not ids:
            return {"mapping": {}}
        r = supabase.table('ensayo_instrumentos') \
            .select('ensayo_id,instrumento,convocado') \
            .in_('ensayo_id', ids).execute()
        out = {}
        for row in (r.data or []):
            eid = row['ensayo_id']
            out.setdefault(eid, {})[row['instrumento']] = bool(row.get('convocado', True))
        return {"mapping": out}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


def _is_convocado(ensayo_instr_map: dict, ensayo_id: str, instrumento: Optional[str]) -> bool:
    """True si el instrumento está convocado en ese ensayo. Default TRUE."""
    if not instrumento:
        return True
    per_ensayo = ensayo_instr_map.get(ensayo_id)
    if not per_ensayo:
        return True  # sin overrides → todos convocados
    # Si existe clave instrumento, respeta su valor; si no, default TRUE
    return bool(per_ensayo.get(instrumento, True))



@router.delete("/eventos/{evento_id}")
async def delete_evento(
    evento_id: str,
    current_user: dict = Depends(get_current_gestor)
):
    """Delete evento (CASCADE deletes asignaciones and ensayos)"""
    try:
        response = supabase.table('eventos').delete().eq('id', evento_id).execute()
        
        return {"message": "Evento eliminado"}
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al eliminar evento: {str(e)}"
        )

# ==================== Asignaciones ====================

@router.post("/asignaciones")
async def create_asignacion(
    data: AsignacionCreate,
    current_user: dict = Depends(get_current_gestor)
):
    """Assign musician to evento"""
    try:
        asignacion_data = {
            **data.model_dump(),
            "estado": "pendiente",
            "estado_pago": "pendiente"
        }
        
        response = supabase.table('asignaciones').insert(asignacion_data).execute()
        
        return {
            "message": "Músico asignado al evento",
            "asignacion": response.data[0] if response.data else None
        }
        
    except Exception as e:
        error_msg = str(e)
        if "duplicate" in error_msg or "unique" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Este músico ya está asignado a este evento"
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear asignación: {error_msg}"
        )

@router.get("/asignaciones/evento/{evento_id}")
async def get_asignaciones_evento(
    evento_id: str,
    current_user: dict = Depends(get_current_gestor)
):
    """Get all asignaciones for an evento"""
    try:
        response = supabase.table('asignaciones') \
            .select('*, usuario:usuarios(*)') \
            .eq('evento_id', evento_id) \
            .execute()
        
        return {"asignaciones": response.data or []}
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al cargar asignaciones: {str(e)}"
        )

@router.delete("/asignaciones/{asignacion_id}")
async def delete_asignacion(
    asignacion_id: str,
    current_user: dict = Depends(get_current_gestor)
):
    """Remove musician from evento"""
    try:
        response = supabase.table('asignaciones').delete().eq('id', asignacion_id).execute()
        
        return {"message": "Asignación eliminada"}
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al eliminar asignación: {str(e)}"
        )

# ==================== Ensayos ====================

@router.post("/ensayos")
async def create_ensayo(
    data: EnsayoCreate,
    current_user: dict = Depends(get_current_gestor)
):
    """Create rehearsal/performance for evento"""
    try:
        response = supabase.table('ensayos').insert(data.model_dump()).execute()
        
        return {
            "message": "Ensayo creado",
            "ensayo": response.data[0] if response.data else None
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear ensayo: {str(e)}"
        )

@router.get("/eventos/{evento_id}/ensayos")
async def get_ensayos_de_evento(
    evento_id: str,
    current_user: dict = Depends(get_current_gestor)
):
    """B3 · Lista los ensayos/funciones de un evento (usado por MontajeRiderSection)."""
    try:
        res = supabase.table('ensayos').select('*') \
            .eq('evento_id', evento_id) \
            .order('fecha').order('hora').execute()
        return {"ensayos": res.data or []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al listar ensayos: {str(e)}")


@router.delete("/ensayos/{ensayo_id}")
async def delete_ensayo(
    ensayo_id: str,
    current_user: dict = Depends(get_current_gestor)
):
    """Delete ensayo"""
    try:
        response = supabase.table('ensayos').delete().eq('id', ensayo_id).execute()
        
        return {"message": "Ensayo eliminado"}
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al eliminar ensayo: {str(e)}"
        )

@router.put("/ensayos/{ensayo_id}")
async def update_ensayo(
    ensayo_id: str,
    data: EnsayoUpdate,
    current_user: dict = Depends(get_current_gestor)
):
    """Update existing ensayo fields (fecha, hora, tipo, obligatorio, lugar, notas)."""
    try:
        payload = data.model_dump(exclude_none=True)
        if not payload:
            return {"message": "Sin cambios", "ensayo": None}
        response = supabase.table('ensayos').update(payload).eq('id', ensayo_id).execute()
        return {
            "message": "Ensayo actualizado",
            "ensayo": response.data[0] if response.data else None
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al actualizar ensayo: {str(e)}"
        )

# ==================== Músicos ====================

@router.get("/musicos")
async def get_musicos(
    q: Optional[str] = None,
    instrumento: Optional[str] = None,
    estado: Optional[str] = None,
    current_user: dict = Depends(get_current_gestor)
):
    """Get all musicians with optional filters.
    
    Query params:
    - q: search by nombre, apellidos or email (ilike)
    - instrumento: filter by instrumento
    - estado: 'activo' | 'inactivo'
    """
    try:
        query = supabase.table('usuarios').select('*').eq('rol', 'musico')
        
        if instrumento:
            query = query.eq('instrumento', instrumento)
        
        if estado:
            query = query.eq('estado', estado)
        
        if q:
            # Supabase OR filter: search on nombre, apellidos, email
            safe = q.replace(',', ' ').strip()
            query = query.or_(f"nombre.ilike.%{safe}%,apellidos.ilike.%{safe}%,email.ilike.%{safe}%")
        
        response = query.order('apellidos', desc=False).execute()
        
        return {"musicos": response.data or []}
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al cargar músicos: {str(e)}"
        )


@router.get("/musicos/{musico_id}")
async def get_musico_detalle(
    musico_id: str,
    current_user: dict = Depends(get_current_gestor)
):
    """Ficha completa de un músico: perfil + historial de eventos + pagos + reclamaciones."""
    try:
        # Perfil
        u_res = supabase.table('usuarios').select('*').eq('id', musico_id).eq('rol', 'musico').single().execute()
        musico = u_res.data
        if not musico:
            raise HTTPException(status_code=404, detail="Músico no encontrado")
        
        # Historial eventos
        asigs_res = supabase.table('asignaciones') \
            .select('*, evento:eventos(id,nombre,temporada,fecha_inicio,fecha_fin,estado)') \
            .eq('usuario_id', musico_id) \
            .order('created_at', desc=True) \
            .execute()
        asignaciones = asigs_res.data or []
        
        # Totales de pago
        total_cobrado = 0.0
        total_pendiente = 0.0
        for a in asignaciones:
            try: imp = float(a.get('importe') or 0)
            except: imp = 0
            if a.get('estado_pago') == 'pagado': total_cobrado += imp
            else: total_pendiente += imp
        
        # Reclamaciones
        recl_res = supabase.table('reclamaciones') \
            .select('*, evento:eventos(nombre,temporada)') \
            .eq('usuario_id', musico_id) \
            .order('fecha_creacion', desc=True) \
            .execute()
        
        return {
            "musico": musico,
            "asignaciones": asignaciones,
            "total_cobrado": round(total_cobrado, 2),
            "total_pendiente": round(total_pendiente, 2),
            "reclamaciones": recl_res.data or []
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@router.delete("/musicos/{musico_id}")
async def delete_musico(
    musico_id: str,
    current_user: dict = Depends(get_current_gestor)
):
    """
    Elimina un músico:
    - Bloquea si tiene asignaciones confirmadas en eventos activos (estado='abierto'/'en_curso').
    - Si no, elimina el perfil de `usuarios` y el usuario de Supabase Auth.
    - Registra la acción en `registro_actividad`.
    """
    import os
    from supabase import create_client
    admin_client = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_KEY'])

    # 1) Cargar perfil
    try:
        u_res = admin_client.table('usuarios').select('id,user_id,email,nombre,apellidos,rol') \
            .eq('id', musico_id).single().execute()
        musico = u_res.data
    except Exception:
        musico = None
    if not musico:
        raise HTTPException(status_code=404, detail="Músico no encontrado")
    if musico.get('rol') != 'musico':
        raise HTTPException(status_code=400, detail="El usuario no es un músico")

    # 2) Comprobar asignaciones confirmadas en eventos activos
    try:
        confirmadas_res = admin_client.table('asignaciones') \
            .select('id, evento:eventos(id,nombre,estado)') \
            .eq('usuario_id', musico_id) \
            .eq('estado', 'confirmado') \
            .execute()
        activas = [
            a for a in (confirmadas_res.data or [])
            if (a.get('evento') or {}).get('estado') in ('abierto', 'en_curso')
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error comprobando asignaciones: {str(e)}")

    if activas:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="No se puede eliminar un músico con convocatorias confirmadas activas"
        )

    # 3) Eliminar perfil (CASCADE se lleva asignaciones, reclamaciones, etc.)
    try:
        admin_client.table('usuarios').delete().eq('id', musico_id).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al eliminar perfil: {str(e)}")

    # 4) Eliminar usuario en Supabase Auth (best-effort)
    auth_deleted = False
    auth_error = None
    if musico.get('user_id'):
        try:
            admin_client.auth.admin.delete_user(musico['user_id'])
            auth_deleted = True
        except Exception as e:
            auth_error = str(e)[:200]

    # 5) Registro de actividad
    try:
        gp = current_user.get('profile') or {}
        gname = f"{gp.get('nombre','')} {gp.get('apellidos','')}".strip()
        mname = f"{musico.get('nombre','')} {musico.get('apellidos','')}".strip()
        supabase.table('registro_actividad').insert({
            'tipo': 'musico_eliminado',
            'descripcion': f"{gname} eliminó al músico {mname} ({musico.get('email')})",
            'usuario_id': gp.get('id'),
            'usuario_nombre': gname,
            'entidad_tipo': 'musico',
            'entidad_id': musico_id,
            'metadata': {
                'email': musico.get('email'),
                'auth_deleted': auth_deleted,
                'auth_error': auth_error,
            }
        }).execute()
    except Exception:
        pass

    return {
        "message": "Músico eliminado correctamente",
        "auth_deleted": auth_deleted,
        "auth_error": auth_error,
    }


@router.get("/pendientes")
async def get_pendientes(current_user: dict = Depends(get_current_gestor)):
    """Contadores de pendientes para el sidebar y el dashboard del gestor."""
    try:
        # Reclamaciones pendientes
        r_res = supabase.table('reclamaciones').select('id', count='exact') \
            .in_('estado', ['pendiente', 'en_gestion']).execute()
        reclamaciones_pendientes = r_res.count or 0
        
        # Perfiles actualizados en últimas 24h (si la migración está aplicada)
        perfiles_actualizados = 0
        try:
            from datetime import timezone, timedelta
            cutoff_24h = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
            p_res = supabase.table('usuarios').select('id', count='exact') \
                .eq('rol', 'musico') \
                .gte('ultima_actualizacion_perfil', cutoff_24h) \
                .execute()
            perfiles_actualizados = p_res.count or 0
        except Exception:
            perfiles_actualizados = 0
        
        # Respuestas nuevas desde el último acceso del gestor
        from datetime import timezone, timedelta
        gestor_profile = current_user.get('profile') or {}
        ultimo_acceso = gestor_profile.get('ultimo_acceso_gestor') or (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
        a_res = supabase.table('asignaciones').select('id', count='exact') \
            .in_('estado', ['confirmado', 'rechazado']) \
            .gte('fecha_respuesta', ultimo_acceso) \
            .execute()
        respuestas_nuevas = a_res.count or 0
        
        # Tareas próximas (24h) - si existe la tabla tareas
        tareas_proximas = 0
        try:
            cutoff_24h_fw = (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat()
            t_res = supabase.table('tareas').select('id', count='exact') \
                .lte('fecha_limite', cutoff_24h_fw) \
                .neq('estado', 'completada') \
                .execute()
            tareas_proximas = t_res.count or 0
        except Exception:
            tareas_proximas = 0

        # Comentarios de equipo en estado 'pendiente' (sólo hilos raíz).
        comentarios_pendientes = 0
        try:
            c_res = supabase.table('comentarios_equipo').select('id', count='exact') \
                .eq('estado', 'pendiente') \
                .is_('parent_id', 'null') \
                .execute()
            comentarios_pendientes = c_res.count or 0
        except Exception:
            comentarios_pendientes = 0

        try:
            sr_res = supabase.table('solicitudes_registro') \
                .select('id', count='exact') \
                .eq('estado', 'pendiente') \
                .execute()
            solicitudes_pendientes = sr_res.count or 0
        except Exception:
            solicitudes_pendientes = 0

        # Iter F1 — Importes provisionales pendientes de validación (solo super admins).
        importes_pendientes_validacion = 0
        if is_super_admin(current_user):
            try:
                ipv = supabase.table('gastos_adicionales').select('id', count='exact') \
                    .or_('cache_extra_provisional.eq.true,transporte_provisional.eq.true').execute()
                importes_pendientes_validacion = ipv.count or 0
            except Exception:
                importes_pendientes_validacion = 0

        return {
            "reclamaciones_pendientes": reclamaciones_pendientes,
            "perfiles_actualizados": perfiles_actualizados,
            "respuestas_nuevas": respuestas_nuevas,
            "tareas_proximas": tareas_proximas,
            "comentarios_pendientes": comentarios_pendientes,
            "solicitudes_pendientes": solicitudes_pendientes,
            "importes_pendientes_validacion": importes_pendientes_validacion,
            "ultimo_acceso_gestor": ultimo_acceso
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# Iter F1 — Validar importes provisionales (solo super admins).
@router.post("/gastos/{gasto_id}/validar")
async def validar_importe_provisional(
    gasto_id: str,
    payload: dict,
    current_user: dict = Depends(get_current_gestor),
):
    """Valida un importe provisional (cache_extra o transporte) de un gasto_adicional.
    Solo super admins. Body: { campo: 'cache_extra' | 'transporte' }.
    """
    if not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Solo el director general o administradores pueden validar importes.")
    campo = (payload or {}).get('campo')
    if campo not in ('cache_extra', 'transporte'):
        raise HTTPException(status_code=400, detail="Campo inválido. Debe ser 'cache_extra' o 'transporte'.")
    prov_col = 'cache_extra_provisional' if campo == 'cache_extra' else 'transporte_provisional'
    val_por_col = 'cache_extra_validado_por' if campo == 'cache_extra' else 'transporte_validado_por'
    val_at_col = 'cache_extra_validado_at' if campo == 'cache_extra' else 'transporte_validado_at'
    importe_col = 'cache_extra' if campo == 'cache_extra' else 'transporte_importe'

    rows = supabase.table('gastos_adicionales').select(
        f'id,usuario_id,evento_id,{importe_col},{prov_col}'
    ).eq('id', gasto_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(status_code=404, detail="Gasto no encontrado.")
    row = rows[0]
    if not row.get(prov_col):
        raise HTTPException(status_code=400, detail="Este importe ya está validado.")

    profile = current_user.get('profile') or {}
    admin_id = profile.get('id')
    admin_nombre = (
        f"{profile.get('nombre','')} {profile.get('apellidos','')}".strip()
        or (profile.get('email') or 'Admin')
    )
    now_iso = datetime.now(timezone.utc).isoformat()
    importe = float(row.get(importe_col) or 0)

    supabase.table('gastos_adicionales').update({
        prov_col: False,
        val_por_col: admin_id,
        val_at_col: now_iso,
        "updated_at": now_iso,
    }).eq('id', gasto_id).execute()

    # Buscar al gestor que lo introdujo (vía registro_actividad) y notificarle.
    try:
        ra_rows = supabase.table('registro_actividad').select('usuario_id') \
            .eq('entidad_tipo', 'gasto').eq('entidad_id', gasto_id) \
            .eq('tipo', 'importe_provisional_creado') \
            .order('created_at', desc=True).limit(1).execute().data or []
        if ra_rows:
            gestor_id = ra_rows[0].get('usuario_id')
            if gestor_id:
                tipo_label = 'caché extra' if campo == 'cache_extra' else 'transporte'
                titulo = "✅ Importe validado"
                body = f"Tu importe de {importe:.2f}€ de {tipo_label} ha sido validado."
                try:
                    supabase.table('notificaciones_gestor').insert({
                        "gestor_id": gestor_id,
                        "tipo": "importe_validado",
                        "titulo": titulo,
                        "descripcion": body,
                        "entidad_tipo": "gasto",
                        "entidad_id": gasto_id,
                        "leida": False,
                    }).execute()
                except Exception as e:
                    logger.warning(f"Iter F1 validar: error insertando notificacion_gestor: {e}")
                try:
                    from routes_push import notify_push as _notify_push
                    _notify_push(gestor_id, titulo, body, '/plantillas-definitivas', tipo='general')
                except Exception as e:
                    logger.warning(f"Iter F1 validar: error notify_push: {e}")
    except Exception as e:
        logger.warning(f"Iter F1 validar: error buscando gestor original: {e}")

    # Registro de actividad de la validación.
    try:
        supabase.table('registro_actividad').insert({
            "tipo": "importe_validado",
            "descripcion": f"Importe {campo}={importe}€ validado por {admin_nombre}",
            "usuario_id": admin_id,
            "usuario_nombre": admin_nombre,
            "entidad_tipo": "gasto",
            "entidad_id": gasto_id,
        }).execute()
    except Exception:
        pass

    return {
        "ok": True,
        "gasto_id": gasto_id,
        "campo": campo,
        "validado_por_nombre": admin_nombre,
        "validado_at": now_iso,
    }


@router.post("/marcar-acceso")
async def marcar_acceso_gestor(current_user: dict = Depends(get_current_gestor)):
    try:
        gestor_id = (current_user.get('profile') or {}).get('id')
        if not gestor_id:
            raise HTTPException(status_code=400, detail="Perfil no encontrado")
        now = datetime.now().isoformat()
        supabase.table('usuarios').update({'ultimo_acceso_gestor': now}).eq('id', gestor_id).execute()
        return {"ultimo_acceso": now}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@router.get("/instrumentos")
async def get_instrumentos_disponibles(current_user: dict = Depends(get_current_gestor)):
    """Return distinct list of instrumentos for filter dropdown"""
    try:
        response = supabase.table('usuarios') \
            .select('instrumento') \
            .eq('rol', 'musico') \
            .execute()
        
        instrumentos = sorted({u.get('instrumento') for u in (response.data or []) if u.get('instrumento')})
        return {"instrumentos": instrumentos}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al cargar instrumentos: {str(e)}"
        )

def _generate_temp_password(length: int = 12) -> str:
    """Generate a secure temp password: letters + digits (meets 8+ upper + digit rule)."""
    alphabet = string.ascii_letters + string.digits
    while True:
        pwd = ''.join(secrets.choice(alphabet) for _ in range(length))
        if any(c.isupper() for c in pwd) and any(c.isdigit() for c in pwd) and any(c.islower() for c in pwd):
            return pwd


# ==================== Importación masiva de músicos ====================

# Columnas aceptadas en la plantilla (orden visible en el Excel).
IMPORT_MUSICOS_HEADERS = [
    "nombre", "apellidos", "email", "telefono", "instrumento",
    "especialidad", "nivel_estudios", "localidad", "baremo",
    "dni", "direccion", "fecha_nacimiento", "nacionalidad", "bio"
]


@router.get("/musicos-import/plantilla")
async def descargar_plantilla_musicos(current_user: dict = Depends(get_current_gestor)):
    """Genera y descarga un Excel con cabeceras + fila de ejemplo + hoja INSTRUCCIONES."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Músicos"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for idx, name in enumerate(IMPORT_MUSICOS_HEADERS, start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = max(14, len(name) + 4)
    # Fila de ejemplo (comentada con color) para guiar al usuario
    ejemplo = ["Ana", "García", "ana@ejemplo.com", "+34600111222", "Violín",
               "Música clásica", "Superior finalizado", "Madrid", "8.5",
               "12345678A", "Calle Mayor 1, Madrid",
               "1990-05-12", "Española", "Breve biografía opcional"]
    for idx, value in enumerate(ejemplo, start=1):
        c = ws.cell(row=2, column=idx, value=value)
        c.font = Font(color="94A3B8", italic=True)
    ws.row_dimensions[2].height = 18

    # ---- Pestaña INSTRUCCIONES ----
    ws2 = wb.create_sheet("INSTRUCCIONES")
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 80
    title_font = Font(bold=True, size=14, color="0F172A")
    bold = Font(bold=True)
    h = ws2.cell(row=1, column=1, value="Plantilla de importación de músicos")
    h.font = title_font
    ws2.merge_cells('A1:B1')

    instrucciones = [
        ("", ""),
        ("Campo", "Descripción / valores aceptados"),
        ("nombre", "Texto libre. Obligatorio."),
        ("apellidos", "Texto libre. Obligatorio."),
        ("email", "Email único. Obligatorio."),
        ("telefono", "Teléfono con prefijo internacional, ej: +34600111222."),
        ("instrumento", "Violín / Viola / Violonchelo / Contrabajo / Flauta / Oboe / Clarinete / Fagot / Trompa / Trompeta / Trombón / Tuba / Percusión / Piano / Arpa / Coro / etc."),
        ("especialidad", "Texto libre. Ej: Música clásica, Jazz, Música antigua…"),
        ("nivel_estudios", "Uno de: Superior finalizado / Superior cursando / Profesional finalizado / Profesional cursando."),
        ("localidad", "Ciudad o localidad de residencia."),
        ("baremo", "Número decimal entre 0 y 10 (ej: 8.5). Se acepta coma o punto decimal."),
        ("dni", "DNI/NIE/Pasaporte. Texto libre."),
        ("direccion", "Dirección postal completa."),
        ("fecha_nacimiento", "Formato ISO YYYY-MM-DD (ej: 1990-05-12) o reconocido por Excel."),
        ("nacionalidad", "Nacionalidad. Texto libre."),
        ("bio", "Biografía breve. Texto libre opcional."),
    ]
    for i, (campo, desc) in enumerate(instrucciones, start=2):
        a = ws2.cell(row=i, column=1, value=campo)
        b = ws2.cell(row=i, column=2, value=desc)
        if i == 3:  # cabecera "Campo / Descripción"
            a.font = bold; b.font = bold
            a.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            b.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        else:
            a.font = bold
            b.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[i].height = 22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_musicos.xlsx"'}
    )


def _parse_musicos_file(raw: bytes, filename: str) -> List[dict]:
    """Lee un fichero xlsx o csv y devuelve una lista de dicts normalizados."""
    name = (filename or "").lower()
    rows: List[dict] = []
    if name.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        # Detectar delimitador (coma o punto y coma)
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except Exception:
            dialect = csv.excel
        reader = csv.DictReader(StringIO(text), dialect=dialect)
        for r in reader:
            rows.append({(k or "").strip().lower(): (v.strip() if isinstance(v, str) else v)
                         for k, v in r.items() if k})
    else:
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        headers: List[str] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [str(c).strip().lower() if c is not None else "" for c in row]
                continue
            if not any(cell not in (None, "") for cell in row):
                continue
            d = {}
            for h, v in zip(headers, row):
                if not h:
                    continue
                if isinstance(v, datetime):
                    v = v.date().isoformat()
                d[h] = (str(v).strip() if v is not None else "")
            rows.append(d)
    # Quitar filas completamente vacías o sin email
    return [r for r in rows if any(v for v in r.values())]


@router.post("/musicos-import/preview")
async def importar_musicos_preview(
    archivo: UploadFile = File(...),
    current_user: dict = Depends(get_current_gestor)
):
    """Preview: devuelve los primeros 5 registros + total + errores de validación."""
    content = await archivo.read()
    try:
        rows = _parse_musicos_file(content, archivo.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo: {str(e)}")

    missing_headers = [h for h in ("nombre", "apellidos", "email") if h not in (rows[0].keys() if rows else [])]
    return {
        "total_filas": len(rows),
        "preview": rows[:5],
        "missing_required_headers": missing_headers,
        "columnas_plantilla": IMPORT_MUSICOS_HEADERS,
    }


@router.post("/musicos-import")
async def importar_musicos(
    archivo: UploadFile = File(...),
    current_user: dict = Depends(get_current_gestor)
):
    """
    Importa masivamente músicos desde Excel/CSV.
    Para cada fila:
      - Crea usuario en Supabase Auth con password temporal aleatorio (8 chars).
      - Crea el perfil en tabla `usuarios` con requiere_cambio_password=True.
      - Si el email ya existe, lo marca como "ya_existente" y continúa.
      - Cualquier otro error se reporta en el informe final.
    """
    import os
    from supabase import create_client

    content = await archivo.read()
    try:
        rows = _parse_musicos_file(content, archivo.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo: {str(e)}")

    admin_client = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_KEY'])

    creados: List[dict] = []
    existentes: List[dict] = []
    errores: List[dict] = []

    for idx, row in enumerate(rows, start=2):  # start=2 porque fila 1 es header
        email = (row.get("email") or "").strip().lower()
        nombre = (row.get("nombre") or "").strip()
        apellidos = (row.get("apellidos") or "").strip()
        if not email or not nombre or not apellidos:
            errores.append({"fila": idx, "email": email, "motivo": "Faltan campos obligatorios (nombre, apellidos, email)"})
            continue

        # Saltar si ya existe en tabla usuarios
        try:
            existente = admin_client.table('usuarios').select('id,email').eq('email', email).limit(1).execute()
            if existente.data:
                existentes.append({"fila": idx, "email": email})
                continue
        except Exception:
            pass

        # Crear en Auth
        temp_password = _generate_temp_password(8)
        created_user_id = None
        try:
            auth_resp = admin_client.auth.admin.create_user({
                "email": email,
                "password": temp_password,
                "email_confirm": True,
                "app_metadata": {"rol": "musico"}
            })
            created_user = auth_resp.user if hasattr(auth_resp, 'user') else None
            if not created_user:
                errores.append({"fila": idx, "email": email, "motivo": "No se pudo crear el usuario en Auth"})
                continue
            created_user_id = created_user.id
        except Exception as e:
            msg = str(e).lower()
            if "already" in msg or "exists" in msg or "duplicate" in msg or "registered" in msg:
                existentes.append({"fila": idx, "email": email})
            else:
                errores.append({"fila": idx, "email": email, "motivo": f"Auth: {str(e)[:150]}"})
            continue

        # Normalizar baremo (acepta coma decimal)
        baremo_raw = row.get("baremo")
        baremo_val = None
        if baremo_raw not in (None, ""):
            try:
                baremo_val = float(str(baremo_raw).strip().replace(",", "."))
            except (ValueError, TypeError):
                baremo_val = None

        # Perfil en usuarios
        profile_payload = {
            "user_id": created_user_id,
            "email": email,
            "nombre": nombre,
            "apellidos": apellidos,
            "telefono": row.get("telefono") or None,
            "instrumento": row.get("instrumento") or None,
            "especialidad": row.get("especialidad") or None,
            "nivel_estudios": row.get("nivel_estudios") or None,
            "localidad": row.get("localidad") or None,
            "baremo": baremo_val,
            "dni": row.get("dni") or None,
            "direccion": row.get("direccion") or None,
            "fecha_nacimiento": row.get("fecha_nacimiento") or None,
            "nacionalidad": row.get("nacionalidad") or None,
            "bio": row.get("bio") or None,
            "rol": "musico",
            "estado": "activo",
            "requiere_cambio_password": True,
        }
        profile_payload = {k: v for k, v in profile_payload.items() if v not in (None, "")}
        try:
            ins = admin_client.table('usuarios').insert(profile_payload).execute()
            profile = ins.data[0] if ins.data else None
        except Exception as e:
            try:
                admin_client.auth.admin.delete_user(created_user_id)
            except Exception:
                pass
            errores.append({"fila": idx, "email": email, "motivo": f"Perfil: {str(e)[:150]}"})
            continue

        creados.append({"fila": idx, "email": email, "id": profile.get("id") if profile else None})

    # Registro de actividad
    try:
        gp = current_user.get('profile') or {}
        gname = f"{gp.get('nombre','')} {gp.get('apellidos','')}".strip()
        supabase.table('registro_actividad').insert({
            'tipo': 'importacion_musicos',
            'descripcion': f"{gname} importó {len(creados)} músicos (existentes: {len(existentes)}, errores: {len(errores)})",
            'usuario_id': gp.get('id'),
            'usuario_nombre': gname,
            'entidad_tipo': 'musico',
            'metadata': {'creados': len(creados), 'existentes': len(existentes), 'errores': len(errores)}
        }).execute()
    except Exception:
        pass

    return {
        "creados": creados,
        "existentes": existentes,
        "errores": errores,
        "resumen": {
            "total_procesadas": len(rows),
            "creados": len(creados),
            "ya_existentes": len(existentes),
            "errores": len(errores),
        }
    }


@router.post("/musicos/crear")
async def crear_musico(
    data: MusicoCreate,
    current_user: dict = Depends(get_current_gestor)
):
    """
    Crea un nuevo músico:
    - Genera contraseña temporal
    - Crea usuario en Supabase Auth (email + password, confirmado)
    - Asigna rol 'musico' en app_metadata
    - Crea perfil en tabla usuarios con requiere_cambio_password=True
    - Envía email con credenciales vía Resend (si está configurado)
    """
    import os
    from supabase import create_client
    # Fresh admin client to avoid any session interference from auth verification flow
    admin_client = create_client(
        os.environ['SUPABASE_URL'],
        os.environ['SUPABASE_KEY']  # service role
    )

    temp_password = _generate_temp_password(12)
    email_str = data.email
    created_user_id = None

    try:
        # 1. Crear usuario en Supabase Auth (email confirmed)
        try:
            auth_resp = admin_client.auth.admin.create_user({
                "email": email_str,
                "password": temp_password,
                "email_confirm": True,
                "app_metadata": {"rol": "musico"}
            })
            created_user = auth_resp.user if hasattr(auth_resp, 'user') else None
            if not created_user:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="No se pudo crear el usuario en Supabase Auth"
                )
            created_user_id = created_user.id
        except HTTPException:
            raise
        except Exception as e:
            msg = str(e).lower()
            if "already" in msg or "exists" in msg or "duplicate" in msg or "registered" in msg:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Este email ya está registrado"
                )
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Error al crear usuario: {str(e)}"
            )

        # 2. Crear perfil en tabla usuarios
        profile_payload = {
            "user_id": created_user_id,
            "email": email_str,
            "nombre": data.nombre,
            "apellidos": data.apellidos,
            "instrumento": data.instrumento,
            "telefono": data.telefono,
            "rol": "musico",
            "estado": "activo",
            "requiere_cambio_password": True
        }
        profile_payload = {k: v for k, v in profile_payload.items() if v is not None}

        try:
            insert_res = admin_client.table('usuarios').insert(profile_payload).execute()
            profile = insert_res.data[0] if insert_res.data else None
        except Exception as e:
            try:
                admin_client.auth.admin.delete_user(created_user_id)
            except Exception:
                pass
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Error al crear perfil: {str(e)}"
            )

        # 3. Enviar email con credenciales (no bloqueante de errores)
        email_result = await send_musico_credentials_email(
            to_email=email_str,
            nombre=data.nombre,
            password_temporal=temp_password,
            usuario_id=created_user_id
        )

        # Registro de actividad
        try:
            gp = current_user.get('profile') or {}
            gname = f"{gp.get('nombre','')} {gp.get('apellidos','')}".strip()
            supabase.table('registro_actividad').insert({
                'tipo': 'musico_creado',
                'descripcion': f"{gname} creó al músico {data.nombre} {data.apellidos}",
                'usuario_id': gp.get('id'),
                'usuario_nombre': gname,
                'entidad_tipo': 'musico',
                'entidad_id': profile['id'] if profile else None
            }).execute()
        except Exception:
            pass

        return {
            "message": "Músico creado correctamente" + (
                " y email de credenciales enviado" if email_result.get("sent") else ". Email NO enviado (configurar RESEND_API_KEY)"
            ),
            "musico": profile,
            "password_temporal": temp_password,
            "email_enviado": email_result.get("sent", False),
            "email_error": email_result.get("reason") if not email_result.get("sent") else None
        }

    except HTTPException:
        raise
    except Exception as e:
        if created_user_id:
            try:
                admin_client.auth.admin.delete_user(created_user_id)
            except Exception:
                pass
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear músico: {str(e)}"
        )


@router.post("/musicos/invite")
async def invite_musico(
    data: MusicoCreate,
    current_user: dict = Depends(get_current_gestor)
):
    """
    (Legacy) Invite musician - sends magic link for first login.
    Creates profile without auth account (created on first magic link login).
    """
    try:
        # Send magic link invitation
        supabase.auth.sign_in_with_otp({
            "email": data.email,
            "options": {
                "email_redirect_to": f"/portal"
            }
        })
        
        # Create profile (user_id will be null until they log in)
        profile_data = {
            **data.model_dump(),
            "rol": "musico",
            "estado": "activo",
            "user_id": None  # Will be linked on first login
        }
        
        response = supabase.table('usuarios').insert(profile_data).execute()
        
        return {
            "message": f"Invitación enviada a {data.email}",
            "musico": response.data[0] if response.data else None
        }
        
    except Exception as e:
        error_msg = str(e)
        if "duplicate" in error_msg or "unique" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Este email ya está registrado"
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al invitar músico: {error_msg}"
        )


# ==================== Export Excel ====================

@router.get("/export/xlsx")
async def export_excel(current_user: dict = Depends(get_current_gestor)):
    """
    Export Usuarios, Eventos y Asignaciones en un fichero .xlsx con 3 hojas.
    """
    try:
        usuarios_res = supabase.table('usuarios').select('*').order('apellidos', desc=False).execute()
        eventos_res = supabase.table('eventos').select('*').order('created_at', desc=True).execute()
        asignaciones_res = supabase.table('asignaciones') \
            .select('*, usuario:usuarios(nombre,apellidos,email,instrumento), evento:eventos(nombre,temporada)') \
            .execute()

        wb = Workbook()

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        def write_sheet(ws, headers, rows):
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            for r_idx, row in enumerate(rows, start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            # Auto width (aprox)
            for col_idx, h in enumerate(headers, start=1):
                max_len = len(str(h))
                for row in rows:
                    val = row[col_idx - 1] if col_idx - 1 < len(row) else ""
                    max_len = max(max_len, len(str(val)) if val is not None else 0)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

        # Hoja 1: Usuarios
        ws_u = wb.active
        ws_u.title = "Usuarios"
        u_headers = ["Nombre", "Apellidos", "Email", "Rol", "Instrumento", "Teléfono", "Estado", "Fecha Alta"]
        u_rows = []
        for u in (usuarios_res.data or []):
            u_rows.append([
                u.get('nombre', ''),
                u.get('apellidos', ''),
                u.get('email', ''),
                u.get('rol', ''),
                u.get('instrumento', '') or '',
                u.get('telefono', '') or '',
                u.get('estado', ''),
                u.get('fecha_alta', '') or u.get('created_at', '') or ''
            ])
        write_sheet(ws_u, u_headers, u_rows)

        # Hoja 2: Eventos
        ws_e = wb.create_sheet("Eventos")
        e_headers = ["Nombre", "Temporada", "Tipo", "Estado", "Fecha Inicio", "Fecha Fin", "Lugar", "Descripción"]
        e_rows = []
        for e in (eventos_res.data or []):
            e_rows.append([
                e.get('nombre', ''),
                e.get('temporada', '') or '',
                e.get('tipo', '') or '',
                e.get('estado', '') or '',
                e.get('fecha_inicio', '') or '',
                e.get('fecha_fin', '') or '',
                e.get('lugar', '') or '',
                e.get('descripcion', '') or ''
            ])
        write_sheet(ws_e, e_headers, e_rows)

        # Hoja 3: Asignaciones
        ws_a = wb.create_sheet("Asignaciones")
        a_headers = ["Evento", "Temporada", "Músico", "Email", "Instrumento", "Estado", "Estado Pago", "Importe"]
        a_rows = []
        for a in (asignaciones_res.data or []):
            ev = a.get('evento') or {}
            us = a.get('usuario') or {}
            nombre_completo = f"{us.get('nombre', '')} {us.get('apellidos', '')}".strip()
            a_rows.append([
                ev.get('nombre', ''),
                ev.get('temporada', '') or '',
                nombre_completo,
                us.get('email', '') or '',
                us.get('instrumento', '') or '',
                a.get('estado', '') or '',
                a.get('estado_pago', '') or '',
                a.get('importe', 0) or 0
            ])
        write_sheet(ws_a, a_headers, a_rows)

        # Guardar en memoria
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"opus_manager_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al exportar Excel: {str(e)}"
        )


# ==================== RECORDATORIOS (Bloque 3) ====================

# Catálogo de los 10 recordatorios predefinidos
RECORDATORIOS_PREDEFINIDOS = [
    {"tipo": "nueva_asignacion", "nombre": "Nueva asignación", "descripcion": "Email al músico al ser asignado a un evento", "destinatario_default": "musico", "dias_default": None},
    {"tipo": "respuesta_7d", "nombre": "Recordatorio respuesta 7 días antes del límite", "descripcion": "Solo a músicos sin responder", "destinatario_default": "musico", "dias_default": 7},
    {"tipo": "respuesta_3d", "nombre": "Recordatorio respuesta 3 días antes", "descripcion": "Solo a músicos sin responder", "destinatario_default": "musico", "dias_default": 3},
    {"tipo": "respuesta_24h", "nombre": "Último aviso 24h", "descripcion": "Solo a músicos sin responder", "destinatario_default": "musico", "dias_default": 1},
    {"tipo": "aviso_ensayo_24h", "nombre": "Aviso ensayo 24h antes", "descripcion": "Solo a músicos confirmados", "destinatario_default": "musico", "dias_default": 1},
    {"tipo": "aviso_funcion_48h", "nombre": "Aviso función 48h antes", "descripcion": "Solo a músicos confirmados", "destinatario_default": "musico", "dias_default": 2},
    {"tipo": "alerta_baja_respuesta", "nombre": "Alerta baja respuesta <50% a 5 días", "descripcion": "Aviso al gestor si menos del 50% ha respondido", "destinatario_default": "gestor", "dias_default": 5},
    {"tipo": "pago_pendiente_3d", "nombre": "Recordatorio pago pendiente 3 días antes", "descripcion": "Aviso al gestor para pagos pendientes", "destinatario_default": "gestor", "dias_default": 3},
    {"tipo": "confirmacion_cobro", "nombre": "Confirmación de cobro al pagar", "descripcion": "Email al músico al marcarse como pagado", "destinatario_default": "musico", "dias_default": None},
    {"tipo": "resumen_diario", "nombre": "Resumen diario 8:00 con eventos activos", "descripcion": "Resumen diario para el gestor", "destinatario_default": "gestor", "dias_default": None},
]


class RecordatorioConfigPayload(BaseModel):
    tipo: str
    activo: Optional[bool] = None
    dias_antes: Optional[int] = None
    mensaje_personalizado: Optional[str] = None
    destinatario: Optional[str] = None


@router.get("/eventos/{evento_id}/recordatorios")
async def get_recordatorios(evento_id: str, current_user: dict = Depends(get_current_gestor)):
    """Devuelve la lista de 10 recordatorios con su config actual (o defaults)."""
    try:
        cfgs_res = supabase.table('recordatorios_config') \
            .select('*') \
            .eq('evento_id', evento_id) \
            .execute()
        cfgs_by_tipo = {c['tipo']: c for c in (cfgs_res.data or [])}

        resultado = []
        for predef in RECORDATORIOS_PREDEFINIDOS:
            cfg = cfgs_by_tipo.get(predef['tipo'], {})
            resultado.append({
                **predef,
                "activo": cfg.get('activo', False),
                "dias_antes": cfg.get('dias_antes', predef['dias_default']),
                "mensaje_personalizado": cfg.get('mensaje_personalizado', ''),
                "destinatario": cfg.get('destinatario', predef['destinatario_default']),
                "config_id": cfg.get('id')
            })
        return {"recordatorios": resultado}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@router.put("/eventos/{evento_id}/recordatorios")
async def upsert_recordatorio(
    evento_id: str,
    payload: RecordatorioConfigPayload,
    current_user: dict = Depends(get_current_gestor)
):
    """Activa/desactiva/edita un recordatorio (upsert por tipo+evento)."""
    try:
        data = payload.model_dump(exclude_none=True)
        data['evento_id'] = evento_id
        data['updated_at'] = datetime.utcnow().isoformat()

        # Upsert: intento update, si no existe inserto
        existing = supabase.table('recordatorios_config') \
            .select('id') \
            .eq('evento_id', evento_id) \
            .eq('tipo', payload.tipo) \
            .execute()
        if existing.data:
            supabase.table('recordatorios_config') \
                .update(data) \
                .eq('id', existing.data[0]['id']) \
                .execute()
        else:
            supabase.table('recordatorios_config').insert(data).execute()
        return {"message": "Recordatorio actualizado"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ==================== EMAIL LOG ====================

@router.get("/emails/log")
async def get_email_log(
    limit: int = 200,
    tipo: Optional[str] = None,
    estado: Optional[str] = None,
    evento_id: Optional[str] = None,
    desde: Optional[str] = None,  # 'YYYY-MM-DD'
    hasta: Optional[str] = None,
    current_user: dict = Depends(get_current_gestor)
):
    """Historial de emails enviados con filtros y contadores."""
    try:
        query = supabase.table('email_log').select('*')
        if tipo:
            query = query.eq('tipo', tipo)
        if estado:
            query = query.eq('estado', estado)
        if evento_id:
            query = query.eq('evento_id', evento_id)
        if desde:
            query = query.gte('created_at', f"{desde}T00:00:00")
        if hasta:
            query = query.lte('created_at', f"{hasta}T23:59:59")
        res = query.order('created_at', desc=True).limit(min(limit, 500)).execute()
        emails = res.data or []

        # Enrichment: get destinatario nombre de tabla usuarios
        dest_emails = list({e['destinatario'] for e in emails if e.get('destinatario')})
        dest_info = {}
        if dest_emails:
            try:
                u_res = supabase.table('usuarios').select('email,nombre,apellidos') \
                    .in_('email', dest_emails).execute()
                dest_info = {u['email']: f"{u.get('nombre','')} {u.get('apellidos','')}".strip() for u in (u_res.data or [])}
            except Exception:
                dest_info = {}
        for e in emails:
            e['destinatario_nombre'] = dest_info.get(e.get('destinatario'), '')

        # Contadores del día actual
        from datetime import timezone, timedelta
        now = datetime.now(timezone.utc)
        hoy_inicio = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
        mes_inicio = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        total_hoy = supabase.table('email_log').select('id', count='exact') \
            .gte('created_at', hoy_inicio).eq('estado', 'enviado').execute().count or 0
        total_error_hoy = supabase.table('email_log').select('id', count='exact') \
            .gte('created_at', hoy_inicio).eq('estado', 'error').execute().count or 0
        total_mes = supabase.table('email_log').select('id', count='exact') \
            .gte('created_at', mes_inicio).eq('estado', 'enviado').execute().count or 0

        return {
            "emails": emails,
            "contadores": {
                "enviados_hoy": total_hoy,
                "errores_hoy": total_error_hoy,
                "enviados_mes": total_mes
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@router.get("/emails/status")
async def get_resend_status(current_user: dict = Depends(get_current_gestor)):
    """Estado de conexión con Resend."""
    from email_service import _get_api_key, _get_sender
    api_key = _get_api_key()
    if not api_key:
        return {
            "conectado": False,
            "sender": _get_sender(),
            "mensaje": "RESEND_API_KEY no configurada",
            "enviados_mes": 0
        }
    try:
        # Validamos la key intentando listar domains. Si Resend responde
        # con un error de "restricted" significa que la key es válida pero
        # sólo permite enviar emails (sending key). Ese caso lo consideramos OK.
        import resend
        resend.api_key = api_key
        import asyncio
        try:
            await asyncio.to_thread(resend.Domains.list)
        except Exception as inner:
            msg = str(inner).lower()
            if 'restricted' not in msg and 'insufficient' not in msg:
                raise
        
        # Contar emails del mes
        from datetime import timezone
        now = datetime.now(timezone.utc)
        mes_inicio = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        c = supabase.table('email_log').select('id', count='exact') \
            .gte('created_at', mes_inicio).eq('estado', 'enviado').execute().count or 0
        
        return {
            "conectado": True,
            "sender": _get_sender(),
            "mensaje": "Conexión establecida con Resend",
            "enviados_mes": c
        }
    except Exception as e:
        return {
            "conectado": False,
            "sender": _get_sender(),
            "mensaje": f"Error: {str(e)[:200]}",
            "enviados_mes": 0
        }


class EmailTestRequest(BaseModel):
    destinatario: EmailStr
    tipo: str = "prueba"  # 'prueba' | 'nueva_convocatoria' | 'recordatorio' | ...
    asunto: Optional[str] = None
    html: Optional[str] = None


EMAIL_TEST_TEMPLATES = {
    "prueba": {
        "asunto": "OPUS MANAGER — Email de prueba",
        "html": "<h2>✅ Email de prueba</h2><p>Si recibes este correo, Resend está correctamente configurado.</p><p>— OPUS MANAGER</p>"
    },
    "nueva_convocatoria": {
        "asunto": "[Prueba] Nueva convocatoria",
        "html": "<h2>🎼 Nueva convocatoria</h2><p>Has sido asignado a un nuevo evento. Accede al portal para confirmar tu asistencia.</p><p>— OPUS MANAGER</p>"
    },
    "recordatorio": {
        "asunto": "[Prueba] Recordatorio de respuesta",
        "html": "<h2>⏰ Recordatorio</h2><p>Aún no has respondido a la convocatoria. Por favor, confirma tu asistencia lo antes posible.</p>"
    },
    "aviso_ensayo": {
        "asunto": "[Prueba] Aviso de ensayo",
        "html": "<h2>🎵 Recordatorio de ensayo</h2><p>Tu próximo ensayo es mañana. Consulta hora y lugar en el portal.</p>"
    },
    "confirmacion_cobro": {
        "asunto": "[Prueba] Confirmación de pago",
        "html": "<h2>💰 Pago procesado</h2><p>Hemos procesado tu pago correctamente. Consulta el detalle en tu historial.</p>"
    }
}


@router.get("/emails/preview")
async def email_preview(tipo: str = "prueba", current_user: dict = Depends(get_current_gestor)):
    """Devuelve la previsualización del email según tipo."""
    tmpl = EMAIL_TEST_TEMPLATES.get(tipo, EMAIL_TEST_TEMPLATES["prueba"])
    return {"asunto": tmpl["asunto"], "html": tmpl["html"]}


@router.post("/emails/test")
async def email_test_send(
    payload: EmailTestRequest,
    current_user: dict = Depends(get_current_gestor)
):
    """Envía un email de prueba."""
    from email_service import _send_email
    tmpl = EMAIL_TEST_TEMPLATES.get(payload.tipo, EMAIL_TEST_TEMPLATES["prueba"])
    asunto = payload.asunto or tmpl["asunto"]
    html = payload.html or tmpl["html"]
    gestor_profile = current_user.get('profile') or {}
    r = await _send_email(
        to_email=payload.destinatario,
        subject=asunto,
        html=html,
        tipo=f"test_{payload.tipo}",
        usuario_id=gestor_profile.get('id')
    )
    return r


class ReenviarEmailRequest(BaseModel):
    email_log_id: str


@router.post("/emails/reenviar")
async def reenviar_email(
    payload: ReenviarEmailRequest,
    current_user: dict = Depends(get_current_gestor)
):
    """Reenvía un email a partir de un log entry."""
    from email_service import _send_email
    try:
        res = supabase.table('email_log').select('*').eq('id', payload.email_log_id).single().execute()
        item = res.data
        if not item:
            raise HTTPException(status_code=404, detail="Email no encontrado")
        # Cuerpo simple si no tenemos el HTML original
        html = f"<p>Reenvío del email: {item.get('asunto')}</p>"
        r = await _send_email(
            to_email=item['destinatario'],
            subject=f"[Reenvío] {item.get('asunto','')}",
            html=html,
            tipo=f"reenvio_{item.get('tipo','')}",
            usuario_id=item.get('usuario_id'),
            evento_id=item.get('evento_id')
        )
        return r
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ==================== RECLAMACIONES (Gestor) ====================

@router.get("/reclamaciones")
async def get_reclamaciones_gestor(current_user: dict = Depends(get_current_gestor)):
    """Todas las reclamaciones para el panel del gestor."""
    try:
        res = supabase.table('reclamaciones') \
            .select('*, usuario:usuarios!reclamaciones_usuario_id_fkey(nombre,apellidos,email), evento:eventos(nombre,temporada)') \
            .order('fecha_creacion', desc=True) \
            .execute()
        return {"reclamaciones": res.data or []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


class ReclamacionUpdatePayload(BaseModel):
    estado: Optional[str] = None
    respuesta_gestor: Optional[str] = None


@router.put("/reclamaciones/{reclamacion_id}")
async def update_reclamacion(
    reclamacion_id: str,
    payload: ReclamacionUpdatePayload,
    current_user: dict = Depends(get_current_gestor)
):
    """El gestor actualiza el estado/respuesta de una reclamación."""
    try:
        gestor_profile = current_user.get('profile') or {}
        data = payload.model_dump(exclude_none=True)
        # Trazabilidad: quién gestiona la reclamación
        data['gestor_id'] = gestor_profile.get('id')
        data['gestor_nombre'] = f"{gestor_profile.get('nombre','')} {gestor_profile.get('apellidos','')}".strip()
        if data.get('estado') in ('resuelta', 'rechazada'):
            data['fecha_resolucion'] = datetime.utcnow().isoformat()
        res = supabase.table('reclamaciones').update(data).eq('id', reclamacion_id).execute()
        recl_actualizada = res.data[0] if res.data else None

        # Push (Bloque PWA): notificar al músico si hay respuesta nueva o cambio de estado
        try:
            if recl_actualizada and (data.get('respuesta_gestor') or data.get('estado')):
                from routes_push import notify_push
                musico_id = recl_actualizada.get('usuario_id')
                if musico_id:
                    estado = data.get('estado') or recl_actualizada.get('estado')
                    notify_push(
                        musico_id,
                        f"📬 Respuesta a tu reclamación ({estado or 'actualizada'})",
                        (data.get('respuesta_gestor') or '')[:140] or 'Tu reclamación ha sido actualizada.',
                        '/portal',
                        tipo='reclamacion',
                    )
        except Exception:
            pass
        
        # Registro de actividad
        try:
            supabase.table('registro_actividad').insert({
                'tipo': f"reclamacion_{data.get('estado','actualizada')}",
                'descripcion': f"Reclamación {data.get('estado') or 'actualizada'} por {data['gestor_nombre']}",
                'usuario_id': gestor_profile.get('id'),
                'usuario_nombre': data['gestor_nombre'],
                'entidad_tipo': 'reclamacion',
                'entidad_id': reclamacion_id
            }).execute()
        except Exception:
            pass
        
        return {"message": "Reclamación actualizada", "reclamacion": recl_actualizada}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ==================== COMENTARIOS INTERNOS ====================

class ComentarioPayload(BaseModel):
    tipo: str  # 'reclamacion' | 'evento'
    entidad_id: str
    contenido: str


@router.get("/comentarios")
async def get_comentarios(
    tipo: str,
    entidad_id: str,
    current_user: dict = Depends(get_current_gestor)
):
    """Lista comentarios internos de una entidad."""
    try:
        res = supabase.table('comentarios_internos') \
            .select('*') \
            .eq('tipo', tipo) \
            .eq('entidad_id', entidad_id) \
            .order('created_at', desc=True) \
            .execute()
        return {"comentarios": res.data or []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@router.post("/comentarios")
async def crear_comentario(
    payload: ComentarioPayload,
    current_user: dict = Depends(get_current_gestor)
):
    """Crea un comentario interno entre gestores."""
    try:
        gestor_profile = current_user.get('profile') or {}
        nombre = f"{gestor_profile.get('nombre','')} {gestor_profile.get('apellidos','')}".strip() or gestor_profile.get('email')
        data = {
            "tipo": payload.tipo,
            "entidad_id": payload.entidad_id,
            "gestor_id": gestor_profile.get('id'),
            "gestor_nombre": nombre,
            "contenido": payload.contenido
        }
        res = supabase.table('comentarios_internos').insert(data).execute()
        comentario_creado = res.data[0] if res.data else None

        # TAREA 3B — Si el comentario es sobre una tarea, notificar al responsable (si es distinto del autor)
        if payload.tipo == 'tarea' and comentario_creado:
            try:
                tarea_row = supabase.table('tareas').select('id,titulo,responsable_id') \
                    .eq('id', payload.entidad_id).limit(1).execute().data or []
                if tarea_row:
                    t = tarea_row[0]
                    resp_id = t.get('responsable_id')
                    if resp_id and resp_id != gestor_profile.get('id'):
                        supabase.table('notificaciones_gestor').insert({
                            "gestor_id": resp_id,
                            "tipo": "comentario_tarea",
                            "titulo": f"Nuevo comentario en: {t.get('titulo')}",
                            "descripcion": f"{nombre}: {payload.contenido[:80]}",
                            "entidad_tipo": "tarea",
                            "entidad_id": t['id'],
                            "leida": False,
                        }).execute()
            except Exception:
                pass

        # Notificar a otros gestores si hay menciones @
        import re
        menciones = re.findall(r'@([\w]+)', payload.contenido)
        if menciones:
            try:
                gs = supabase.table('usuarios').select('id,nombre,apellidos') \
                    .eq('rol', 'gestor').execute().data or []
                for g in gs:
                    if g['id'] == gestor_profile.get('id'):
                        continue
                    full = f"{g.get('nombre','')}{g.get('apellidos','')}".lower()
                    if any(m.lower() in full for m in menciones):
                        supabase.table('notificaciones_gestor').insert({
                            "gestor_id": g['id'],
                            "tipo": "mencion_comentario",
                            "titulo": f"{nombre} te ha mencionado",
                            "descripcion": payload.contenido[:200],
                            "entidad_tipo": "comentario",
                            "entidad_id": res.data[0]['id']
                        }).execute()
            except Exception:
                pass
        
        return {"comentario": comentario_creado}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ==================== NOTIFICACIONES ====================

@router.get("/notificaciones")
async def get_notificaciones(
    limit: int = 50,
    current_user: dict = Depends(get_current_gestor)
):
    """Lista notificaciones del gestor actual."""
    try:
        gestor_profile = current_user.get('profile') or {}
        gestor_id = gestor_profile.get('id')
        res = supabase.table('notificaciones_gestor') \
            .select('*') \
            .eq('gestor_id', gestor_id) \
            .order('created_at', desc=True) \
            .limit(min(limit, 100)) \
            .execute()
        items = res.data or []
        no_leidas = supabase.table('notificaciones_gestor').select('id', count='exact') \
            .eq('gestor_id', gestor_id).eq('leida', False).execute().count or 0
        return {"notificaciones": items, "no_leidas": no_leidas}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@router.put("/notificaciones/{notif_id}/leer")
async def marcar_leida(notif_id: str, current_user: dict = Depends(get_current_gestor)):
    gestor_id = (current_user.get('profile') or {}).get('id')
    supabase.table('notificaciones_gestor').update({"leida": True}) \
        .eq('id', notif_id).eq('gestor_id', gestor_id).execute()
    return {"message": "Leída"}


@router.post("/notificaciones/leer-todas")
async def marcar_todas_leidas(current_user: dict = Depends(get_current_gestor)):
    gestor_id = (current_user.get('profile') or {}).get('id')
    supabase.table('notificaciones_gestor').update({"leida": True}) \
        .eq('gestor_id', gestor_id).eq('leida', False).execute()
    return {"message": "Todas marcadas como leídas"}


# ==================== REGISTRO DE ACTIVIDAD ====================

@router.get("/actividad")
async def get_actividad(
    limit: int = 100,
    current_user: dict = Depends(get_current_gestor)
):
    """Registro de actividad global."""
    try:
        res = supabase.table('registro_actividad') \
            .select('*') \
            .order('created_at', desc=True) \
            .limit(min(limit, 500)) \
            .execute()
        return {"actividad": res.data or []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ==================================================================
# BLOQUE 1 — Presupuestos (CRUD real sobre tabla `presupuestos`)
# ==================================================================

# ==================================================================
# BLOQUE 4 — Presupuestos (Bloque B)
# NOTE: Modelos PresupuestoItem / PresupuestoBulkItem y endpoints /presupuestos
# movidos a routes_economia.py durante el refactor de feb 2026.
# ==================================================================


# ==================================================================
# BLOQUE 5 — Gestión Económica (contabilidad, marcar pagos, export xlsx)
# ==================================================================

@router.get("/gestion-economica")
async def get_gestion_economica(
    temporada: Optional[str] = None,
    current_user: dict = Depends(get_current_gestor),
):
    """
    Devuelve la misma estructura que /plantillas-definitivas pero añadiendo
    datos de contabilidad por músico: iban, swift, titulaciones y estado_pago.
    """
    try:
        # Arrancamos con asignaciones confirmadas
        # Iter E2 — añadidos campos de cierre (plantilla + económico).
        a_res = supabase.table('asignaciones') \
            .select('id,usuario_id,evento_id,estado,estado_pago,cache_presupuestado,importe,numero_atril,letra,comentarios,nivel_estudios,porcentaje_asistencia,estado_cierre,cerrado_plantilla_por,cerrado_plantilla_at,cerrado_economico_por,cerrado_economico_at') \
            .eq('estado', 'confirmado') \
            .execute()
        confirmadas = a_res.data or []
        if not confirmadas:
            return {"eventos": [], "total_temporada": 0}

        evento_ids = list({a['evento_id'] for a in confirmadas})
        usuario_ids = list({a['usuario_id'] for a in confirmadas})

        ev_q = supabase.table('eventos').select('*').in_('id', evento_ids)
        if temporada:
            ev_q = ev_q.eq('temporada', temporada)
        eventos_raw = ev_q.order('fecha_inicio', desc=False).execute().data or []
        evento_ids_vistos = [e['id'] for e in eventos_raw]

        # Iter E2 — Resolver cierre por evento (mismo patrón que plantillas-definitivas).
        cierre_by_evento_econ: Dict[str, Dict] = {}
        for a in confirmadas:
            evid = a['evento_id']
            if evid not in evento_ids_vistos:
                continue
            ec = a.get('estado_cierre') or 'abierto'
            cur = cierre_by_evento_econ.get(evid)
            # Priorizar el "más cerrado": cerrado_economico > cerrado_plantilla > abierto.
            rank = {'abierto': 0, 'cerrado_plantilla': 1, 'cerrado_economico': 2}
            if cur is None or rank.get(ec, 0) > rank.get(cur.get('estado_cierre'), 0):
                cierre_by_evento_econ[evid] = {
                    "estado_cierre": ec,
                    "cerrado_plantilla_por": a.get('cerrado_plantilla_por'),
                    "cerrado_plantilla_at": a.get('cerrado_plantilla_at'),
                    "cerrado_economico_por": a.get('cerrado_economico_por'),
                    "cerrado_economico_at": a.get('cerrado_economico_at'),
                }
        # Resolver nombres de gestores que cerraron.
        gestor_ids_cierre = set()
        for v in cierre_by_evento_econ.values():
            for k in ('cerrado_plantilla_por', 'cerrado_economico_por'):
                if v.get(k):
                    gestor_ids_cierre.add(v[k])
        gestores_cierre_econ_by_id: Dict[str, str] = {}
        if gestor_ids_cierre:
            try:
                gres = supabase.table('usuarios').select('id,nombre,apellidos') \
                    .in_('id', list(gestor_ids_cierre)).execute().data or []
                gestores_cierre_econ_by_id = {
                    g['id']: f"{g.get('nombre','')} {g.get('apellidos','')}".strip() or g.get('id')
                    for g in gres
                }
            except Exception:
                pass
        # Iter E2 — flag de historial (4 tipos: cierre y reapertura de plantilla y económico).
        eventos_con_historial_econ: set = set()
        try:
            ra = supabase.table('registro_actividad').select('entidad_id,tipo') \
                .eq('entidad_tipo', 'evento') \
                .in_('tipo', ['evento_concluido', 'evento_reabierto', 'economico_cerrado', 'economico_reabierto']) \
                .in_('entidad_id', evento_ids_vistos).execute().data or []
            for r in ra:
                if r.get('entidad_id'):
                    eventos_con_historial_econ.add(r['entidad_id'])
        except Exception:
            pass

        ens_res = supabase.table('ensayos').select('*').in_('evento_id', evento_ids_vistos).execute()
        ensayos_by_evento = {}
        for e in (ens_res.data or []):
            ensayos_by_evento.setdefault(e['evento_id'], []).append(e)
        for evid, lst in ensayos_by_evento.items():
            lst.sort(key=lambda x: (0 if (x.get('tipo') or 'ensayo') == 'ensayo' else 1,
                                     x.get('fecha') or '', x.get('hora') or ''))

        usuarios_res = supabase.table('usuarios') \
            .select('id,nombre,apellidos,email,instrumento,especialidad,nivel_estudios,baremo,localidad,direccion,anos_experiencia,iban,swift,titulaciones') \
            .in_('id', usuario_ids).execute()
        usuarios_by_id = {u['id']: u for u in (usuarios_res.data or [])}

        ensayo_ids = [e['id'] for evs in ensayos_by_evento.values() for e in evs]
        disp_by_pair = {}
        if ensayo_ids:
            d_res = supabase.table('disponibilidad') \
                .select('id,usuario_id,ensayo_id,asiste,asistencia_real') \
                .in_('ensayo_id', ensayo_ids) \
                .in_('usuario_id', usuario_ids) \
                .execute()
            for d in (d_res.data or []):
                disp_by_pair[(d['usuario_id'], d['ensayo_id'])] = d

        gastos_res = supabase.table('gastos_adicionales').select('*') \
            .in_('evento_id', evento_ids_vistos).in_('usuario_id', usuario_ids).execute()
        gastos_by_pair = {(g['usuario_id'], g['evento_id']): g for g in (gastos_res.data or [])}

        # Convocatoria por instrumento por ensayo
        ensayo_instr_map = {}
        if ensayo_ids:
            ei_res = supabase.table('ensayo_instrumentos') \
                .select('ensayo_id,instrumento,convocado') \
                .in_('ensayo_id', ensayo_ids).execute()
            for row in (ei_res.data or []):
                ensayo_instr_map.setdefault(row['ensayo_id'], {})[row['instrumento']] = bool(row.get('convocado', True))

        cachets_res = supabase.table('cachets_config').select('*') \
            .in_('evento_id', evento_ids_vistos).execute()
        cachets_by_evento = {}
        for c in (cachets_res.data or []):
            cachets_by_evento.setdefault(c['evento_id'], []).append(c)
        cachets_base = (supabase.table('cachets_config').select('*').is_('evento_id', 'null').execute().data or [])

        total_temporada = 0.0
        eventos_out = []
        for ev in eventos_raw:
            ensayos = ensayos_by_evento.get(ev['id'], [])
            asigs_ev = [a for a in confirmadas if a['evento_id'] == ev['id']]
            secciones_map = {key: [] for key, _label in SECCIONES_ORDER}
            sin_seccion = []
            total_ev = {"cache_previsto": 0.0, "cache_real": 0.0, "extras": 0.0,
                        "transporte": 0.0, "alojamiento": 0.0, "otros": 0.0, "total": 0.0}
            for a in asigs_ev:
                u = usuarios_by_id.get(a['usuario_id'])
                if not u:
                    continue
                instr_musico = u.get('instrumento')
                asist_list = []
                for e in ensayos:
                    convocado = _is_convocado(ensayo_instr_map, e['id'], instr_musico)
                    d = disp_by_pair.get((u['id'], e['id']))
                    asist_list.append({
                        "ensayo_id": e['id'],
                        "asistencia_real": d.get('asistencia_real') if d else None,
                        "convocado": convocado,
                    })
                # Sólo ensayos convocados cuentan para %
                convocados_ensayos = [e for e in ensayos if _is_convocado(ensayo_instr_map, e['id'], instr_musico)]
                total_conv = len(convocados_ensayos)
                si_disp = sum(1 for e in convocados_ensayos if (disp_by_pair.get((u['id'], e['id'])) or {}).get('asiste') is True)
                pct_disp = round((si_disp / total_conv) * 100) if total_conv else 0
                vals = [float(x['asistencia_real']) for x in asist_list if x['convocado'] and x['asistencia_real'] is not None]
                pct_real = round(sum(vals) / len(vals), 2) if vals else 0

                # TAREA 1 — Gestión Económica lee nivel_estudios DIRECTAMENTE de usuarios
                # (sin fallback a 'especialidad' ni priorizar asignaciones.nivel_estudios para evitar valores desactualizados como "Música clásica").
                nivel_efectivo = u.get('nivel_estudios') or None
                combined = cachets_by_evento.get(ev['id'], []) + cachets_base
                lookup = _cachet_lookup_with_source(combined, u.get('instrumento'), nivel_efectivo)
                if lookup is not None:
                    cache_prev, cache_fuente = lookup
                else:
                    cache_prev = float(a.get('cache_presupuestado') or a.get('importe') or 0)
                    cache_fuente = 'asignacion' if cache_prev > 0 else 'sin_datos'
                cache_real = round(cache_prev * (pct_real / 100.0), 2)
                g = gastos_by_pair.get((u['id'], ev['id'])) or {}
                extras_visibles = float(g.get('cache_extra') or 0)
                transp_visibles = float(g.get('transporte_importe') or 0)
                # Iter F1 — provisionales NO suman al TOTAL.
                cache_extra_prov = bool(g.get('cache_extra_provisional'))
                transporte_prov = bool(g.get('transporte_provisional'))
                extras = 0.0 if cache_extra_prov else extras_visibles
                transp = 0.0 if transporte_prov else transp_visibles
                aloj = float(g.get('alojamiento_importe') or 0)
                otros = float(g.get('otros_importe') or 0)
                total = round(cache_real + extras + transp + aloj + otros, 2)

                total_ev["cache_previsto"] += cache_prev
                total_ev["cache_real"]     += cache_real
                total_ev["extras"]         += extras
                total_ev["transporte"]     += transp
                total_ev["alojamiento"]    += aloj
                total_ev["otros"]          += otros
                total_ev["total"]          += total

                musico_row = {
                    "asignacion_id": a['id'],
                    "usuario_id": u['id'],
                    "nombre": u.get('nombre') or '',
                    "apellidos": u.get('apellidos') or '',
                    "email": u.get('email') or '',
                    "instrumento": u.get('instrumento'),
                    "especialidad": u.get('especialidad'),
                    "nivel_estudios": nivel_efectivo,
                    "iban": u.get('iban'),
                    "swift": u.get('swift'),
                    "titulaciones": u.get('titulaciones') or [],
                    "numero_atril": a.get('numero_atril'),
                    "letra": a.get('letra'),
                    "estado_pago": a.get('estado_pago') or 'pendiente',
                    "asistencia": asist_list,
                    "porcentaje_disponibilidad": pct_disp,
                    "porcentaje_asistencia_real": pct_real,
                    "cache_previsto": round(cache_prev, 2),
                    "cache_fuente": cache_fuente,
                    "cache_real": cache_real,
                    "cache_extra": extras_visibles,
                    "transporte_importe": transp_visibles,
                    "alojamiento_importe": aloj,
                    "otros_importe": otros,
                    "total": total,
                    # Iter F1 — Provisional / Validación
                    "cache_extra_provisional": cache_extra_prov,
                    "transporte_provisional": transporte_prov,
                    "_extras_efectivo": extras,
                    "_transp_efectivo": transp,
                }
                sec_key = seccion_de_instrumento(u.get('instrumento'))
                if sec_key:
                    secciones_map[sec_key].append(musico_row)
                else:
                    sin_seccion.append(musico_row)

            secciones_out = []
            for key, label in SECCIONES_ORDER:
                lst = secciones_map[key]
                if not lst:
                    continue
                sec_tot = {
                    "cache_previsto": round(sum(m["cache_previsto"] for m in lst), 2),
                    "cache_real":     round(sum(m["cache_real"] for m in lst), 2),
                    "transporte":     round(sum(m["_transp_efectivo"] for m in lst), 2),
                    "alojamiento":    round(sum(m["alojamiento_importe"] for m in lst), 2),
                    "otros":          round(sum(m["otros_importe"] for m in lst), 2),
                    "total":          round(sum(m["total"] for m in lst), 2),
                    "extras":         round(sum(m["_extras_efectivo"] for m in lst), 2),
                }
                secciones_out.append({
                    "key": key, "label": label,
                    "count": len(lst),
                    "musicos": lst,
                    "totales": sec_tot,
                })
            if sin_seccion:
                sec_tot = {
                    "cache_previsto": round(sum(m["cache_previsto"] for m in sin_seccion), 2),
                    "cache_real":     round(sum(m["cache_real"] for m in sin_seccion), 2),
                    "transporte":     round(sum(m["_transp_efectivo"] for m in sin_seccion), 2),
                    "alojamiento":    round(sum(m["alojamiento_importe"] for m in sin_seccion), 2),
                    "otros":          round(sum(m["otros_importe"] for m in sin_seccion), 2),
                    "total":          round(sum(m["total"] for m in sin_seccion), 2),
                    "extras":         round(sum(m["_extras_efectivo"] for m in sin_seccion), 2),
                }
                secciones_out.append({"key":"otros","label":"Sin sección","count":len(sin_seccion),
                                      "musicos":sin_seccion,"totales":sec_tot})

            # Iter F1 — Limpiar campos auxiliares antes de serializar.
            for sec in secciones_out:
                for m in sec.get('musicos', []):
                    m.pop('_extras_efectivo', None)
                    m.pop('_transp_efectivo', None)

            cierre_info_e = cierre_by_evento_econ.get(ev['id']) or {}
            eventos_out.append({
                "id": ev['id'],
                "nombre": ev.get('nombre'),
                "fecha_inicio": ev.get('fecha_inicio'),
                "hora_inicio": ev.get('hora_inicio'),
                "estado": ev.get('estado'),
                "temporada": ev.get('temporada'),
                "ensayos": ensayos,
                "total_musicos": len(asigs_ev),
                "totales": {k: round(v, 2) for k, v in total_ev.items()},
                "secciones": secciones_out,
                # Iter E2 — cierre plantilla + económico
                "estado_cierre": cierre_info_e.get("estado_cierre") or "abierto",
                "cerrado_plantilla_at": cierre_info_e.get("cerrado_plantilla_at"),
                "cerrado_plantilla_por_nombre": gestores_cierre_econ_by_id.get(
                    cierre_info_e.get("cerrado_plantilla_por")
                ) if cierre_info_e.get("cerrado_plantilla_por") else None,
                "cerrado_economico_at": cierre_info_e.get("cerrado_economico_at"),
                "cerrado_economico_por_nombre": gestores_cierre_econ_by_id.get(
                    cierre_info_e.get("cerrado_economico_por")
                ) if cierre_info_e.get("cerrado_economico_por") else None,
                "tiene_historial_cierre": ev['id'] in eventos_con_historial_econ,
            })
            total_temporada += total_ev["total"]

        return {"eventos": eventos_out, "total_temporada": round(total_temporada, 2)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en gestión económica: {str(e)}")


@router.put("/asignaciones/{asignacion_id}/pago")
async def marcar_pago(asignacion_id: str, payload: dict, background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_gestor)):
    """Marca estado_pago de una asignación. {estado_pago: 'pagado'|'pendiente'|'anulado'}"""
    try:
        # Iter E2 — Defensa backend: no permitir cambios si el económico del evento está cerrado.
        try:
            asig_row = supabase.table('asignaciones') \
                .select('id,evento_id,estado_cierre') \
                .eq('id', asignacion_id).limit(1).execute().data or []
            if asig_row and (asig_row[0].get('estado_cierre') or 'abierto') == 'cerrado_economico':
                ev_nombre = ''
                try:
                    er = supabase.table('eventos').select('nombre') \
                        .eq('id', asig_row[0].get('evento_id')).limit(1).execute().data or []
                    ev_nombre = (er[0].get('nombre') if er else '') or ''
                except Exception:
                    pass
                raise HTTPException(
                    status_code=403,
                    detail=(
                        f"No se permiten cambios: el económico del evento "
                        f"'{ev_nombre}' está cerrado." if ev_nombre
                        else "No se permiten cambios: el económico del evento está cerrado."
                    ),
                )
        except HTTPException:
            raise
        except Exception:
            pass

        estado = payload.get('estado_pago') or 'pendiente'
        r = supabase.table('asignaciones').update({
            "estado_pago": estado,
            "updated_at": datetime.now().isoformat(),
        }).eq('id', asignacion_id).execute()
        if estado == 'pagado':
            try:
                from routes_documentos import hook_pago_marcado
                background_tasks.add_task(hook_pago_marcado, asignacion_id)
            except Exception:
                pass
        return {"ok": True, "asignacion": r.data[0] if r.data else None}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al marcar pago: {str(e)}")


@router.post("/eventos/{evento_id}/pagos-bulk")
async def marcar_pagos_bulk(evento_id: str, payload: dict, background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_gestor)):
    """TAREA 2 — Marca estado_pago en bulk para todas las asignaciones confirmadas de un evento.
    Body: {estado_pago: 'pagado' | 'pendiente' | 'anulado'}
    Devuelve: {ok: True, actualizadas: N, evento_id, estado_pago}
    """
    try:
        estado = (payload or {}).get('estado_pago') or 'pendiente'
        if estado not in ('pagado', 'pendiente', 'anulado'):
            raise HTTPException(status_code=400, detail="estado_pago inválido")

        # Iter E2 — Defensa backend: no permitir bulk si el económico del evento está cerrado.
        try:
            cerr = supabase.table('asignaciones').select('estado_cierre') \
                .eq('evento_id', evento_id).eq('estado', 'confirmado').execute().data or []
            if any((c.get('estado_cierre') or 'abierto') == 'cerrado_economico' for c in cerr):
                ev_nombre = ''
                try:
                    er = supabase.table('eventos').select('nombre') \
                        .eq('id', evento_id).limit(1).execute().data or []
                    ev_nombre = (er[0].get('nombre') if er else '') or ''
                except Exception:
                    pass
                raise HTTPException(
                    status_code=403,
                    detail=(
                        f"No se permiten cambios: el económico del evento "
                        f"'{ev_nombre}' está cerrado." if ev_nombre
                        else "No se permiten cambios: el económico del evento está cerrado."
                    ),
                )
        except HTTPException:
            raise
        except Exception:
            pass

        r = supabase.table('asignaciones').update({
            "estado_pago": estado,
            "updated_at": datetime.now().isoformat(),
        }).eq('evento_id', evento_id).eq('estado', 'confirmado').execute()
        if estado == 'pagado':
            try:
                from routes_documentos import hook_pagos_bulk
                background_tasks.add_task(hook_pagos_bulk, evento_id)
            except Exception:
                pass
        return {"ok": True, "actualizadas": len(r.data or []), "evento_id": evento_id, "estado_pago": estado}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al marcar pagos en bulk: {str(e)}")



@router.get("/gestion-economica/export")
async def export_gestion_xlsx(
    evento_id: Optional[str] = None,
    temporada: Optional[str] = None,
    current_user: dict = Depends(get_current_gestor),
):
    """Descarga .xlsx con los datos económicos. Si evento_id, solo ese evento."""
    from fastapi.responses import StreamingResponse as SR
    # Reutilizamos la consulta GET
    data = await get_gestion_economica(temporada=temporada, current_user=current_user)
    eventos = data.get("eventos", [])
    if evento_id:
        eventos = [e for e in eventos if e['id'] == evento_id]
    wb = Workbook()
    ws = wb.active
    ws.title = "Economía"
    headers = ["Evento","Sección","Apellidos, Nombre","Instrumento","Especialidad","Nivel",
               "IBAN","SWIFT","% Disp.","% Asist. Real","Caché Prev.","Caché Real","Extras",
               "Transporte","Alojamiento","Otros","TOTAL","Estado Pago"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for ev in eventos:
        for sec in ev.get("secciones", []):
            for m in sec.get("musicos", []):
                ws.append([
                    ev.get("nombre",""), sec.get("label",""),
                    f"{m.get('apellidos','')}, {m.get('nombre','')}",
                    m.get("instrumento") or "",
                    m.get("especialidad") or "",
                    m.get("nivel_estudios") or "",
                    m.get("iban") or "",
                    m.get("swift") or "",
                    m.get("porcentaje_disponibilidad"),
                    m.get("porcentaje_asistencia_real"),
                    m.get("cache_previsto"),
                    m.get("cache_real"),
                    m.get("cache_extra"),
                    m.get("transporte_importe"),
                    m.get("alojamiento_importe"),
                    m.get("otros_importe"),
                    m.get("total"),
                    m.get("estado_pago"),
                ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"gestion_economica_{evento_id or temporada or 'todos'}.xlsx"
    return SR(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
              headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ==================================================================
# BLOQUE 6 — Análisis económico + export SEPA XML pain.001
# ==================================================================

@router.get("/analisis/resumen")
async def get_analisis_resumen(
    temporada: Optional[str] = None,
    current_user: dict = Depends(get_current_gestor),
):
    """Devuelve métricas agregadas de la temporada para la página de análisis."""
    try:
        data = await get_gestion_economica(temporada=temporada, current_user=current_user)
        eventos = data.get("eventos", [])
        total_prev = sum(ev["totales"]["cache_previsto"] for ev in eventos)
        total_real = sum(ev["totales"]["total"] for ev in eventos)
        total_musicos = 0
        asistencias_pct = []
        por_evento = []
        por_seccion = {}
        for ev in eventos:
            musicos = sum(sec["count"] for sec in ev["secciones"])
            total_musicos += musicos
            pct_medio = 0
            n = 0
            for sec in ev["secciones"]:
                for m in sec["musicos"]:
                    pct_medio += m["porcentaje_asistencia_real"]
                    n += 1
                por_seccion[sec["label"]] = por_seccion.get(sec["label"], 0) + sec["totales"]["total"]
            pct_medio = round(pct_medio / n, 2) if n else 0
            asistencias_pct.append(pct_medio)
            por_evento.append({
                "id": ev["id"],
                "nombre": ev["nombre"],
                "fecha_inicio": ev.get("fecha_inicio"),
                "musicos": musicos,
                "pct_asistencia_medio": pct_medio,
                "cache_previsto": ev["totales"]["cache_previsto"],
                "total": ev["totales"]["total"],
            })
        pct_asistencia_temporada = round(sum(asistencias_pct) / len(asistencias_pct), 2) if asistencias_pct else 0

        # Total músicos convocados (asignaciones en eventos de la temporada)
        convocados_q = supabase.table('asignaciones').select('usuario_id, evento_id')
        if temporada:
            ev_ids = [e['id'] for e in eventos]
            if ev_ids:
                convocados_q = convocados_q.in_('evento_id', ev_ids)
        convocados_res = convocados_q.execute()
        total_convocados = len({a['usuario_id'] for a in (convocados_res.data or [])})

        return {
            "total_eventos": len(eventos),
            "total_musicos_convocados": total_convocados,
            "total_musicos_confirmados": total_musicos,
            "pct_asistencia_medio": pct_asistencia_temporada,
            "coste_previsto": round(total_prev, 2),
            "coste_real": round(total_real, 2),
            "diferencia": round(total_real - total_prev, 2),
            "por_evento": por_evento,
            "por_seccion": [{"seccion": k, "importe": round(v, 2)} for k, v in por_seccion.items()],
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en análisis: {str(e)}")


@router.get("/analisis/sepa-xml")
async def export_sepa_xml(
    temporada: Optional[str] = None,
    evento_id: Optional[str] = None,
    current_user: dict = Depends(get_current_gestor),
):
    """Genera XML SEPA pain.001.001.03 con las liquidaciones por músico.
    Si se pasa `evento_id`, sólo incluye músicos de ese evento."""
    from fastapi.responses import Response
    data = await get_gestion_economica(temporada=temporada, current_user=current_user)
    eventos = data.get("eventos", [])
    if evento_id:
        eventos = [e for e in eventos if e['id'] == evento_id]
    # Agrupar por músico: acumular el total a percibir
    por_musico = {}
    for ev in eventos:
        for sec in ev.get("secciones", []):
            for m in sec.get("musicos", []):
                uid = m["usuario_id"]
                if uid not in por_musico:
                    por_musico[uid] = {
                        "nombre": f"{m['nombre']} {m['apellidos']}".strip(),
                        "iban": m.get("iban") or "",
                        "swift": m.get("swift") or "",
                        "total": 0.0,
                    }
                por_musico[uid]["total"] += m.get("total", 0)
    now = datetime.now()
    msg_id = f"OPUS-{now.strftime('%Y%m%d%H%M%S')}"
    total_general = round(sum(p["total"] for p in por_musico.values()), 2)
    tx_count = sum(1 for p in por_musico.values() if p["total"] > 0 and p["iban"])

    def esc(s):
        return (s or "").replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    txs = []
    idx = 0
    for p in por_musico.values():
        if not p["iban"] or p["total"] <= 0:
            continue
        idx += 1
        txs.append(f"""
    <CdtTrfTxInf>
      <PmtId><EndToEndId>OPUS-{idx:04d}</EndToEndId></PmtId>
      <Amt><InstdAmt Ccy="EUR">{p["total"]:.2f}</InstdAmt></Amt>
      {f'<CdtrAgt><FinInstnId><BIC>{esc(p["swift"])}</BIC></FinInstnId></CdtrAgt>' if p["swift"] else ''}
      <Cdtr><Nm>{esc(p["nombre"])}</Nm></Cdtr>
      <CdtrAcct><Id><IBAN>{esc(p["iban"].replace(' ',''))}</IBAN></Id></CdtrAcct>
      <RmtInf><Ustrd>Liquidación Opus Manager {esc(temporada or '')}</Ustrd></RmtInf>
    </CdtTrfTxInf>""")

    xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<Document xmlns="urn:iso:std:iso:20022:tech:xsd:pain.001.001.03">
  <CstmrCdtTrfInitn>
    <GrpHdr>
      <MsgId>{msg_id}</MsgId>
      <CreDtTm>{now.strftime('%Y-%m-%dT%H:%M:%S')}</CreDtTm>
      <NbOfTxs>{tx_count}</NbOfTxs>
      <CtrlSum>{total_general:.2f}</CtrlSum>
      <InitgPty><Nm>Opus Manager</Nm></InitgPty>
    </GrpHdr>
    <PmtInf>
      <PmtInfId>{msg_id}-PMT</PmtInfId>
      <PmtMtd>TRF</PmtMtd>
      <NbOfTxs>{tx_count}</NbOfTxs>
      <CtrlSum>{total_general:.2f}</CtrlSum>
      <PmtTpInf><SvcLvl><Cd>SEPA</Cd></SvcLvl></PmtTpInf>
      <ReqdExctnDt>{now.strftime('%Y-%m-%d')}</ReqdExctnDt>
      <Dbtr><Nm>Opus Manager</Nm></Dbtr>
      <DbtrAcct><Id><IBAN>ES0000000000000000000000</IBAN></Id></DbtrAcct>
      <DbtrAgt><FinInstnId><Othr><Id>NOTPROVIDED</Id></Othr></FinInstnId></DbtrAgt>
      {''.join(txs)}
    </PmtInf>
  </CstmrCdtTrfInitn>
</Document>"""
    return Response(
        content=xml,
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="sepa_liquidaciones_{temporada or "todas"}.xml"'}
    )


# ==================================================================
# ==================================================================
# BLOQUE 7 — Tareas (CRUD)
# NOTE: Movido a routes_tareas.py durante el refactor de feb 2026.
# ==================================================================


@router.get("/gestores")
async def list_gestores(current_user: dict = Depends(get_current_gestor)):
    try:
        r = supabase.table('usuarios').select('id,nombre,apellidos,email').eq('rol','gestor').order('apellidos').execute()
        return {"gestores": r.data or []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al listar gestores: {str(e)}")


# NOTE: Endpoints de /incidencias movidos a routes_incidencias.py (refactor feb 2026)


# ==================================================================
# LOGÍSTICA del evento (transportes + alojamientos)
# ==================================================================

class LogisticaItem(BaseModel):
    id: Optional[str] = None
    tipo: Literal['transporte_ida', 'transporte_vuelta', 'alojamiento']
    orden: Optional[int] = 1
    fecha: Optional[str] = None
    hora_salida: Optional[str] = None
    lugar_salida: Optional[str] = None
    hora_llegada: Optional[str] = None
    lugar_llegada: Optional[str] = None
    recogida_1_lugar: Optional[str] = None
    recogida_1_hora: Optional[str] = None
    recogida_2_lugar: Optional[str] = None
    recogida_2_hora: Optional[str] = None
    recogida_3_lugar: Optional[str] = None
    recogida_3_hora: Optional[str] = None
    hotel_nombre: Optional[str] = None
    hotel_direccion: Optional[str] = None
    fecha_checkin: Optional[str] = None
    fecha_checkout: Optional[str] = None
    fecha_limite_confirmacion: Optional[str] = None
    notas: Optional[str] = None


class LogisticaBulkRequest(BaseModel):
    items: List[LogisticaItem] = []
    eliminar_ids: List[str] = []


@router.get("/eventos/{evento_id}/logistica")
async def get_logistica_evento(evento_id: str, current_user: dict = Depends(get_current_gestor)):
    try:
        r = supabase.table('evento_logistica').select('*') \
            .eq('evento_id', evento_id) \
            .order('tipo', desc=False) \
            .order('orden', desc=False) \
            .execute()
        return {"logistica": r.data or []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al listar logística: {str(e)}")


@router.put("/eventos/{evento_id}/logistica")
async def put_logistica_evento(evento_id: str, data: LogisticaBulkRequest, current_user: dict = Depends(get_current_gestor)):
    """Guardado masivo de la logística de un evento."""
    try:
        now = datetime.now().isoformat()
        creados = 0
        actualizados = 0
        for it in data.items:
            base = it.model_dump(exclude_none=True)
            base.pop('id', None)
            base['evento_id'] = evento_id
            base['updated_at'] = now
            # Limpiar strings vacíos para columnas DATE/TIME (Postgres lanza error si recibe '')
            for k in ('fecha', 'fecha_checkin', 'fecha_checkout', 'fecha_limite_confirmacion',
                      'hora_salida', 'hora_llegada',
                      'recogida_1_hora', 'recogida_2_hora', 'recogida_3_hora'):
                if base.get(k) == '':
                    base[k] = None
            if it.id:
                supabase.table('evento_logistica').update(base).eq('id', it.id).execute()
                actualizados += 1
            else:
                supabase.table('evento_logistica').insert(base).execute()
                creados += 1
        borrados = 0
        for lid in (data.eliminar_ids or []):
            supabase.table('evento_logistica').delete().eq('id', lid).execute()
            borrados += 1
        return {"ok": True, "creados": creados, "actualizados": actualizados, "borrados": borrados}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al guardar logística: {str(e)}")


@router.delete("/logistica/{logistica_id}")
async def delete_logistica(logistica_id: str, current_user: dict = Depends(get_current_gestor)):
    try:
        supabase.table('evento_logistica').delete().eq('id', logistica_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al eliminar logística: {str(e)}")


@router.get("/logistica/{logistica_id}/confirmaciones")
async def get_confirmaciones_logistica(logistica_id: str, current_user: dict = Depends(get_current_gestor)):
    """Para una pieza de logística, devuelve la lista de músicos confirmados,
    rechazados y los que aún no han respondido (entre los músicos asignados al evento)."""
    try:
        log_row = supabase.table('evento_logistica').select('evento_id') \
            .eq('id', logistica_id).limit(1).execute().data
        if not log_row:
            raise HTTPException(status_code=404, detail="Logística no encontrada")
        evento_id = log_row[0]['evento_id']

        # Músicos asignados al evento
        asigs = supabase.table('asignaciones').select('usuario_id') \
            .eq('evento_id', evento_id).execute().data or []
        usuario_ids = list({a['usuario_id'] for a in asigs})

        # Confirmaciones existentes
        confs = supabase.table('confirmaciones_logistica').select('*') \
            .eq('logistica_id', logistica_id).execute().data or []
        conf_by_user = {c['usuario_id']: c for c in confs}

        # Datos de músicos
        users = supabase.table('usuarios') \
            .select('id,nombre,apellidos,instrumento,email') \
            .in_('id', usuario_ids).execute().data if usuario_ids else []

        confirmados = []; rechazados = []; sin_respuesta = []
        for u in users:
            c = conf_by_user.get(u['id'])
            entry = {**u, "respuesta_at": (c or {}).get('updated_at')}
            if not c:
                sin_respuesta.append(entry)
            elif c.get('confirmado') is True:
                confirmados.append(entry)
            elif c.get('confirmado') is False:
                rechazados.append(entry)
            else:
                sin_respuesta.append(entry)
        return {
            "logistica_id": logistica_id,
            "total_asignados": len(usuario_ids),
            "confirmados": confirmados,
            "rechazados": rechazados,
            "sin_respuesta": sin_respuesta,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@router.get("/logistica")
async def listar_logistica_global(current_user: dict = Depends(get_current_gestor)):
    """Vista global de logística para `/asistencia/logistica`.
    Devuelve eventos con logística (estado=abierto) + músicos confirmados
    en cada pieza de logística para mostrar en acordeón.
    """
    try:
        # 1. Eventos abiertos con logística asociada
        log_rows = supabase.table('evento_logistica') \
            .select('*, evento:eventos(id,nombre,fecha_inicio,fecha_fin,estado)') \
            .order('evento_id').execute().data or []

        # Filtrar solo eventos con estado=abierto (o sin estado=cerrado)
        log_rows = [
            l for l in log_rows
            if (l.get('evento') or {}).get('estado') != 'cerrado'
        ]
        if not log_rows:
            return {"eventos": []}

        # Agrupar por evento_id
        eventos_map: Dict[str, dict] = {}
        for l in log_rows:
            ev = l.pop('evento', None) or {}
            eid = l.get('evento_id')
            if not eid:
                continue
            if eid not in eventos_map:
                eventos_map[eid] = {
                    "evento_id": eid,
                    "nombre": ev.get('nombre'),
                    "fecha_inicio": ev.get('fecha_inicio'),
                    "fecha_fin": ev.get('fecha_fin'),
                    "items_logistica": [],
                }
            eventos_map[eid]["items_logistica"].append(l)

        # 2. Cargar todas las confirmaciones de cada pieza
        all_log_ids = [l['id'] for l in log_rows]
        confs_all: List[dict] = []
        for i in range(0, len(all_log_ids), 200):
            chunk = all_log_ids[i:i + 200]
            res = supabase.table('confirmaciones_logistica').select('*') \
                .in_('logistica_id', chunk).execute().data or []
            confs_all.extend(res)
        confs_by_log: Dict[str, List[dict]] = {}
        for c in confs_all:
            confs_by_log.setdefault(c['logistica_id'], []).append(c)

        # 3. Asignaciones (músicos confirmados por evento)
        ev_ids = list(eventos_map.keys())
        asigs_all: List[dict] = []
        for i in range(0, len(ev_ids), 200):
            chunk = ev_ids[i:i + 200]
            res = supabase.table('asignaciones') \
                .select('id,evento_id,usuario_id,estado') \
                .in_('evento_id', chunk).eq('estado', 'confirmado').execute().data or []
            asigs_all.extend(res)
        asigs_by_ev: Dict[str, List[dict]] = {}
        all_user_ids: set = set()
        for a in asigs_all:
            asigs_by_ev.setdefault(a['evento_id'], []).append(a)
            all_user_ids.add(a['usuario_id'])

        # 4. Datos de usuarios
        users_map: Dict[str, dict] = {}
        if all_user_ids:
            uid_list = list(all_user_ids)
            for i in range(0, len(uid_list), 200):
                chunk = uid_list[i:i + 200]
                res = supabase.table('usuarios') \
                    .select('id,nombre,apellidos,instrumento,email') \
                    .in_('id', chunk).execute().data or []
                for u in res:
                    users_map[u['id']] = u

        # 5. Construir respuesta agregada
        out: List[dict] = []
        for eid, ev in eventos_map.items():
            asigs = asigs_by_ev.get(eid, [])
            # Para cada músico, calcular estado de cada pieza
            confs_by_user_log: Dict[str, dict] = {}
            for it in ev["items_logistica"]:
                for c in confs_by_log.get(it['id'], []):
                    confs_by_user_log[f"{c['usuario_id']}|{it['id']}"] = c

            musicos = []
            tot_ida_ok = tot_vuelta_ok = tot_aloja_ok = 0
            for a in asigs:
                u = users_map.get(a['usuario_id'])
                if not u:
                    continue
                fila = {
                    "usuario_id": u['id'],
                    "nombre": u.get('nombre'),
                    "apellidos": u.get('apellidos'),
                    "instrumento": u.get('instrumento'),
                    "email": u.get('email'),
                    "ida": None,
                    "vuelta": None,
                    "alojamiento": None,
                    "punto_recogida": None,
                    "fecha_confirmacion": None,
                }
                for it in ev["items_logistica"]:
                    c = confs_by_user_log.get(f"{u['id']}|{it['id']}")
                    confirmado = c.get('confirmado') if c else None
                    estado = '✅' if confirmado is True else ('❌' if confirmado is False else '⏳')
                    if it['tipo'] == 'transporte_ida':
                        fila['ida'] = estado
                        if confirmado is True: tot_ida_ok += 1
                    elif it['tipo'] == 'transporte_vuelta':
                        fila['vuelta'] = estado
                        if confirmado is True: tot_vuelta_ok += 1
                    elif it['tipo'] == 'alojamiento':
                        fila['alojamiento'] = estado
                        if confirmado is True: tot_aloja_ok += 1
                    if c and c.get('updated_at') and not fila.get('fecha_confirmacion'):
                        fila['fecha_confirmacion'] = c['updated_at']
                    if c and c.get('punto_recogida') and not fila.get('punto_recogida'):
                        fila['punto_recogida'] = c.get('punto_recogida')
                musicos.append(fila)

            # Fecha límite mínima entre todos los items
            limites = [it.get('fecha_limite_confirmacion') for it in ev["items_logistica"] if it.get('fecha_limite_confirmacion')]
            ev["fecha_limite_min"] = min(limites) if limites else None
            ev["totales"] = {
                "asignados_confirmados": len(asigs),
                "ida_confirmada": tot_ida_ok,
                "vuelta_confirmada": tot_vuelta_ok,
                "alojamiento_confirmado": tot_aloja_ok,
            }
            ev["musicos"] = musicos
            out.append(ev)

        out.sort(key=lambda e: (e.get('fecha_inicio') or '9999-99-99'))
        return {"eventos": out}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en logística global: {str(e)}")


# ==================================================================
# COMIDAS / Servicio de comedor del evento (Iter 19)
# ==================================================================

class ComidaItem(BaseModel):
    id: Optional[str] = None
    orden: Optional[int] = 1
    fecha: Optional[str] = None
    hora_inicio: Optional[str] = None
    hora_fin: Optional[str] = None
    lugar: Optional[str] = None
    menu: Optional[str] = None
    precio_menu: Optional[float] = 0
    incluye_cafe: Optional[bool] = False
    precio_cafe: Optional[float] = 0
    fecha_limite_confirmacion: Optional[str] = None
    notas: Optional[str] = None
    opciones_menu: Optional[List[Dict[str, Any]]] = None  # [{id, nombre}]


class ComidasBulkRequest(BaseModel):
    items: List[ComidaItem] = []
    eliminar_ids: List[str] = []


@router.get("/eventos/{evento_id}/comidas")
async def get_comidas_evento(evento_id: str, current_user: dict = Depends(get_current_gestor)):
    try:
        r = supabase.table('evento_comidas').select('*') \
            .eq('evento_id', evento_id) \
            .order('fecha', desc=False).order('orden', desc=False).execute()
        return {"comidas": r.data or []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al listar comidas: {str(e)}")


@router.put("/eventos/{evento_id}/comidas")
async def put_comidas_evento(evento_id: str, data: ComidasBulkRequest, current_user: dict = Depends(get_current_gestor)):
    """Guardado masivo del servicio de comedor de un evento."""
    try:
        now = datetime.now().isoformat()
        creados = 0
        actualizados = 0
        for it in data.items:
            base = it.model_dump(exclude_none=True)
            base.pop('id', None)
            base['evento_id'] = evento_id
            base['updated_at'] = now
            for k in ('fecha', 'hora_inicio', 'hora_fin', 'fecha_limite_confirmacion'):
                if base.get(k) == '':
                    base[k] = None
            if it.id:
                supabase.table('evento_comidas').update(base).eq('id', it.id).execute()
                actualizados += 1
            else:
                supabase.table('evento_comidas').insert(base).execute()
                creados += 1
        borrados = 0
        for cid in (data.eliminar_ids or []):
            supabase.table('evento_comidas').delete().eq('id', cid).execute()
            borrados += 1
        return {"ok": True, "creados": creados, "actualizados": actualizados, "borrados": borrados}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al guardar comidas: {str(e)}")


@router.delete("/comidas/{comida_id}")
async def delete_comida(comida_id: str, current_user: dict = Depends(get_current_gestor)):
    try:
        supabase.table('evento_comidas').delete().eq('id', comida_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al eliminar comida: {str(e)}")


@router.get("/comidas/{comida_id}/confirmaciones")
async def get_confirmaciones_comida(comida_id: str, current_user: dict = Depends(get_current_gestor)):
    """Para un servicio de comedor, devuelve músicos confirmados, rechazados y sin respuesta."""
    try:
        cr = supabase.table('evento_comidas').select('evento_id,incluye_cafe,precio_menu,precio_cafe,opciones_menu') \
            .eq('id', comida_id).limit(1).execute().data
        if not cr:
            raise HTTPException(status_code=404, detail="Comida no encontrada")
        com = cr[0]
        evento_id = com['evento_id']

        asigs = supabase.table('asignaciones').select('usuario_id') \
            .eq('evento_id', evento_id).execute().data or []
        usuario_ids = list({a['usuario_id'] for a in asigs})

        confs = supabase.table('confirmaciones_comida').select('*') \
            .eq('comida_id', comida_id).execute().data or []
        conf_by_user = {c['usuario_id']: c for c in confs}

        users = supabase.table('usuarios') \
            .select('id,nombre,apellidos,instrumento,email') \
            .in_('id', usuario_ids).execute().data if usuario_ids else []

        confirmados = []; rechazados = []; sin_respuesta = []
        total_recaudado = 0.0
        desglose_opciones: Dict[str, int] = {}  # id opcion -> nº confirmados
        for u in users:
            c = conf_by_user.get(u['id'])
            entry = {
                **u,
                "respuesta_at": (c or {}).get('updated_at'),
                "toma_cafe": (c or {}).get('toma_cafe'),
                "opcion_menu": (c or {}).get('opcion_menu_seleccionada'),
                "notas": (c or {}).get('notas'),
            }
            if c is None or c.get('confirmado') is None:
                sin_respuesta.append(entry)
            elif c.get('confirmado') is True:
                confirmados.append(entry)
                total_recaudado += float(com.get('precio_menu') or 0)
                if com.get('incluye_cafe') and c.get('toma_cafe'):
                    total_recaudado += float(com.get('precio_cafe') or 0)
                op = c.get('opcion_menu_seleccionada')
                if op:
                    desglose_opciones[op] = desglose_opciones.get(op, 0) + 1
            else:
                rechazados.append(entry)
        # Resolver nombres de las opciones
        opciones_defs = {o.get('id'): o.get('nombre', o.get('id')) for o in (com.get('opciones_menu') or []) if isinstance(o, dict)}
        desglose_por_opcion = [
            {"id": oid, "nombre": opciones_defs.get(oid, oid), "cantidad": n}
            for oid, n in sorted(desglose_opciones.items(), key=lambda x: -x[1])
        ]
        return {
            "confirmados": confirmados,
            "rechazados": rechazados,
            "sin_respuesta": sin_respuesta,
            "total_recaudado": round(total_recaudado, 2),
            "desglose_por_opcion": desglose_por_opcion,
            "opciones_menu": com.get('opciones_menu') or [],
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener confirmaciones de comida: {str(e)}")


@router.get("/comidas")
async def get_comidas_globales(current_user: dict = Depends(get_current_gestor)):
    """Vista global de comidas (consumida por LogisticaPage o similares)."""
    try:
        com_rows = supabase.table('evento_comidas') \
            .select('*, evento:eventos(id,nombre,fecha_inicio,temporada,estado,lugar)') \
            .order('fecha', desc=False).execute().data or []
        # Agrupar por evento
        eventos: Dict[str, Any] = {}
        for c in com_rows:
            ev = c.pop('evento', None) or {}
            eid = ev.get('id') or c.get('evento_id')
            if not eid:
                continue
            if eid not in eventos:
                eventos[eid] = {**ev, "comidas": [], "evento_id": eid}
            eventos[eid]["comidas"].append(c)
        # Adjuntar confirmaciones por comida
        com_ids = [c['id'] for ev in eventos.values() for c in ev["comidas"]]
        confs_by_comida: Dict[str, List[Dict]] = {}
        if com_ids:
            confs = supabase.table('confirmaciones_comida').select('*') \
                .in_('comida_id', com_ids).execute().data or []
            for cf in confs:
                confs_by_comida.setdefault(cf['comida_id'], []).append(cf)
        out = []
        for ev in eventos.values():
            tot_ok = 0
            tot_no = 0
            for c in ev["comidas"]:
                cs = confs_by_comida.get(c['id'], [])
                c["confirmados_n"] = sum(1 for x in cs if x.get('confirmado') is True)
                c["rechazados_n"] = sum(1 for x in cs if x.get('confirmado') is False)
                tot_ok += c["confirmados_n"]
                tot_no += c["rechazados_n"]
            ev["totales"] = {"confirmados": tot_ok, "rechazados": tot_no, "n_servicios": len(ev["comidas"])}
            out.append(ev)
        out.sort(key=lambda e: (e.get('fecha_inicio') or '9999-99-99'))
        return {"eventos": out}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en comidas globales: {str(e)}")
