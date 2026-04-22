"""Tests for Bloque C: musicos DB, import Excel/CSV, seguimiento, DELETE musico."""
import io
import os
import uuid
import csv
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://contact-conductor.preview.emergentagent.com').rstrip('/')

GESTOR_EMAIL = "admin@convocatorias.com"
GESTOR_PASSWORD = "Admin123!"

JESUS_MUSICO_ID = "8bf521fa-dc27-4c5b-8069-d36d3d4eaad3"
NONEXISTENT_ID = "00000000-0000-0000-0000-000000000000"


@pytest.fixture(scope="session")
def gestor_token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"email": GESTOR_EMAIL, "password": GESTOR_PASSWORD}, timeout=20)
    if r.status_code != 200:
        pytest.skip(f"Login gestor falló: {r.status_code} {r.text[:200]}")
    data = r.json()
    token = data.get("access_token") or data.get("token") or (data.get("session") or {}).get("access_token")
    assert token, f"No token in login response: {data}"
    return token


@pytest.fixture(scope="session")
def auth_headers(gestor_token):
    return {"Authorization": f"Bearer {gestor_token}"}


# ============== C-2.1 Plantilla ==============
def test_plantilla_descarga_xlsx(auth_headers):
    r = requests.get(f"{BASE_URL}/api/gestor/musicos-import/plantilla", headers=auth_headers, timeout=30)
    assert r.status_code == 200, r.text[:300]
    ct = r.headers.get("content-type", "")
    assert "spreadsheetml" in ct or "xlsx" in ct, f"unexpected content-type: {ct}"
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    expected = ["nombre","apellidos","email","telefono","instrumento","especialidad","dni","direccion","fecha_nacimiento","nacionalidad","bio"]
    assert headers == expected, f"headers mismatch: {headers}"


# ============== C-2.2 Preview ==============
def _build_csv(rows):
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["nombre","apellidos","email","telefono","instrumento"])
    writer.writeheader()
    for r in rows:
        writer.writerow(r)
    return buf.getvalue().encode("utf-8")


def test_preview_csv_3_filas(auth_headers):
    rows = [
        {"nombre":"TestA","apellidos":"Prev","email":f"test_prev_a_{uuid.uuid4().hex[:6]}@example.com","telefono":"600","instrumento":"Violín"},
        {"nombre":"TestB","apellidos":"Prev","email":f"test_prev_b_{uuid.uuid4().hex[:6]}@example.com","telefono":"601","instrumento":"Viola"},
        {"nombre":"TestC","apellidos":"Prev","email":f"test_prev_c_{uuid.uuid4().hex[:6]}@example.com","telefono":"602","instrumento":"Chelo"},
    ]
    csv_bytes = _build_csv(rows)
    files = {"archivo": ("preview.csv", csv_bytes, "text/csv")}
    r = requests.post(f"{BASE_URL}/api/gestor/musicos-import/preview", headers=auth_headers, files=files, timeout=30)
    assert r.status_code == 200, r.text[:300]
    j = r.json()
    assert j["total_filas"] == 3
    assert isinstance(j["preview"], list) and len(j["preview"]) <= 5
    assert j["missing_required_headers"] == []


# ============== C-2.3 Import + cleanup ==============
@pytest.fixture(scope="module")
def created_test_emails():
    emails = []
    yield emails
    # Teardown: best-effort cleanup
    from supabase import create_client
    try:
        admin = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_KEY'])
        for em in emails:
            try:
                u = admin.table('usuarios').select('id,user_id').eq('email', em).limit(1).execute()
                if u.data:
                    uid = u.data[0]['user_id']
                    admin.table('usuarios').delete().eq('email', em).execute()
                    if uid:
                        try:
                            admin.auth.admin.delete_user(uid)
                        except Exception:
                            pass
            except Exception:
                pass
    except Exception:
        pass


def test_importar_csv_crea_usuarios(auth_headers, created_test_emails):
    e1 = f"test_import_a_{uuid.uuid4().hex[:8]}@example.com"
    e2 = f"test_import_b_{uuid.uuid4().hex[:8]}@example.com"
    rows = [
        {"nombre":"TestImp","apellidos":"Uno","email":e1,"telefono":"600111","instrumento":"Violín"},
        {"nombre":"TestImp","apellidos":"Dos","email":e2,"telefono":"600222","instrumento":"Viola"},
    ]
    created_test_emails.extend([e1, e2])
    csv_bytes = _build_csv(rows)
    files = {"archivo": ("import.csv", csv_bytes, "text/csv")}
    r = requests.post(f"{BASE_URL}/api/gestor/musicos-import", headers=auth_headers, files=files, timeout=60)
    assert r.status_code == 200, r.text[:500]
    j = r.json()
    assert "resumen" in j
    assert j["resumen"]["creados"] >= 1, f"resumen={j['resumen']} errors={j.get('errores')}"


def test_importar_email_repetido_marca_existente(auth_headers, created_test_emails):
    if not created_test_emails:
        pytest.skip("No previous emails imported")
    em = created_test_emails[0]
    rows = [{"nombre":"Repeat","apellidos":"X","email":em,"telefono":"","instrumento":"Violín"}]
    csv_bytes = _build_csv(rows)
    files = {"archivo": ("rep.csv", csv_bytes, "text/csv")}
    r = requests.post(f"{BASE_URL}/api/gestor/musicos-import", headers=auth_headers, files=files, timeout=30)
    assert r.status_code == 200, r.text[:300]
    j = r.json()
    assert j["resumen"]["ya_existentes"] >= 1


# ============== C-3.1 Seguimiento ==============
def test_seguimiento_estructura(auth_headers):
    r = requests.get(f"{BASE_URL}/api/gestor/seguimiento", headers=auth_headers, timeout=30)
    assert r.status_code == 200, r.text[:300]
    j = r.json()
    assert "eventos" in j and "musicos" in j and "asignaciones" in j
    assert isinstance(j["eventos"], list)
    if j["eventos"]:
        ev = j["eventos"][0]
        assert "funciones" in ev
        assert isinstance(ev["funciones"], list)
        assert len(ev["funciones"]) >= 1
        assert ev["funciones"][0].get("fecha")


# ============== C-3.2 Bulk acciones ==============
def test_bulk_5_estados(auth_headers):
    r = requests.get(f"{BASE_URL}/api/gestor/seguimiento", headers=auth_headers, timeout=30)
    assert r.status_code == 200
    j = r.json()
    if not j["eventos"] or not j["musicos"]:
        pytest.skip("No hay eventos o músicos")
    evento_id = j["eventos"][0]["id"]
    musico_ids = [m["id"] for m in j["musicos"] if m["id"] != JESUS_MUSICO_ID][:1]
    if not musico_ids:
        pytest.skip("No musicos sin Jesús")
    # Save original
    orig_key = f"{musico_ids[0]}_{evento_id}"
    orig = j["asignaciones"].get(orig_key)
    orig_estado = orig["estado"] if orig else None

    estados = ["pendiente", "confirmado", "rechazado", "no_disponible", "excluido"]
    for est in estados:
        r2 = requests.post(f"{BASE_URL}/api/gestor/seguimiento/bulk", headers=auth_headers,
                           json={"evento_id": evento_id, "usuario_ids": musico_ids, "estado": est}, timeout=30)
        assert r2.status_code == 200, f"estado={est} -> {r2.status_code} {r2.text[:200]}"
        jj = r2.json()
        assert "actualizados" in jj and "creados" in jj
        assert (jj["actualizados"] + jj["creados"]) >= 1

    # Restore
    if orig_estado:
        requests.post(f"{BASE_URL}/api/gestor/seguimiento/bulk", headers=auth_headers,
                      json={"evento_id": evento_id, "usuario_ids": musico_ids, "estado": orig_estado}, timeout=30)


# ============== DELETE músico ==============
def test_delete_musico_ok(auth_headers):
    # Create a fresh musico via crear_musico endpoint
    em = f"test_del_{uuid.uuid4().hex[:8]}@example.com"
    payload = {"nombre": "Delete", "apellidos": "Test", "email": em, "instrumento": "Violín"}
    r = requests.post(f"{BASE_URL}/api/gestor/musicos/crear", headers=auth_headers, json=payload, timeout=30)
    assert r.status_code == 200, r.text[:500]
    j = r.json()
    new_id = (j.get("musico") or {}).get("id") or j.get("id")
    assert new_id, f"No id en respuesta crear: {j}"

    # Now delete
    rd = requests.delete(f"{BASE_URL}/api/gestor/musicos/{new_id}", headers=auth_headers, timeout=30)
    assert rd.status_code == 200, rd.text[:500]
    jd = rd.json()
    assert "message" in jd
    assert "auth_deleted" in jd
    assert "auth_error" in jd

    # Verify GET musicos no longer contains
    rg = requests.get(f"{BASE_URL}/api/gestor/musicos", headers=auth_headers, timeout=30)
    assert rg.status_code == 200
    items = rg.json() if isinstance(rg.json(), list) else rg.json().get("musicos", [])
    assert all(m.get("id") != new_id for m in items)


def test_delete_musico_409_jesus(auth_headers):
    r = requests.delete(f"{BASE_URL}/api/gestor/musicos/{JESUS_MUSICO_ID}", headers=auth_headers, timeout=30)
    assert r.status_code == 409, f"got {r.status_code}: {r.text[:300]}"
    j = r.json()
    assert "confirmadas" in (j.get("detail") or "").lower()


def test_delete_musico_404(auth_headers):
    r = requests.delete(f"{BASE_URL}/api/gestor/musicos/{NONEXISTENT_ID}", headers=auth_headers, timeout=30)
    assert r.status_code == 404, f"got {r.status_code}: {r.text[:300]}"


# ============== Registro de actividad ==============
def test_registro_actividad_musico_eliminado(auth_headers):
    # Create + delete to ensure fresh entry
    em = f"test_act_{uuid.uuid4().hex[:8]}@example.com"
    rc = requests.post(f"{BASE_URL}/api/gestor/musicos/crear", headers=auth_headers,
                       json={"nombre":"Act","apellidos":"Test","email":em,"instrumento":"Violín"}, timeout=30)
    assert rc.status_code == 200
    new_id = (rc.json().get("musico") or {}).get("id") or rc.json().get("id")
    rd = requests.delete(f"{BASE_URL}/api/gestor/musicos/{new_id}", headers=auth_headers, timeout=30)
    assert rd.status_code == 200

    # Query actividad
    ra = requests.get(f"{BASE_URL}/api/gestor/actividad", headers=auth_headers, timeout=30)
    assert ra.status_code == 200, ra.text[:300]
    items = ra.json() if isinstance(ra.json(), list) else ra.json().get("items") or ra.json().get("actividad") or []
    found = any(it.get("tipo") == "musico_eliminado" and it.get("entidad_id") == new_id for it in items)
    assert found, f"No se encontró registro_actividad para {new_id}"
