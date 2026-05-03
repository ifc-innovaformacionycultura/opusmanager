"""
Router del módulo Archivo Musical.
Endpoints: catálogo de obras, partes, originales, préstamos, programas de eventos,
verificación de atriles, importación masiva y generación de etiquetas PDF.
"""
import io
import re
import math
import unicodedata
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from supabase_client import supabase
from auth_utils import get_current_gestor
from papeles_archivo import PAPELES_ARCHIVO, PAPELES_POR_SECCION
from instrumentos import INSTRUMENTO_A_SECCION

router = APIRouter(prefix="/api/gestor/archivo", tags=["archivo"])

GENERO_VALIDOS = ['SINF.', 'SINF.COR.', 'ESC.', 'COR.']
PROCEDENCIAS_VALIDAS = ['PROPIO', 'COMPRADO', 'ALQUILER', 'INTERNET', 'INTERNET-LIBRE', 'CESIÓN']


# ============================================================
# Pydantic models
# ============================================================
class ObraIn(BaseModel):
    autor: str
    arreglista: Optional[str] = None
    co_autor: Optional[str] = None
    titulo: str
    movimiento: Optional[str] = None
    genero: Optional[str] = None
    subgenero: Optional[str] = None
    procedencia: Optional[str] = None
    fecha_registro: Optional[str] = None
    observaciones: Optional[str] = None
    estado: Optional[str] = 'activo'
    obra_provisional: Optional[bool] = False
    codigo: Optional[str] = None


class PartePayload(BaseModel):
    papel: str
    copias_fisicas: Optional[int] = 0
    copia_digital: Optional[bool] = False
    enlace_drive: Optional[str] = None
    estado: Optional[str] = 'pendiente'
    notas: Optional[str] = None


class OriginalPayload(BaseModel):
    tipo: str  # general | partes | arcos
    estado: str  # si | no | necesita_revision
    notas: Optional[str] = None


class PrestamoIn(BaseModel):
    obra_id: str
    tipo: str  # interno | externo
    evento_id: Optional[str] = None
    entidad_externa: Optional[str] = None
    contacto_externo: Optional[str] = None
    partes_prestadas: Optional[List[str]] = []
    fecha_salida: str
    fecha_prevista_devolucion: Optional[str] = None
    notas: Optional[str] = None


class PrestamoUpdate(BaseModel):
    estado: Optional[str] = None
    fecha_devolucion_real: Optional[str] = None
    notas: Optional[str] = None
    fecha_prevista_devolucion: Optional[str] = None


class EventoObraIn(BaseModel):
    obra_id: Optional[str] = None
    titulo_provisional: Optional[str] = None
    estado: Optional[str] = 'provisional'
    orden_programa: Optional[int] = 1
    notas: Optional[str] = None
    duracion_display: Optional[str] = None
    autor_display: Optional[str] = None


class EventoObraPatch(BaseModel):
    """Patch parcial de una fila de evento_obras (Iter F3)."""
    obra_id: Optional[str] = None
    titulo_provisional: Optional[str] = None
    estado: Optional[str] = None
    orden_programa: Optional[int] = None
    notas: Optional[str] = None
    duracion_display: Optional[str] = None
    autor_display: Optional[str] = None


class ListaObrasFavoritaItem(BaseModel):
    obra_id: Optional[str] = None
    titulo_provisional: Optional[str] = None
    duracion_display: Optional[str] = None
    autor_display: Optional[str] = None
    notas: Optional[str] = None
    orden: Optional[int] = None


class ListaObrasFavoritaIn(BaseModel):
    nombre: str
    descripcion: Optional[str] = None
    obras: List[ListaObrasFavoritaItem] = []


class EtiquetasReq(BaseModel):
    incluye_general: bool = False
    incluye_partes: bool = False
    incluye_arcos: bool = False
    incluye_documentacion: bool = False
    incluye_atriles: bool = True
    incluye_atril_coro: bool = False
    incluye_atril_cuerda: bool = False
    incluye_atril_viento: bool = False
    incluye_atril_percusion: bool = False


# ============================================================
# Helpers
# ============================================================
def _strip_accents(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFKD', s or '') if not unicodedata.combining(c))


def _iniciales_autor(autor: str) -> str:
    """De 'Mozart, Wolfgang Amadeus' → 'MW'. De 'John Williams' → 'JW'."""
    autor = (autor or '').strip()
    if not autor:
        return 'XX'
    if ',' in autor:
        ap, nom = autor.split(',', 1)
        a = _strip_accents(ap.strip())[:1].upper() or 'X'
        n = _strip_accents(nom.strip())[:1].upper() or 'X'
        return a + n
    parts = [p for p in re.split(r'\s+', _strip_accents(autor)) if p]
    if len(parts) == 1:
        return (parts[0][:2] or 'XX').upper()
    return ((parts[0][:1] + parts[-1][:1]) or 'XX').upper()


def _generar_codigo(autor: str) -> str:
    iniciales = _iniciales_autor(autor)
    res = supabase.table('obras').select('codigo').like('codigo', f'{iniciales}/N%').execute().data or []
    nums = []
    for r in res:
        m = re.search(r'/Nº?(\d+)', r.get('codigo') or '')
        if m:
            nums.append(int(m.group(1)))
    nxt = (max(nums) + 1) if nums else 1
    return f"{iniciales}/Nº{nxt:03d}"


def _validar_obra(p: dict) -> dict:
    if p.get('genero') and p['genero'] not in GENERO_VALIDOS:
        raise HTTPException(status_code=400, detail=f"genero inválido. Usa uno de {GENERO_VALIDOS}")
    if p.get('procedencia') and p['procedencia'] not in PROCEDENCIAS_VALIDAS:
        raise HTTPException(status_code=400, detail=f"procedencia inválida. Usa uno de {PROCEDENCIAS_VALIDAS}")
    return p


# ============================================================
# CATÁLOGO
# ============================================================
@router.get("/obras")
async def listar_obras(
    q: Optional[str] = None,
    genero: Optional[str] = None,
    subgenero: Optional[str] = None,
    procedencia: Optional[str] = None,
    estado: Optional[str] = None,
    current_user: dict = Depends(get_current_gestor),
):
    """Lista obras del catálogo con filtros opcionales.

    - `q`: búsqueda full-text sobre `tsv` (titulo+autor+codigo) con
      stemming en español + unaccent (índice GIN). Fallback automático a
      ILIKE en Python si la columna `tsv` aún no existe.
    - Devuelve `total_copias_atril` por obra (suma de `obra_partes.copias_fisicas`).
    """
    sel = supabase.table('obras').select('*').order('autor').order('codigo').limit(2000)
    if genero:
        sel = sel.eq('genero', genero)
    if procedencia:
        sel = sel.eq('procedencia', procedencia)
    if estado:
        sel = sel.eq('estado', estado)
    if subgenero:
        sel = sel.ilike('subgenero', f'%{subgenero}%')

    obras: list = []
    if q:
        # Estrategia híbrida:
        #   - Branch A (full-text con stemming spanish): cubre títulos.
        #     'navidad' → indexa 'navid' y query 'navid':* matchea.
        #   - Branch B (ILIKE substring): cubre autor + codigo (nombres
        #     propios y códigos donde el stemming es contraproducente).
        # Unión de ambos por id. Si la columna `tsv` no existiera, sólo
        # contribuye la branch B.
        q_norm = _strip_accents(q).strip()
        ids_a: set = set()
        ids_b: set = set()

        # Branch A: full-text. Usamos plain ('plfts') para que los espacios
        # se traten como AND y no rompan la sintaxis tsquery.
        try:
            sel_a = supabase.table('obras').select('id')
            if genero: sel_a = sel_a.eq('genero', genero)
            if procedencia: sel_a = sel_a.eq('procedencia', procedencia)
            if estado: sel_a = sel_a.eq('estado', estado)
            if subgenero: sel_a = sel_a.ilike('subgenero', f'%{subgenero}%')
            ra = sel_a.text_search(
                'tsv', q_norm,
                options={'config': 'spanish', 'type': 'plain'},
            ).execute().data or []
            ids_a = {r['id'] for r in ra}
        except Exception:
            ids_a = set()

        # Branch B: ILIKE sobre autor + codigo (sin stemming). Sustituye
        # también de fallback total si la columna tsv no estuviera creada.
        sel_b = supabase.table('obras').select('id,autor,codigo,titulo')
        if genero: sel_b = sel_b.eq('genero', genero)
        if procedencia: sel_b = sel_b.eq('procedencia', procedencia)
        if estado: sel_b = sel_b.eq('estado', estado)
        if subgenero: sel_b = sel_b.ilike('subgenero', f'%{subgenero}%')
        # Escapar comas y comillas en el patrón ILIKE para PostgREST or_
        q_safe = q.replace(',', '').replace('(', '').replace(')', '')
        try:
            rb = sel_b.or_(
                f'autor.ilike.*{q_safe}*,codigo.ilike.*{q_safe}*,titulo.ilike.*{q_safe}*'
            ).execute().data or []
            ids_b = {r['id'] for r in rb}
        except Exception:
            # Último fallback: traer todo y filtrar en Python
            todo = sel.execute().data or []
            ql = q.lower()
            return {
                "obras": [
                    {**o, 'total_copias_atril': 0}
                    for o in todo
                    if ql in (o.get('titulo') or '').lower()
                    or ql in (o.get('autor') or '').lower()
                    or ql in (o.get('codigo') or '').lower()
                ],
                "total": 0,
            }

        ids_match = ids_a | ids_b
        if not ids_match:
            obras = []
        else:
            ids_list = list(ids_match)
            obras = []
            for i in range(0, len(ids_list), 200):
                chunk = ids_list[i:i + 200]
                sel_c = supabase.table('obras').select('*').in_('id', chunk).order('autor').order('codigo')
                obras.extend(sel_c.execute().data or [])
    else:
        obras = sel.execute().data or []

    # Calcular Nº copias atril (suma de obra_partes.copias_fisicas) por obra
    obra_ids = [o['id'] for o in obras]
    copias_por_obra: Dict[str, int] = {}
    for i in range(0, len(obra_ids), 200):
        chunk = obra_ids[i:i + 200]
        r = supabase.table('obra_partes').select('obra_id,copias_fisicas').in_('obra_id', chunk).execute()
        for p in (r.data or []):
            oid = p['obra_id']
            copias_por_obra[oid] = copias_por_obra.get(oid, 0) + (p.get('copias_fisicas') or 0)
    for o in obras:
        o['total_copias_atril'] = copias_por_obra.get(o['id'], 0)

    return {"obras": obras, "total": len(obras)}


@router.post("/obras")
async def crear_obra(data: ObraIn, current_user: dict = Depends(get_current_gestor)):
    payload = data.model_dump(exclude_none=True)
    # Normalizar strings vacías a None para no violar CHECK constraints
    for k in ('genero', 'procedencia', 'arreglista', 'co_autor', 'movimiento',
              'subgenero', 'fecha_registro', 'observaciones'):
        if k in payload and (payload[k] is None or str(payload[k]).strip() == ''):
            payload.pop(k, None)
    payload = _validar_obra(payload)
    if not payload.get('codigo'):
        payload['codigo'] = _generar_codigo(payload['autor'])

    # ---- Insert atómico de obra + 3 obra_originales (general/partes/arcos) ----
    # Si cualquiera de los 3 inserts secundarios falla, se hace rollback de
    # todo (incluyendo la obra) para no dejar registros huérfanos.
    try:
        res = supabase.table('obras').insert(payload).execute()
    except Exception as e:
        # Posible colisión con UNIQUE(codigo) → reintentar una vez con nuevo número
        if 'codigo' in str(e).lower() and ('unique' in str(e).lower() or 'duplicate' in str(e).lower()):
            payload['codigo'] = _generar_codigo(payload['autor'])
            res = supabase.table('obras').insert(payload).execute()
        else:
            raise HTTPException(status_code=400, detail=f"Error creando obra: {str(e)}")

    obra = (res.data or [None])[0]
    if not obra:
        raise HTTPException(status_code=500, detail="No se pudo crear la obra.")

    obra_id = obra['id']
    originales_payload = [
        {"obra_id": obra_id, "tipo": tipo, "estado": 'no'}
        for tipo in ('general', 'partes', 'arcos')
    ]
    try:
        # Insert batch — Supabase lo procesa en una sola query al PostgREST.
        supabase.table('obra_originales').insert(originales_payload).execute()
    except Exception as e:
        # Rollback compensatorio: borrar la obra recién creada para no dejar huérfana
        try:
            supabase.table('obras').delete().eq('id', obra_id).execute()
        except Exception:
            pass
        raise HTTPException(
            status_code=500,
            detail=f"Error creando originales (rollback aplicado): {str(e)}",
        )

    return {"obra": obra}


@router.get("/obras/{obra_id}")
async def detalle_obra(obra_id: str, current_user: dict = Depends(get_current_gestor)):
    o = supabase.table('obras').select('*').eq('id', obra_id).limit(1).execute().data or []
    if not o:
        raise HTTPException(status_code=404, detail="Obra no encontrada")
    originales = supabase.table('obra_originales').select('*').eq('obra_id', obra_id).execute().data or []
    partes = supabase.table('obra_partes').select('*').eq('obra_id', obra_id).execute().data or []
    prestamos = supabase.table('obra_prestamos').select('*').eq('obra_id', obra_id) \
        .order('fecha_salida', desc=True).execute().data or []
    eventos = supabase.table('evento_obras').select('*, evento:eventos(id,nombre,fecha_inicio,estado)') \
        .eq('obra_id', obra_id).execute().data or []
    return {"obra": o[0], "originales": originales, "partes": partes, "prestamos": prestamos, "eventos": eventos}


@router.put("/obras/{obra_id}")
async def actualizar_obra(obra_id: str, data: ObraIn, current_user: dict = Depends(get_current_gestor)):
    payload = _validar_obra(data.model_dump(exclude_none=True))
    payload['updated_at'] = datetime.now().isoformat()
    supabase.table('obras').update(payload).eq('id', obra_id).execute()
    # Actualizar originales si vienen
    return {"ok": True}


@router.put("/obras/{obra_id}/originales")
async def actualizar_originales(obra_id: str, data: List[OriginalPayload], current_user: dict = Depends(get_current_gestor)):
    for o in data:
        existing = supabase.table('obra_originales').select('id').eq('obra_id', obra_id).eq('tipo', o.tipo).limit(1).execute().data or []
        payload = {"obra_id": obra_id, "tipo": o.tipo, "estado": o.estado, "notas": o.notas}
        if existing:
            supabase.table('obra_originales').update(payload).eq('id', existing[0]['id']).execute()
        else:
            supabase.table('obra_originales').insert(payload).execute()
    return {"ok": True}


@router.put("/obras/{obra_id}/partes")
async def actualizar_partes(obra_id: str, data: List[PartePayload], current_user: dict = Depends(get_current_gestor)):
    """Reemplaza todas las partes de una obra. Insert/update por papel."""
    existentes = supabase.table('obra_partes').select('id,papel').eq('obra_id', obra_id).execute().data or []
    by_papel = {r['papel']: r['id'] for r in existentes}
    for p in data:
        meta = PAPELES_ARCHIVO.get(p.papel)
        if not meta:
            continue
        payload = {
            "obra_id": obra_id,
            "papel": p.papel,
            "instrumento": meta['instrumento'],
            "seccion": meta['seccion'],
            "copias_fisicas": p.copias_fisicas or 0,
            "copia_digital": bool(p.copia_digital),
            "enlace_drive": p.enlace_drive,
            "estado": p.estado or 'pendiente',
            "notas": p.notas,
        }
        if p.papel in by_papel:
            supabase.table('obra_partes').update(payload).eq('id', by_papel[p.papel]).execute()
        else:
            supabase.table('obra_partes').insert(payload).execute()
    return {"ok": True, "total": len(data)}


# ============================================================
# Bloque 3 — Cálculo de atriles
# ============================================================
def _instrumento_canon(s: str) -> str:
    if not s:
        return ''
    return INSTRUMENTO_A_SECCION.__class__  # placeholder


def calcular_atriles_necesarios(obra_id: str, evento_id: str) -> Dict[str, Any]:
    """Calcula cuántos atriles físicos se necesitan por papel y compara con copias_fisicas.
    Devuelve {atriles: {papel: necesarios}, alertas: [{papel, necesarios, copias, deficit}]}.
    """
    # 1. Músicos confirmados del evento
    asigs = supabase.table('asignaciones').select('usuario_id,estado') \
        .eq('evento_id', evento_id).eq('estado', 'confirmado').execute().data or []
    user_ids = list({a['usuario_id'] for a in asigs if a.get('usuario_id')})
    cnt_por_instrumento: Dict[str, int] = {}
    if user_ids:
        users = supabase.table('usuarios').select('id,instrumento').in_('id', user_ids).execute().data or []
        for u in users:
            instr = (u.get('instrumento') or '').strip().lower()
            instr = INSTRUMENTO_A_SECCION_KEY_NORM(instr)
            if instr:
                cnt_por_instrumento[instr] = cnt_por_instrumento.get(instr, 0) + 1

    atriles: Dict[str, int] = {}

    # 2. Reglas de cuerda (excepto contrabajo): ceil(N/2)
    cuerda_pares = {'violin_1': 'violin', 'violin_2': 'violin', 'viola': 'viola', 'violonchelo': 'violonchelo'}
    for papel, instr_key in cuerda_pares.items():
        if papel in ('violin_1', 'violin_2'):
            # Violín se reparte: half a violin_1, half a violin_2
            n = cnt_por_instrumento.get('violin', 0)
            half = math.ceil(n / 2) if n > 0 else 0
            # Para 1ª y 2ª: cada una asume mitad y atriles = ceil(mitad/2)
            atriles[papel] = max(1, math.ceil((n / 2) / 2)) if n > 0 else 0
        else:
            n = cnt_por_instrumento.get(instr_key, 0)
            atriles[papel] = math.ceil(n / 2) if n > 0 else 0

    # Contrabajo: 1 atril si hay >= 1
    atriles['contrabajo'] = 1 if cnt_por_instrumento.get('contrabajo', 0) > 0 else 0

    # 3. Trompa: regla especial 1_3 / 2_4
    n_trompa = cnt_por_instrumento.get('trompa', 0)
    atriles['trompa_1_3'] = 1 if n_trompa >= 1 else 0
    atriles['trompa_2_4'] = 1 if n_trompa >= 2 else 0

    # 4. Viento madera: 1ª y 2ª
    for instr_key in ('flauta', 'oboe', 'clarinete', 'fagot'):
        n = cnt_por_instrumento.get(instr_key, 0)
        atriles[f'{instr_key}_1'] = 1 if n >= 1 else 0
        atriles[f'{instr_key}_2'] = 1 if n >= 2 else 0

    # Trompeta y Trombón: 4 papeles
    for instr_key, base in (('trompeta', 'trompeta'), ('trombon', 'trombón')):
        n = cnt_por_instrumento.get(instr_key, 0)
        for i in range(1, 5):
            atriles[f'{base}_{i}'] = 1 if n >= i else 0

    # Tuba
    atriles['tuba'] = 1 if cnt_por_instrumento.get('tuba', 0) > 0 else 0

    # Percusión y coro: 1 por papel si hay al menos 1 músico de la sección/instrumento
    n_perc = cnt_por_instrumento.get('percusion', 0) + cnt_por_instrumento.get('timbales', 0)
    atriles['timbales'] = 1 if cnt_por_instrumento.get('timbales', 0) > 0 else 0
    atriles['bombo'] = 1 if n_perc > 0 else 0
    atriles['caja'] = 1 if n_perc >= 1 else 0
    for i in range(1, 4):
        atriles[f'percusion_{i}'] = 1 if n_perc >= i else 0
    atriles['laminas'] = 1 if n_perc >= 1 else 0

    for voz in ('soprano', 'alto', 'tenor', 'baritono'):
        atriles[voz] = 1 if cnt_por_instrumento.get(voz, 0) > 0 else 0

    for tecla in ('piano', 'organo', 'clave'):
        atriles[tecla] = 1 if cnt_por_instrumento.get(tecla, 0) > 0 else 0

    # 5. Comparar con copias_fisicas
    partes_obra = supabase.table('obra_partes').select('papel,copias_fisicas').eq('obra_id', obra_id).execute().data or []
    by_papel = {p['papel']: int(p.get('copias_fisicas') or 0) for p in partes_obra}
    alertas = []
    for papel, necesarios in atriles.items():
        if necesarios <= 0:
            continue
        copias = by_papel.get(papel, 0)
        if copias < necesarios:
            alertas.append({
                "papel": papel,
                "label": (PAPELES_ARCHIVO.get(papel) or {}).get('label') or papel,
                "necesarios": necesarios,
                "copias": copias,
                "deficit": necesarios - copias,
            })

    return {
        "atriles": atriles,
        "alertas": alertas,
        "musicos_por_instrumento": cnt_por_instrumento,
    }


def INSTRUMENTO_A_SECCION_KEY_NORM(s: str) -> str:
    """Normaliza un instrumento de 'usuarios.instrumento' a la clave usada en PAPELES_ARCHIVO."""
    s = (s or '').strip().lower()
    # Quitar acentos para matching tolerante
    s_na = _strip_accents(s)
    table = {
        'violin': 'violin', 'violines': 'violin',
        'viola': 'viola', 'violas': 'viola',
        'cello': 'violonchelo', 'chelo': 'violonchelo', 'violonchelo': 'violonchelo', 'violoncello': 'violonchelo',
        'contrabajo': 'contrabajo', 'contrabajos': 'contrabajo',
        'flauta': 'flauta', 'flauta travesera': 'flauta', 'flautin': 'flauta',
        'oboe': 'oboe', 'corno ingles': 'oboe',
        'clarinete': 'clarinete', 'clarinete bajo': 'clarinete',
        'fagot': 'fagot', 'contrafagot': 'fagot',
        'trompa': 'trompa', 'corno': 'trompa',
        'trompeta': 'trompeta',
        'trombon': 'trombon',
        'tuba': 'tuba',
        'percusion': 'percusion', 'timbales': 'timbales',
        'piano': 'piano', 'organo': 'organo', 'clave': 'clave',
        'soprano': 'soprano', 'alto': 'alto', 'tenor': 'tenor',
        'baritono': 'baritono',
    }
    return table.get(s_na, s_na)


@router.get("/obras/{obra_id}/atriles-evento/{evento_id}")
async def atriles_evento(obra_id: str, evento_id: str, current_user: dict = Depends(get_current_gestor)):
    return calcular_atriles_necesarios(obra_id, evento_id)


# ============================================================
# Bloque 5 — Importación masiva + plantilla
# ============================================================
@router.get("/plantilla-obras")
async def descargar_plantilla(current_user: dict = Depends(get_current_gestor)):
    """Genera y devuelve un Excel plantilla para importar obras."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")
    wb = Workbook()
    ws = wb.active
    ws.title = "Obras"
    headers = [
        'autor', 'arreglista', 'co_autor', 'titulo', 'movimiento', 'genero', 'subgenero',
        'procedencia', 'fecha_registro', 'original_general', 'original_partes', 'original_arcos',
        'enlace_digital', 'observaciones',
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F2937')
        c.alignment = Alignment(horizontal='center')
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
    # Hoja instrucciones
    inst = wb.create_sheet('INSTRUCCIONES')
    inst['A1'] = 'INSTRUCCIONES DE IMPORTACIÓN'
    inst['A1'].font = Font(bold=True, size=14)
    rows = [
        ('autor (obligatorio)', 'Texto. Ej: "Mozart, Wolfgang Amadeus"'),
        ('titulo (obligatorio)', 'Texto. Ej: "Sinfonía nº 40"'),
        ('genero', f"Uno de: {', '.join(GENERO_VALIDOS)}"),
        ('procedencia', f"Uno de: {', '.join(PROCEDENCIAS_VALIDAS)}"),
        ('fecha_registro', 'Formato YYYY-MM-DD. Opcional'),
        ('original_general/partes/arcos', "Uno de: si, no, necesita_revision"),
        ('enlace_digital', 'URL del archivo en Google Drive (opcional)'),
        ('código', 'Se genera automáticamente con formato XX/Nº001'),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        inst.cell(row=i, column=1, value=k).font = Font(bold=True)
        inst.cell(row=i, column=2, value=v)
    inst.column_dimensions['A'].width = 30
    inst.column_dimensions['B'].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=plantilla_obras.xlsx'},
    )


def _norm_estado_original(v: Any) -> str:
    s = (str(v or '').strip().lower())
    if s in ('si', 'sí', 'yes', 'true', '1'):
        return 'si'
    if s in ('necesita revision', 'necesita_revision', 'needs review', 'revisar', 'revision'):
        return 'necesita_revision'
    return 'no'


@router.post("/obras/importar")
async def importar_obras(
    archivo: UploadFile = File(...),
    confirmar: bool = False,
    current_user: dict = Depends(get_current_gestor),
):
    """Importa obras desde Excel. Si confirmar=False, devuelve solo preview + validación."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")
    raw = await archivo.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel inválido: {e}")
    headers = [str(c.value or '').strip() for c in ws[1]]
    rows: List[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append(dict(zip(headers, row)))

    obras_existentes = supabase.table('obras').select('autor,titulo').execute().data or []
    existentes_set = {(o.get('autor', '').strip().lower(), o.get('titulo', '').strip().lower()) for o in obras_existentes}

    preview, errors, ya_existen = [], [], 0
    for idx, r in enumerate(rows, start=2):
        autor = (r.get('autor') or '').strip()
        titulo = (r.get('titulo') or '').strip()
        if not autor or not titulo:
            errors.append({"fila": idx, "error": "Falta autor o titulo"})
            continue
        if (autor.lower(), titulo.lower()) in existentes_set:
            ya_existen += 1
            continue
        genero = (r.get('genero') or '').strip() or None
        if genero and genero not in GENERO_VALIDOS:
            errors.append({"fila": idx, "error": f"genero inválido '{genero}'"})
            continue
        procedencia = (r.get('procedencia') or '').strip() or None
        if procedencia and procedencia not in PROCEDENCIAS_VALIDAS:
            errors.append({"fila": idx, "error": f"procedencia inválida '{procedencia}'"})
            continue
        fecha_reg = r.get('fecha_registro')
        if fecha_reg and hasattr(fecha_reg, 'strftime'):
            fecha_reg = fecha_reg.strftime('%Y-%m-%d')
        elif fecha_reg:
            fecha_reg = str(fecha_reg)
        preview.append({
            "fila": idx,
            "autor": autor,
            "arreglista": (r.get('arreglista') or '').strip() or None,
            "co_autor": (r.get('co_autor') or '').strip() or None,
            "titulo": titulo,
            "movimiento": (r.get('movimiento') or '').strip() or None,
            "genero": genero,
            "subgenero": (r.get('subgenero') or '').strip() or None,
            "procedencia": procedencia,
            "fecha_registro": fecha_reg,
            "observaciones": (r.get('observaciones') or '').strip() or None,
            "_originales": {
                'general': _norm_estado_original(r.get('original_general')),
                'partes':  _norm_estado_original(r.get('original_partes')),
                'arcos':   _norm_estado_original(r.get('original_arcos')),
            },
            "_enlace_digital": (r.get('enlace_digital') or '').strip() or None,
        })

    if not confirmar:
        return {
            "preview": preview[:5],
            "total": len(preview),
            "errores": errors,
            "ya_existentes": ya_existen,
        }

    # Insertar
    insertadas = 0
    for p in preview:
        originales_data = p.pop('_originales', {})
        enlace = p.pop('_enlace_digital', None)
        p.pop('fila', None)
        p['codigo'] = _generar_codigo(p['autor'])
        try:
            res = supabase.table('obras').insert(p).execute()
            obra_id = res.data[0]['id'] if res.data else None
            if obra_id:
                for tipo, est in originales_data.items():
                    supabase.table('obra_originales').insert({
                        "obra_id": obra_id, "tipo": tipo, "estado": est,
                    }).execute()
                if enlace:
                    supabase.table('obras').update({"observaciones": (p.get('observaciones') or '') + f" | Drive: {enlace}"}).eq('id', obra_id).execute()
            insertadas += 1
        except Exception as e:
            errors.append({"autor": p['autor'], "titulo": p['titulo'], "error": str(e)})
    return {"importadas": insertadas, "errores": errors, "ya_existentes": ya_existen}


# ============================================================
# Bloque 6 — Etiquetas PDF
# ============================================================
@router.post("/obras/{obra_id}/etiquetas")
async def generar_etiquetas(obra_id: str, req: EtiquetasReq, current_user: dict = Depends(get_current_gestor)):
    """Genera un PDF con etiquetas según los tipos seleccionados."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab no instalado")

    obra = supabase.table('obras').select('*').eq('id', obra_id).limit(1).execute().data or []
    if not obra:
        raise HTTPException(status_code=404, detail="Obra no encontrada")
    obra = obra[0]
    partes = supabase.table('obra_partes').select('*').eq('obra_id', obra_id).execute().data or []

    etiquetas: List[str] = []
    if req.incluye_general:       etiquetas.append('Copia ORIGINAL — General')
    if req.incluye_partes:        etiquetas.extend(['Copia ORIGINAL — Partes'] * len(partes) or ['Copia ORIGINAL — Partes'])
    if req.incluye_arcos:         etiquetas.append('Copia ORIGINAL — Arcos de cuerda')
    if req.incluye_documentacion: etiquetas.append('Documentación de registro')
    if req.incluye_atriles:
        for p in partes:
            n = int(p.get('copias_fisicas') or 0)
            if n > 0:
                label = (PAPELES_ARCHIVO.get(p['papel']) or {}).get('label') or p['papel']
                etiquetas.extend([f'ATRIL — {label}'] * n)
    if req.incluye_atril_coro:      etiquetas.append('ATRIL — Coro')
    if req.incluye_atril_cuerda:    etiquetas.append('ATRIL — Cuerda')
    if req.incluye_atril_viento:    etiquetas.append('ATRIL — Viento')
    if req.incluye_atril_percusion: etiquetas.append('ATRIL — Percusión')

    if not etiquetas:
        raise HTTPException(status_code=400, detail="Selecciona al menos un tipo de etiqueta")

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 12 * mm
    cols, lbl_w, lbl_h = 2, (width - 2 * margin - 6 * mm) / 2, 50 * mm
    rows_per_page = int((height - 2 * margin) // lbl_h)

    for idx, txt in enumerate(etiquetas):
        i = idx % (cols * rows_per_page)
        if idx > 0 and i == 0:
            c.showPage()
        col = i % cols
        row = i // cols
        x = margin + col * (lbl_w + 6 * mm)
        y = height - margin - (row + 1) * lbl_h
        # Borde
        c.rect(x, y, lbl_w, lbl_h)
        # Contenido
        c.setFont('Helvetica', 9)
        c.drawString(x + 4 * mm, y + lbl_h - 8 * mm, f"AUTOR: {obra.get('autor', '')[:40]}")
        c.setFont('Helvetica-Bold', 11)
        c.drawString(x + 4 * mm, y + lbl_h - 16 * mm, f"TÍTULO: {obra.get('titulo', '')[:36]}")
        c.setFont('Helvetica', 7)
        c.line(x + 4 * mm, y + lbl_h - 20 * mm, x + lbl_w - 4 * mm, y + lbl_h - 20 * mm)
        c.drawString(x + 4 * mm, y + lbl_h - 25 * mm, f"CÓDIGO: {obra.get('codigo', '')}")
        c.drawString(x + lbl_w / 2, y + lbl_h - 25 * mm, f"GÉNERO: {obra.get('genero') or '—'}")
        c.setFont('Helvetica-Bold', 10)
        c.drawCentredString(x + lbl_w / 2, y + 6 * mm, txt)
    c.save()
    buf.seek(0)
    fn = f"etiquetas_{(obra.get('codigo') or 'obra').replace('/', '_')}.pdf"
    return StreamingResponse(buf, media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={fn}'})


# ============================================================
# PRÉSTAMOS
# ============================================================
@router.get("/prestamos")
async def listar_prestamos(estado: Optional[str] = None, tipo: Optional[str] = None,
                            current_user: dict = Depends(get_current_gestor)):
    sel = supabase.table('obra_prestamos') \
        .select('*, obra:obras(id,codigo,titulo,autor), evento:eventos(id,nombre,fecha_inicio)') \
        .order('fecha_prevista_devolucion', desc=False)
    if estado:
        sel = sel.eq('estado', estado)
    if tipo:
        sel = sel.eq('tipo', tipo)
    data = sel.execute().data or []
    return {"prestamos": data}


@router.post("/prestamos")
async def crear_prestamo(data: PrestamoIn, current_user: dict = Depends(get_current_gestor)):
    payload = data.model_dump(exclude_none=True)
    payload['gestor_id'] = (current_user.get('profile') or {}).get('id')
    res = supabase.table('obra_prestamos').insert(payload).execute()
    return {"prestamo": (res.data or [None])[0]}


@router.put("/prestamos/{prestamo_id}")
async def actualizar_prestamo(prestamo_id: str, data: PrestamoUpdate,
                                current_user: dict = Depends(get_current_gestor)):
    payload = data.model_dump(exclude_none=True)
    payload['updated_at'] = datetime.now().isoformat()
    supabase.table('obra_prestamos').update(payload).eq('id', prestamo_id).execute()
    return {"ok": True}


# ============================================================
# Bloque 4 — Alertas dashboard
# ============================================================
@router.get("/alertas")
async def alertas_archivo(current_user: dict = Depends(get_current_gestor)):
    from datetime import date
    today = date.today()
    # Pendientes de registrar
    obras_provisionales = supabase.table('evento_obras') \
        .select('*, evento:eventos(id,nombre,fecha_inicio)') \
        .eq('estado', 'provisional').execute().data or []
    # Préstamos vencidos / próximos
    prestamos_act = supabase.table('obra_prestamos') \
        .select('*, obra:obras(id,codigo,titulo,autor)') \
        .eq('estado', 'activo').execute().data or []
    vencidos, proximos = [], []
    for p in prestamos_act:
        fp = p.get('fecha_prevista_devolucion')
        if not fp:
            continue
        try:
            d = date.fromisoformat(fp)
        except Exception:
            continue
        if d < today:
            vencidos.append(p)
        elif (d - today).days <= 7:
            proximos.append(p)
    # Material incompleto: partes con estado='incompleto'
    incompletas = supabase.table('obra_partes') \
        .select('obra_id,papel,copias_fisicas,estado') \
        .eq('estado', 'incompleto').execute().data or []
    # Originales que necesitan revisión (importados del Excel histórico)
    orig_revision = supabase.table('obra_originales') \
        .select('obra_id,tipo,estado,notas, obra:obras(id,codigo,titulo,autor,genero)') \
        .eq('estado', 'necesita_revision') \
        .order('obra_id').execute().data or []
    return {
        "obras_pendientes_registro": obras_provisionales,
        "prestamos_vencidos": vencidos,
        "prestamos_proximos": proximos,
        "partes_incompletas": incompletas,
        "originales_necesita_revision": orig_revision,
    }


# ============================================================
# Bloque 7 — Programa del evento (vinculación con catálogo)
# ============================================================
@router.get("/evento/{evento_id}/programa")
async def programa_evento(evento_id: str, current_user: dict = Depends(get_current_gestor)):
    rows = supabase.table('evento_obras') \
        .select('*, obra:obras(id,codigo,titulo,autor,genero)') \
        .eq('evento_id', evento_id).order('orden_programa').execute().data or []
    return {"programa": rows}


@router.post("/evento/{evento_id}/obras")
async def agregar_obra_evento(evento_id: str, data: EventoObraIn, current_user: dict = Depends(get_current_gestor)):
    """Vincula una obra al programa de un evento.
    Si se pasa solo `titulo_provisional`, se busca match en obras y, si no existe,
    se crea como provisional + notificación a archiveros.
    """
    payload = data.model_dump(exclude_none=True)
    payload['evento_id'] = evento_id
    if not payload.get('obra_id') and payload.get('titulo_provisional'):
        # Match por título exacto
        candidato = supabase.table('obras').select('id,titulo,autor') \
            .ilike('titulo', payload['titulo_provisional'].strip()).limit(1).execute().data or []
        if candidato:
            payload['obra_id'] = candidato[0]['id']
            payload['estado'] = 'confirmada'
            payload.pop('titulo_provisional', None)
        else:
            payload['estado'] = 'provisional'
    res = supabase.table('evento_obras').insert(payload).execute()
    eo = (res.data or [None])[0]

    # Notificar a archiveros si la obra es provisional
    if eo and eo.get('estado') == 'provisional':
        try:
            evt = supabase.table('eventos').select('nombre').eq('id', evento_id).limit(1).execute().data or []
            evt_nombre = evt[0]['nombre'] if evt else 'Evento'
            archiveros = supabase.table('usuarios').select('id').eq('rol', 'archivero').execute().data or []
            for a in archiveros:
                supabase.table('notificaciones_gestor').insert({
                    "gestor_id": a['id'],
                    "tipo": "obra_pendiente_registro",
                    "titulo": f"Nueva obra pendiente: {eo.get('titulo_provisional')}",
                    "descripcion": f"Solicitada para evento: {evt_nombre}",
                    "entidad_tipo": "evento_obra",
                    "entidad_id": eo['id'],
                    "leida": False,
                }).execute()
        except Exception:
            pass
    return {"evento_obra": eo}


# ============================================================
# Iter F3 — PATCH/DELETE de filas del programa + migración silenciosa
# ============================================================
def _archivo_is_super_admin(current_user: dict) -> bool:
    try:
        from auth_utils import is_super_admin as _isa
        return _isa(current_user)
    except Exception:
        profile = current_user.get('profile') or {}
        rol = profile.get('rol')
        if rol in ('admin', 'director_general'):
            return True
        email = (profile.get('email') or '').lower()
        return email == 'admin@convocatorias.com'


@router.patch("/evento/{evento_id}/obras/{eo_id}")
async def actualizar_obra_evento(evento_id: str, eo_id: str, data: EventoObraPatch,
                                  current_user: dict = Depends(get_current_gestor)):
    """Edición parcial de una fila del programa (Iter F3)."""
    row = supabase.table('evento_obras').select('id,evento_id,titulo_provisional,obra_id,estado') \
        .eq('id', eo_id).eq('evento_id', evento_id).limit(1).execute().data or []
    if not row:
        raise HTTPException(status_code=404, detail="Fila de programa no encontrada")
    payload = data.model_dump(exclude_none=True)
    if not payload:
        return {"ok": True, "evento_obra": row[0]}

    # Si llega un nuevo titulo_provisional sin obra_id, intentar match con catálogo
    notify_archiveros = False
    if payload.get('titulo_provisional') and not payload.get('obra_id') and not row[0].get('obra_id'):
        candidato = supabase.table('obras').select('id') \
            .ilike('titulo', payload['titulo_provisional'].strip()).limit(1).execute().data or []
        if candidato:
            payload['obra_id'] = candidato[0]['id']
            payload['estado'] = 'confirmada'
            payload['titulo_provisional'] = None
        else:
            # Solo notificar si la obra antes no era provisional o el título cambió
            if row[0].get('estado') != 'provisional' or row[0].get('titulo_provisional') != payload['titulo_provisional']:
                payload['estado'] = 'provisional'
                notify_archiveros = True

    supabase.table('evento_obras').update(payload).eq('id', eo_id).execute()

    if notify_archiveros:
        try:
            evt = supabase.table('eventos').select('nombre').eq('id', evento_id).limit(1).execute().data or []
            evt_nombre = evt[0]['nombre'] if evt else 'Evento'
            archiveros = supabase.table('usuarios').select('id').eq('rol', 'archivero').execute().data or []
            for a in archiveros:
                supabase.table('notificaciones_gestor').insert({
                    "gestor_id": a['id'],
                    "tipo": "obra_pendiente_registro",
                    "titulo": f"Obra pendiente actualizada: {payload.get('titulo_provisional')}",
                    "descripcion": f"Solicitada para evento: {evt_nombre}",
                    "entidad_tipo": "evento_obra",
                    "entidad_id": eo_id,
                    "leida": False,
                }).execute()
        except Exception:
            pass

    nuevo = supabase.table('evento_obras').select('*, obra:obras(id,codigo,titulo,autor,genero)') \
        .eq('id', eo_id).limit(1).execute().data or []
    return {"ok": True, "evento_obra": (nuevo[0] if nuevo else None)}


@router.delete("/evento/{evento_id}/obras/{eo_id}")
async def eliminar_obra_evento(evento_id: str, eo_id: str,
                                current_user: dict = Depends(get_current_gestor)):
    """Borra una fila del programa (Iter F3 — botón borrar fila)."""
    row = supabase.table('evento_obras').select('id') \
        .eq('id', eo_id).eq('evento_id', evento_id).limit(1).execute().data or []
    if not row:
        raise HTTPException(status_code=404, detail="Fila de programa no encontrada")
    supabase.table('evento_obras').delete().eq('id', eo_id).execute()
    return {"ok": True}


@router.post("/evento/{evento_id}/programa/migrar")
async def migrar_programa_legacy(evento_id: str,
                                  current_user: dict = Depends(get_current_gestor)):
    """Migración silenciosa e idempotente del antiguo eventos.program JSON local
    a filas reales de evento_obras. Solo se ejecuta si evento_obras está vacío
    para este evento. No borra eventos.program (rollback seguro).
    """
    existentes = supabase.table('evento_obras').select('id') \
        .eq('evento_id', evento_id).limit(1).execute().data or []
    if existentes:
        return {"migrado": False, "motivo": "ya_tiene_filas"}

    evt = supabase.table('eventos').select('program,nombre') \
        .eq('id', evento_id).limit(1).execute().data or []
    if not evt:
        return {"migrado": False, "motivo": "evento_no_encontrado"}
    legacy = evt[0].get('program') or []
    if not isinstance(legacy, list) or len(legacy) == 0:
        return {"migrado": False, "motivo": "sin_legacy"}

    creados = []
    notificar_titulos = []
    for idx, item in enumerate(legacy):
        if not isinstance(item, dict):
            continue
        titulo = (item.get('obra') or '').strip()
        autor = (item.get('author') or '').strip() or None
        duracion = (item.get('duration') or '').strip() or None
        notas = (item.get('observaciones') or '').strip() or None
        if not titulo and not autor and not duracion and not notas:
            continue

        payload = {
            "evento_id": evento_id,
            "orden_programa": idx + 1,
            "duracion_display": duracion,
            "autor_display": autor,
            "notas": notas,
        }
        # Match catálogo por título
        obra_id = None
        if titulo:
            candidato = supabase.table('obras').select('id') \
                .ilike('titulo', titulo).limit(1).execute().data or []
            if candidato:
                obra_id = candidato[0]['id']
        if obra_id:
            payload['obra_id'] = obra_id
            payload['estado'] = 'confirmada'
        else:
            payload['titulo_provisional'] = titulo or '(sin título)'
            payload['estado'] = 'provisional'
            if titulo:
                notificar_titulos.append((None, titulo))

        try:
            res = supabase.table('evento_obras').insert(payload).execute()
            creado = (res.data or [None])[0]
            if creado:
                creados.append(creado)
                if creado.get('estado') == 'provisional':
                    notificar_titulos.append((creado.get('id'), creado.get('titulo_provisional')))
        except Exception:
            continue

    # Notificar a archiveros (1 sola vez por título migrado)
    if notificar_titulos:
        try:
            evt_nombre = evt[0].get('nombre') or 'Evento'
            archiveros = supabase.table('usuarios').select('id').eq('rol', 'archivero').execute().data or []
            for eo_id, titulo in notificar_titulos:
                if not eo_id:
                    continue
                for a in archiveros:
                    supabase.table('notificaciones_gestor').insert({
                        "gestor_id": a['id'],
                        "tipo": "obra_pendiente_registro",
                        "titulo": f"Nueva obra pendiente (migración): {titulo}",
                        "descripcion": f"Solicitada para evento: {evt_nombre}",
                        "entidad_tipo": "evento_obra",
                        "entidad_id": eo_id,
                        "leida": False,
                    }).execute()
        except Exception:
            pass

    return {"migrado": True, "creadas": len(creados)}


# ============================================================
# Iter F3 — Listas de obras favoritas (globales)
# ============================================================
@router.get("/listas-obras-favoritas")
async def listar_obras_favoritas(current_user: dict = Depends(get_current_gestor)):
    rows = supabase.table('listas_obras_favoritas').select('*') \
        .order('nombre').execute().data or []
    return {"listas": rows}


@router.post("/listas-obras-favoritas")
async def crear_obras_favorita(data: ListaObrasFavoritaIn,
                                current_user: dict = Depends(get_current_gestor)):
    profile = current_user.get('profile') or {}
    payload = {
        "nombre": data.nombre,
        "descripcion": data.descripcion,
        "creado_por": profile.get('id'),
        "obras": [it.model_dump(exclude_none=True) for it in data.obras],
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    res = supabase.table('listas_obras_favoritas').insert(payload).execute()
    return {"lista": (res.data or [None])[0]}


@router.put("/listas-obras-favoritas/{lista_id}")
async def actualizar_obras_favorita(lista_id: str, data: ListaObrasFavoritaIn,
                                     current_user: dict = Depends(get_current_gestor)):
    profile = current_user.get('profile') or {}
    row = supabase.table('listas_obras_favoritas').select('id,creado_por') \
        .eq('id', lista_id).limit(1).execute().data or []
    if not row:
        raise HTTPException(status_code=404, detail="Lista no encontrada")
    # Permisos: creador o super-admin
    if row[0].get('creado_por') and row[0]['creado_por'] != profile.get('id'):
        if not _archivo_is_super_admin(current_user):
            raise HTTPException(status_code=403, detail="Solo el creador o un administrador puede editar esta lista.")
    payload = {
        "nombre": data.nombre,
        "descripcion": data.descripcion,
        "obras": [it.model_dump(exclude_none=True) for it in data.obras],
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    supabase.table('listas_obras_favoritas').update(payload).eq('id', lista_id).execute()
    return {"ok": True, "lista_id": lista_id}


@router.delete("/listas-obras-favoritas/{lista_id}")
async def eliminar_obras_favorita(lista_id: str,
                                   current_user: dict = Depends(get_current_gestor)):
    profile = current_user.get('profile') or {}
    row = supabase.table('listas_obras_favoritas').select('id,creado_por') \
        .eq('id', lista_id).limit(1).execute().data or []
    if not row:
        raise HTTPException(status_code=404, detail="Lista no encontrada")
    if row[0].get('creado_por') and row[0]['creado_por'] != profile.get('id'):
        if not _archivo_is_super_admin(current_user):
            raise HTTPException(status_code=403, detail="Solo el creador o un administrador puede eliminar esta lista.")
    supabase.table('listas_obras_favoritas').delete().eq('id', lista_id).execute()
    return {"ok": True}


@router.get("/evento/{evento_id}/programa/pdf")
async def exportar_programa_pdf(evento_id: str, current_user: dict = Depends(get_current_gestor)):
    """B-PDF (2026-05-03) — Exporta el Programa Musical como PDF (ReportLab vía pdf_renderer).
    Incluye cabecera del evento (nombre, lugar, fechas, temporada), tabla de obras
    (orden, autor, obra, duración, notas) y total de duración al pie.
    """
    from html import escape as _esc
    from datetime import datetime as _dt
    try:
        from pdf_renderer import html_to_pdf_bytes
    except Exception:
        raise HTTPException(status_code=500, detail="pdf_renderer no disponible")

    # 1. Cargar evento
    ev_rows = supabase.table('eventos').select('id,nombre,lugar,fecha_inicio,fecha_fin,temporada') \
        .eq('id', evento_id).limit(1).execute().data or []
    if not ev_rows:
        raise HTTPException(status_code=404, detail="Evento no encontrado")
    ev = ev_rows[0]

    # 2. Cargar programa
    obras_rows = supabase.table('evento_obras') \
        .select('orden_programa,titulo_provisional,duracion_display,autor_display,notas,obra:obras(titulo,autor)') \
        .eq('evento_id', evento_id) \
        .order('orden_programa').execute().data or []

    # 3. Construir HTML
    def _row_html(idx, r):
        autor = (r.get('obra') or {}).get('autor') or r.get('autor_display') or ''
        titulo = (r.get('obra') or {}).get('titulo') or r.get('titulo_provisional') or '(sin título)'
        dur = r.get('duracion_display') or ''
        notas = r.get('notas') or ''
        return f"<tr><td>{idx}</td><td>{_esc(autor)}</td><td>{_esc(titulo)}</td><td>{_esc(dur)}</td><td>{_esc(notas)}</td></tr>"

    # 4. Suma tolerante de duración (acepta "15", "15'", "15:30", "1h 20'", etc.)
    def _to_minutes(s):
        if not s:
            return 0
        s = str(s).strip().lower().replace("′", "'").replace("’", "'")
        total = 0
        # h
        mh = re.search(r"(\d+)\s*h", s)
        if mh:
            total += int(mh.group(1)) * 60
            s = re.sub(r"\d+\s*h", "", s)
        # mm:ss → solo mm
        mc = re.search(r"(\d+)\s*:\s*\d+", s)
        if mc:
            total += int(mc.group(1))
            return total
        # primer número que aparezca = minutos
        mn = re.search(r"\d+", s)
        if mn:
            total += int(mn.group())
        return total

    total_min = sum(_to_minutes(r.get('duracion_display')) for r in obras_rows)
    total_str = f"{total_min // 60}h {total_min % 60:02d}'" if total_min >= 60 else f"{total_min}'"
    ahora = _dt.now().strftime('%d/%m/%Y %H:%M')

    fecha_str = ''
    if ev.get('fecha_inicio') and ev.get('fecha_fin') and ev['fecha_inicio'] != ev['fecha_fin']:
        fecha_str = f"{ev['fecha_inicio']} – {ev['fecha_fin']}"
    elif ev.get('fecha_inicio'):
        fecha_str = ev['fecha_inicio']

    info_partes = [p for p in [ev.get('lugar'), fecha_str, ev.get('temporada')] if p]
    info_linea = ' · '.join(info_partes)

    rows_html = "\n".join(_row_html(i + 1, r) for i, r in enumerate(obras_rows)) or \
        '<tr><td colspan="5"><i>Sin obras en el programa.</i></td></tr>'

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"></head><body>
<h1>Programa Musical</h1>
<h2>{_esc(ev.get('nombre') or 'Evento')}</h2>
<p>{_esc(info_linea)}</p>
<table>
  <tr><th>Nº</th><th>Autor</th><th>Obra</th><th>Duración</th><th>Notas</th></tr>
  {rows_html}
</table>
<p><b>Duración total estimada:</b> {_esc(total_str)}</p>
<p><i>Generado el {_esc(ahora)} por OPUS Manager.</i></p>
</body></html>"""

    pdf_bytes = html_to_pdf_bytes(html)
    safe_name = re.sub(r'[^\w\-]+', '_', (ev.get('nombre') or 'evento'))[:60]
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="programa_{safe_name}.pdf"'},
    )


@router.post("/evento/{evento_id}/programa/aplicar-lista/{lista_id}")
async def aplicar_lista_obras(evento_id: str, lista_id: str,
                               current_user: dict = Depends(get_current_gestor)):
    """Vuelca los items de una lista favorita como filas reales en evento_obras.
    No borra filas existentes; las añade al final del orden actual.
    """
    lst = supabase.table('listas_obras_favoritas').select('*') \
        .eq('id', lista_id).limit(1).execute().data or []
    if not lst:
        raise HTTPException(status_code=404, detail="Lista no encontrada")
    items = lst[0].get('obras') or []
    if not isinstance(items, list) or not items:
        return {"ok": True, "creadas": 0}

    # Calcular orden inicial
    existentes = supabase.table('evento_obras').select('orden_programa') \
        .eq('evento_id', evento_id).execute().data or []
    base = max([int(r.get('orden_programa') or 0) for r in existentes] or [0])

    creadas = []
    notificar = []
    for idx, item in enumerate(items):
        if not isinstance(item, dict):
            continue
        payload = {
            "evento_id": evento_id,
            "orden_programa": base + idx + 1,
            "duracion_display": item.get('duracion_display'),
            "autor_display": item.get('autor_display'),
            "notas": item.get('notas'),
        }
        if item.get('obra_id'):
            payload['obra_id'] = item['obra_id']
            payload['estado'] = 'confirmada'
        else:
            titulo = (item.get('titulo_provisional') or '').strip()
            if not titulo:
                continue
            # Re-match catálogo por si fue añadido recientemente
            cand = supabase.table('obras').select('id') \
                .ilike('titulo', titulo).limit(1).execute().data or []
            if cand:
                payload['obra_id'] = cand[0]['id']
                payload['estado'] = 'confirmada'
            else:
                payload['titulo_provisional'] = titulo
                payload['estado'] = 'provisional'
        try:
            res = supabase.table('evento_obras').insert(payload).execute()
            c = (res.data or [None])[0]
            if c:
                creadas.append(c)
                if c.get('estado') == 'provisional':
                    notificar.append((c.get('id'), c.get('titulo_provisional')))
        except Exception:
            continue

    # Notificar archiveros
    if notificar:
        try:
            evt = supabase.table('eventos').select('nombre').eq('id', evento_id).limit(1).execute().data or []
            evt_nombre = evt[0]['nombre'] if evt else 'Evento'
            archiveros = supabase.table('usuarios').select('id').eq('rol', 'archivero').execute().data or []
            for eo_id, titulo in notificar:
                for a in archiveros:
                    supabase.table('notificaciones_gestor').insert({
                        "gestor_id": a['id'],
                        "tipo": "obra_pendiente_registro",
                        "titulo": f"Nueva obra pendiente: {titulo}",
                        "descripcion": f"Solicitada para evento: {evt_nombre}",
                        "entidad_tipo": "evento_obra",
                        "entidad_id": eo_id,
                        "leida": False,
                    }).execute()
        except Exception:
            pass

    return {"ok": True, "creadas": len(creadas)}



# ============================================================
# Bloque 5 — Conflictos de préstamo de obra con fechas de evento
# ============================================================
@router.get("/obras/{obra_id}/conflictos-evento/{evento_id}")
async def conflictos_obra_evento(obra_id: str, evento_id: str,
                                  current_user: dict = Depends(get_current_gestor)):
    """Devuelve préstamos activos de la obra que solapan con las fechas de ensayos/funciones del evento."""
    # Fechas del evento (inicio y fin)
    evt = supabase.table('eventos').select('fecha_inicio,fecha_fin,nombre') \
        .eq('id', evento_id).limit(1).execute().data or []
    if not evt:
        return {"conflictos": [], "evento": None}
    ev = evt[0]
    f_ini = (ev.get('fecha_inicio') or '')[:10]
    f_fin = (ev.get('fecha_fin') or f_ini)[:10] or f_ini
    # Ensayos
    try:
        ens = supabase.table('rehearsals').select('fecha').eq('event_id', evento_id).execute().data or []
        for e in ens:
            d = (e.get('fecha') or '')[:10]
            if d:
                if not f_ini or d < f_ini: f_ini = d
                if not f_fin or d > f_fin: f_fin = d
    except Exception:
        pass
    # Préstamos activos de esa obra (tabla obra_prestamos)
    try:
        prest = supabase.table('obra_prestamos').select('*').eq('obra_id', obra_id) \
            .neq('estado', 'devuelto').execute().data or []
    except Exception:
        prest = []
    conflictos = []
    for p in prest:
        ps = (p.get('fecha_salida') or '')[:10]
        pe = (p.get('fecha_devolucion_real') or p.get('fecha_prevista_devolucion') or '')[:10]
        if not ps:
            continue
        # Solape: ps <= f_fin AND (pe == '' OR pe >= f_ini)
        if ps <= f_fin and (not pe or pe >= f_ini):
            conflictos.append(p)
    return {"conflictos": conflictos, "fechas_evento": {"desde": f_ini, "hasta": f_fin}}


# ============================================================
# Estado de material de una obra (utilidad para B5)
# ============================================================
@router.get("/obras/{obra_id}/estado-material")
async def estado_material_obra(obra_id: str, evento_id: Optional[str] = None,
                                current_user: dict = Depends(get_current_gestor)):
    """Calcula el estado de material de una obra: completo / incompleto / necesita_revision.
    Si se pasa evento_id, indica si las copias_fisicas son suficientes para los atriles del evento.
    """
    partes = supabase.table('obra_partes').select('*').eq('obra_id', obra_id).execute().data or []
    if not partes:
        return {"estado": "sin_partes", "copias_total": 0, "partes_count": 0,
                "copias_suficientes": None, "deficit_por_seccion": []}
    # Estados individuales: revisar campo `estado` de cada parte
    estados = [p.get('estado') for p in partes if p.get('estado')]
    necesita_revision = any(e == 'necesita_revision' for e in estados)
    incompleto = any(e == 'no' for e in estados) or any((p.get('copias_fisicas') or 0) == 0 for p in partes)
    if necesita_revision:
        estado_global = 'necesita_revision'
    elif incompleto:
        estado_global = 'incompleto'
    else:
        estado_global = 'completo'
    copias_total = sum(int(p.get('copias_fisicas') or 0) for p in partes)
    deficit = []
    if evento_id:
        # Atriles necesarios por sección — contar instrumentation del evento
        try:
            ev = supabase.table('eventos').select('instrumentation').eq('id', evento_id).limit(1).execute().data or []
            inst = (ev[0].get('instrumentation') if ev else {}) or {}
            for sec, cantidad_str in inst.items():
                # 'cuerda' o '6' etc — extraer entero
                try:
                    cantidad = int(str(cantidad_str).strip())
                except Exception:
                    continue
                copias_sec = sum(int(p.get('copias_fisicas') or 0) for p in partes
                                 if (p.get('seccion') or '').lower() == sec.lower())
                if copias_sec < cantidad:
                    deficit.append({"seccion": sec, "copias": copias_sec, "necesarias": cantidad,
                                    "deficit": cantidad - copias_sec})
        except Exception:
            pass
    return {"estado": estado_global, "copias_total": copias_total, "partes_count": len(partes),
            "copias_suficientes": len(deficit) == 0 if evento_id else None,
            "deficit_por_seccion": deficit}
